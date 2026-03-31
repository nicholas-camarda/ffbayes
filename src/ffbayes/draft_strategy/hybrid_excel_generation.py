#!/usr/bin/env python3
"""
Hybrid Excel File Generation

This module creates Excel files that follow the VOR structure but include
additional uncertainty data from Bayesian analysis, plus a Statistics Guide.
"""

import logging
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Configure logging
logger = logging.getLogger(__name__)


def create_early_draft_sheet(workbook: openpyxl.Workbook, risk_adjusted_data: pd.DataFrame) -> None:
    """
    Create Early Draft sheet (Rounds 1-5) with VOR + uncertainty data.
    
    Args:
        workbook: OpenPyXL workbook to add sheet to
        risk_adjusted_data: DataFrame with risk-adjusted rankings
    """
    logger.info("Creating Early Draft sheet (Rounds 1-5)...")
    
    # Create sheet
    ws = workbook.create_sheet('Early Draft (Rounds 1-5)')
    
    # Filter for top players (roughly rounds 1-5 in 10-team league = top 50)
    early_data = risk_adjusted_data.head(50).copy()
    
    # Select columns for display
    display_columns = [
        'PLAYER', 'POS', 'FPTS', 'VOR', 'VALUERANK', 
        'uncertainty_score', 'composite_score', 'composite_rank',
        'risk_category', 'mean_projection', 'std_projection'
    ]
    
    # Filter to available columns
    available_columns = [col for col in display_columns if col in early_data.columns]
    sheet_data = early_data[available_columns].copy()
    
    # Add data to sheet
    for r in dataframe_to_rows(sheet_data, index=False, header=True):
        ws.append(r)
    
    # Apply formatting
    apply_vor_formatting(ws)
    
    # Add conditional formatting for risk categories
    apply_risk_color_formatting(ws, 'risk_category')
    
    logger.info(f"✅ Early Draft sheet created with {len(sheet_data)} players")


def create_mid_draft_sheet(workbook: openpyxl.Workbook, risk_adjusted_data: pd.DataFrame) -> None:
    """
    Create Mid Draft sheet (Rounds 6-10) with VOR + uncertainty data.
    
    Args:
        workbook: OpenPyXL workbook to add sheet to
        risk_adjusted_data: DataFrame with risk-adjusted rankings
    """
    logger.info("Creating Mid Draft sheet (Rounds 6-10)...")
    
    # Create sheet
    ws = workbook.create_sheet('Mid Draft (Rounds 6-10)')
    
    # Filter for mid-round players (roughly rounds 6-10 in 10-team league = players 51-100)
    if len(risk_adjusted_data) >= 100:
        mid_data = risk_adjusted_data.iloc[50:100].copy()
    else:
        mid_data = risk_adjusted_data.iloc[50:].copy()
    
    # Select columns for display
    display_columns = [
        'PLAYER', 'POS', 'FPTS', 'VOR', 'VALUERANK', 
        'uncertainty_score', 'composite_score', 'composite_rank',
        'risk_category', 'mean_projection', 'std_projection'
    ]
    
    # Filter to available columns
    available_columns = [col for col in display_columns if col in mid_data.columns]
    sheet_data = mid_data[available_columns].copy()
    
    # Add data to sheet
    for r in dataframe_to_rows(sheet_data, index=False, header=True):
        ws.append(r)
    
    # Apply formatting
    apply_vor_formatting(ws)
    
    # Add conditional formatting for risk categories
    apply_risk_color_formatting(ws, 'risk_category')
    
    logger.info(f"✅ Mid Draft sheet created with {len(sheet_data)} players")


def create_late_draft_sheet(workbook: openpyxl.Workbook, risk_adjusted_data: pd.DataFrame) -> None:
    """
    Create Late Draft sheet (Rounds 11-16) with VOR + uncertainty data.
    
    Args:
        workbook: OpenPyXL workbook to add sheet to
        risk_adjusted_data: DataFrame with risk-adjusted rankings
    """
    logger.info("Creating Late Draft sheet (Rounds 11-16)...")
    
    # Create sheet
    ws = workbook.create_sheet('Late Draft (Rounds 11-16)')
    
    # Filter for late-round players (roughly rounds 11-16 in 10-team league = players 101+)
    if len(risk_adjusted_data) >= 100:
        late_data = risk_adjusted_data.iloc[100:].copy()
    elif len(risk_adjusted_data) >= 50:
        # If we have at least 50 players, put some in late draft
        late_data = risk_adjusted_data.iloc[50:].copy()
    else:
        late_data = pd.DataFrame()  # Empty if not enough data
    
    if not late_data.empty:
        # Select columns for display
        display_columns = [
            'PLAYER', 'POS', 'FPTS', 'VOR', 'VALUERANK', 
            'uncertainty_score', 'composite_score', 'composite_rank',
            'risk_category', 'mean_projection', 'std_projection'
        ]
        
        # Filter to available columns
        available_columns = [col for col in display_columns if col in late_data.columns]
        sheet_data = late_data[available_columns].copy()
        
        # Add data to sheet
        for r in dataframe_to_rows(sheet_data, index=False, header=True):
            ws.append(r)
        
        # Apply formatting
        apply_vor_formatting(ws)
        
        # Add conditional formatting for risk categories
        apply_risk_color_formatting(ws, 'risk_category')
        
        logger.info(f"✅ Late Draft sheet created with {len(sheet_data)} players")
    else:
        # Add header only
        headers = ['PLAYER', 'POS', 'FPTS', 'VOR', 'VALUERANK', 
                  'uncertainty_score', 'composite_score', 'composite_rank',
                  'risk_category', 'mean_projection', 'std_projection']
        ws.append(headers)
        apply_vor_formatting(ws)
        logger.info("✅ Late Draft sheet created (no late-round data available)")


def create_round_by_round_strategy_sheet(workbook: openpyxl.Workbook, risk_adjusted_data: pd.DataFrame) -> None:
    """
    Create Round-by-Round Strategy sheet with enhanced uncertainty recommendations.
    
    Args:
        workbook: OpenPyXL workbook to add sheet to
        risk_adjusted_data: DataFrame with risk-adjusted rankings
    """
    logger.info("Creating Round-by-Round Strategy sheet...")
    
    # Create sheet
    ws = workbook.create_sheet('Round-by-Round Strategy')
    
    # Create round-by-round strategy data
    strategy_data = []
    
    # Get top 160 players (16 rounds * 10 teams)
    top_players = risk_adjusted_data.head(160).copy()
    
    for round_num in range(1, 17):
        # Calculate pick range for this round
        start_pick = (round_num - 1) * 10 + 1
        end_pick = round_num * 10
        
        # Get players in this pick range
        round_players = top_players.iloc[start_pick-1:end_pick].copy()
        
        for _, player in round_players.iterrows():
            strategy_data.append({
                'Round': round_num,
                'Pick Range': f"{start_pick}-{end_pick}",
                'PLAYER': player['PLAYER'],
                'POS': player['POS'],
                'VOR': player['VOR'],
                'composite_score': player['composite_score'],
                'uncertainty_score': player['uncertainty_score'],
                'risk_category': player['risk_category'],
                'Strategy': get_strategy_recommendation(player)
            })
    
    # Convert to DataFrame
    strategy_df = pd.DataFrame(strategy_data)
    
    # Add data to sheet
    for r in dataframe_to_rows(strategy_df, index=False, header=True):
        ws.append(r)
    
    # Apply formatting
    apply_vor_formatting(ws)
    
    # Add conditional formatting for risk categories
    apply_risk_color_formatting(ws, 'risk_category')
    
    logger.info(f"✅ Round-by-Round Strategy sheet created with {len(strategy_data)} recommendations")


def get_strategy_recommendation(player: pd.Series) -> str:
    """
    Generate strategy recommendation based on player data.
    
    Args:
        player: Player data series
        
    Returns:
        Strategy recommendation string
    """
    uncertainty = player['uncertainty_score']
    risk_category = player['risk_category']
    composite_score = player['composite_score']
    
    if risk_category == 'Low Risk':
        if composite_score > 50:
            return "Safe Stud - Excellent value with low risk"
        else:
            return "Safe Pick - Reliable production, good value"
    elif risk_category == 'Medium Risk':
        if composite_score > 50:
            return "Value Pick - Good upside with moderate risk"
        else:
            return "Balanced Pick - Decent value, moderate risk"
    else:  # High Risk
        if composite_score > 50:
            return "High Upside - Great value but high risk"
        else:
            return "Lottery Ticket - High risk, potential upside"


def create_strategy_summary_sheet(workbook: openpyxl.Workbook, risk_adjusted_data: pd.DataFrame) -> None:
    """
    Create Strategy Summary sheet with risk-adjusted overview.
    
    Args:
        workbook: OpenPyXL workbook to add sheet to
        risk_adjusted_data: DataFrame with risk-adjusted rankings
    """
    logger.info("Creating Strategy Summary sheet...")
    
    # Create sheet
    ws = workbook.create_sheet('Strategy Summary')
    
    # Create summary data
    summary_data = []
    
    # Position breakdown
    for pos in risk_adjusted_data['POS'].unique():
        pos_data = risk_adjusted_data[risk_adjusted_data['POS'] == pos]
        
        summary_data.append({
            'Category': 'Position Breakdown',
            'Position': pos,
            'Player Count': len(pos_data),
            'Avg VOR': pos_data['VOR'].mean(),
            'Avg Uncertainty': pos_data['uncertainty_score'].mean(),
            'Avg Composite Score': pos_data['composite_score'].mean(),
            'Top Player': pos_data.iloc[0]['PLAYER'] if not pos_data.empty else 'N/A'
        })
    
    # Risk breakdown
    for risk in ['Low Risk', 'Medium Risk', 'High Risk']:
        risk_data = risk_adjusted_data[risk_adjusted_data['risk_category'] == risk]
        
        if not risk_data.empty:
            summary_data.append({
                'Category': 'Risk Breakdown',
                'Position': risk,
                'Player Count': len(risk_data),
                'Avg VOR': risk_data['VOR'].mean(),
                'Avg Uncertainty': risk_data['uncertainty_score'].mean(),
                'Avg Composite Score': risk_data['composite_score'].mean(),
                'Top Player': risk_data.iloc[0]['PLAYER']
            })
    
    # Top players by composite score
    top_composite = risk_adjusted_data.head(10)
    for i, (_, player) in enumerate(top_composite.iterrows(), 1):
        summary_data.append({
            'Category': 'Top Composite Players',
            'Position': f"#{i}",
            'Player Count': player['PLAYER'],
            'Avg VOR': player['VOR'],
            'Avg Uncertainty': player['uncertainty_score'],
            'Avg Composite Score': player['composite_score'],
            'Top Player': player['risk_category']
        })
    
    # Convert to DataFrame
    summary_df = pd.DataFrame(summary_data)
    
    # Add data to sheet
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws.append(r)
    
    # Apply formatting
    apply_vor_formatting(ws)
    
    logger.info(f"✅ Strategy Summary sheet created with {len(summary_data)} summary rows")


def create_statistics_guide_sheet(workbook: openpyxl.Workbook) -> None:
    """
    Create Statistics Guide sheet explaining all new hybrid statistics.
    
    Args:
        workbook: OpenPyXL workbook to add sheet to
    """
    logger.info("Creating Statistics Guide sheet...")
    
    # Create sheet
    ws = workbook.create_sheet('Statistics Guide')
    
    # Define guide content
    guide_content = [
        ['Statistic', 'What It Means', 'How to Use', 'Draft Strategy'],
        ['', '', '', ''],
        ['Uncertainty Score (0.0-1.0)', 'How uncertain we are about this player\'s projection', 'Low (0.0-0.3): Safe picks\nMedium (0.3-0.6): Balanced\nHigh (0.6-1.0): High risk', 'Lower uncertainty = safer picks\nHigher uncertainty = potential breakout'],
        ['', '', '', ''],
        ['Mean Projection', 'Average projected fantasy points from Monte Carlo simulations', 'Compare with FPTS to see if Bayesian analysis agrees with VOR projections', 'Higher mean projection = more confidence in upside'],
        ['', '', '', ''],
        ['Standard Deviation (Std Projection)', 'How much the player\'s projections vary across simulations', 'Low std: Consistent performance\nHigh std: Volatile, boom-or-bust potential', 'Lower std = safer floor\nHigher std = higher ceiling but more risk'],
        ['', '', '', ''],
        ['Composite Score', 'Risk-adjusted ranking that combines VOR value with uncertainty', 'This is the primary ranking to use for draft decisions', 'Higher composite score = better risk-adjusted value'],
        ['', '', '', ''],
        ['Risk Categories', 'Color-coded risk levels based on uncertainty scores', 'Green: Low Risk (0.0-0.3)\nYellow: Medium Risk (0.3-0.6)\nRed: High Risk (0.6-1.0)', 'Balance safe picks with upside players'],
        ['', '', '', ''],
        ['Early Rounds (1-5)', 'Focus on players with low to medium uncertainty', 'Use composite score as primary ranking\nAvoid high-uncertainty players unless exceptional upside', 'Build solid foundation with reliable production'],
        ['', '', '', ''],
        ['Middle Rounds (6-10)', 'Balance VOR value with uncertainty', 'Look for medium uncertainty with good composite scores\nConsider mean projection vs FPTS for upside', 'Build depth while maintaining value'],
        ['', '', '', ''],
        ['Late Rounds (11-16)', 'Target high-uncertainty players with upside', 'Use standard deviation to identify boom-or-bust candidates\nFocus on mean projection for sleepers', 'Take calculated risks for potential breakout players']
    ]
    
    # Add content to sheet
    for row in guide_content:
        ws.append(row)
    
    # Apply formatting
    apply_guide_formatting(ws)
    
    logger.info("✅ Statistics Guide sheet created")


def apply_vor_formatting(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """
    Apply VOR-style formatting to worksheet.
    
    Args:
        worksheet: OpenPyXL worksheet to format
    """
    # Header formatting
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    
    # Apply header formatting
    for cell in worksheet[1]:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data formatting
    data_font = Font(size=11)
    
    # Apply data formatting
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def apply_guide_formatting(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """
    Apply formatting specific to the Statistics Guide sheet.
    
    Args:
        worksheet: OpenPyXL worksheet to format
    """
    # Header formatting
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Section header formatting
    section_font = Font(bold=True, size=12, color="FFFFFF")
    section_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    # Apply formatting
    for row_num, row in enumerate(worksheet.iter_rows(), 1):
        for cell in row:
            if cell.value:
                if row_num == 1:  # Main header
                    cell.font = header_font
                    cell.fill = header_fill
                elif 'Rounds' in str(cell.value) or cell.value in ['Early Rounds (1-5)', 'Middle Rounds (6-10)', 'Late Rounds (11-16)']:
                    cell.font = section_font
                    cell.fill = section_fill
                else:
                    cell.font = Font(size=11)
                
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        
        adjusted_width = min(max_length + 2, 60)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def apply_risk_color_formatting(worksheet: openpyxl.worksheet.worksheet.Worksheet, risk_column: str) -> None:
    """
    Apply color formatting based on risk categories.
    
    Args:
        worksheet: OpenPyXL worksheet to format
        risk_column: Column name containing risk categories
    """
    # Find risk column index
    risk_col_idx = None
    for col_idx, cell in enumerate(worksheet[1], 1):
        if cell.value == risk_column:
            risk_col_idx = col_idx
            break
    
    if risk_col_idx is None:
        return
    
    # Define colors for risk categories
    risk_colors = {
        'Low Risk': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        'Medium Risk': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        'High Risk': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    }
    
    # Apply colors to risk column
    for row_num in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_num, column=risk_col_idx)
        if cell.value in risk_colors:
            cell.fill = risk_colors[cell.value]


def validate_excel_structure(excel_file_path: str) -> None:
    """
    Validate that Excel file has the required structure.
    
    Args:
        excel_file_path: Path to Excel file to validate
        
    Raises:
        ValueError: If validation fails
    """
    logger.info(f"Validating Excel structure: {excel_file_path}")
    
    if not Path(excel_file_path).exists():
        raise ValueError(f"Excel file not found: {excel_file_path}")
    
    try:
        wb = openpyxl.load_workbook(excel_file_path)
        
        # Check for required sheets
        required_sheets = [
            'Early Draft (Rounds 1-5)',
            'Mid Draft (Rounds 6-10)',
            'Late Draft (Rounds 11-16)',
            'Round-by-Round Strategy',
            'Strategy Summary',
            'Statistics Guide'
        ]
        
        # Add Pick-by-Pick Strategy sheet if it exists (optional)
        if 'Pick-by-Pick Strategy' in wb.sheetnames:
            required_sheets.append('Pick-by-Pick Strategy')
        
        missing_sheets = [sheet for sheet in required_sheets if sheet not in wb.sheetnames]
        if missing_sheets:
            raise ValueError(f"Missing required sheets: {missing_sheets}")
        
        # Check that sheets have data
        for sheet_name in required_sheets[:-1]:  # Exclude Statistics Guide
            sheet = wb[sheet_name]
            if sheet.max_row <= 1:
                logger.warning(f"Sheet '{sheet_name}' has no data")
        
        logger.info("✅ Excel structure validation passed")
        
    except Exception as e:
        raise ValueError(f"Excel validation failed: {e}")


def create_hybrid_excel_file(risk_adjusted_data: pd.DataFrame, output_file_path: str, recommendations: Optional[pd.DataFrame] = None) -> None:
    """
    Create complete hybrid Excel file following VOR structure with uncertainty data.
    
    Args:
        risk_adjusted_data: DataFrame with risk-adjusted rankings
        output_file_path: Path to save Excel file
        recommendations: Optional DataFrame with pick-by-pick recommendations
    """
    logger.info(f"Creating hybrid Excel file: {output_file_path}")
    
    try:
        # Create new workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create all sheets
        create_early_draft_sheet(wb, risk_adjusted_data)
        create_mid_draft_sheet(wb, risk_adjusted_data)
        create_late_draft_sheet(wb, risk_adjusted_data)
        create_round_by_round_strategy_sheet(wb, risk_adjusted_data)
        create_strategy_summary_sheet(wb, risk_adjusted_data)
        create_statistics_guide_sheet(wb)
        
        # Add pick-by-pick recommendations if provided
        if recommendations is not None:
            from ffbayes.draft_strategy.pick_by_pick_strategy import (
                create_pick_by_pick_sheet,
            )
            create_pick_by_pick_sheet(wb, risk_adjusted_data)
        
        # Save workbook
        wb.save(output_file_path)
        
        # Validate structure
        validate_excel_structure(output_file_path)
        
        logger.info("🎉 Hybrid Excel file created successfully!")
        
    except Exception as e:
        logger.error(f"❌ Failed to create hybrid Excel file: {e}")
        raise


def main():
    """Main function for testing hybrid Excel generation."""
    try:
        # Load risk-adjusted data
        from ffbayes.data_pipeline.hybrid_data_integration import (
            create_hybrid_dataset,
            load_hybrid_data_sources,
        )
        from ffbayes.draft_strategy.risk_adjusted_rankings import (
            create_risk_adjusted_rankings,
        )

        # Load and create risk-adjusted data
        vor_data, bayesian_data = load_hybrid_data_sources()
        hybrid_data = create_hybrid_dataset(vor_data, bayesian_data)
        risk_adjusted_data = create_risk_adjusted_rankings(hybrid_data)
        
        # Create Excel file
        output_path = "test_hybrid_draft_strategy.xlsx"
        create_hybrid_excel_file(risk_adjusted_data, output_path)
        
        print("🎉 Hybrid Excel generation successful!")
        print(f"📊 Excel file created: {output_path}")
        print(f"📈 Risk-adjusted dataset: {len(risk_adjusted_data)} players")
        
        return output_path
        
    except Exception as e:
        logger.error(f"❌ Hybrid Excel generation failed: {e}")
        raise


if __name__ == "__main__":
    main()
