#!/usr/bin/env python3
"""
Draft decision system for FFBayes.

This module replaces the old "rank players and call it Bayesian" flow with a
draft-facing decision stack:

* normalize player data from multiple public/free sources
* build a player-level decision table
* score live snake-draft recommendations conditional on roster state
* simulate strategy backtests on historical seasons
* export a portable workbook and dashboard payload for draft day

The implementation is intentionally conservative and data-first. When inputs
are missing, the code falls back to transparent heuristics instead of failing
open with invented values.
"""

from __future__ import annotations

import json
import math
import shutil
from dataclasses import asdict, dataclass, field, replace
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.stats import spearmanr

from ffbayes.analysis.bayesian_player_model import (
    aggregate_season_player_table,
    build_posterior_projection_table,
)

POSITION_ORDER = ['QB', 'RB', 'WR', 'TE', 'DST', 'K']
FANTASY_DRAFT_POSITIONS = tuple(POSITION_ORDER)
OFFENSIVE_POSITIONS = ('QB', 'RB', 'WR', 'TE')
DEFAULT_FLEX_WEIGHTS = {'RB': 0.45, 'WR': 0.45, 'TE': 0.10}
DEFAULT_ROSTER_TEMPLATE = {
    'QB': 1,
    'RB': 2,
    'WR': 2,
    'TE': 1,
    'FLEX': 1,
    'DST': 1,
    'K': 1,
}
SCORING_PRESETS = {
    'standard': {'label': 'Standard', 'scoring_type': 'Standard', 'ppr_value': 0.0},
    'half_ppr': {'label': 'Half PPR', 'scoring_type': 'Half-PPR', 'ppr_value': 0.5},
    'ppr': {'label': 'Full PPR', 'scoring_type': 'PPR', 'ppr_value': 1.0},
}
LATE_SPECIALIST_BUFFER_ROUNDS = 2
CONSERVATIVE_WAIT_SURVIVAL_THRESHOLD = 0.74
CONSERVATIVE_WAIT_LINEUP_LOSS_THRESHOLD = 0.28
SECONDARY_QB_TE_VALUE_EDGE = 0.45
METRIC_GLOSSARY = {
    'draft_score': {
        'label': 'Board value score',
        'summary': 'Base player value before draft-slot action rules are applied.',
        'detail': 'Combines posterior projection, starter edge, replacement edge, and a light market-gap term into a cleaner board ranking.',
    },
    'board_value_score': {
        'label': 'Board value score',
        'summary': 'Base player value before draft-slot action rules are applied.',
        'detail': 'Used as the policy foundation. The live recommendation then layers roster urgency and wait risk on top of this score.',
    },
    'availability_to_next_pick': {
        'label': 'Availability to next pick',
        'summary': 'Estimated chance the player survives until your next turn.',
        'detail': 'Uses fantasy-only ADP dispersion within the player pool plus uncertainty. Higher means waiting is safer.',
    },
    'expected_regret': {
        'label': 'Expected regret',
        'summary': 'Penalty for waiting instead of taking the player now.',
        'detail': 'Rises when the player is a top value and unlikely to make it back. Lower is better if you want to wait.',
    },
    'fragility_score': {
        'label': 'Fragility score',
        'summary': 'Risk score for uncertainty, injury, volatility, and thin history.',
        'detail': 'Higher values mean the profile is shakier. In the contextual model this acts as a penalty, especially in low-risk mode.',
    },
    'upside_score': {
        'label': 'Upside score',
        'summary': 'How much ceiling and breakout potential the player offers.',
        'detail': 'Uses ceiling-over-mean, raw projection, and market timing. Higher is more swing-for-the-fences.',
    },
    'starter_delta': {
        'label': 'Starter delta',
        'summary': 'Projected edge over a typical starter at the same position.',
        'detail': 'Positive values mean the player is clearly starter-worthy in your league shape.',
    },
    'wait_signal': {
        'label': 'Wait signal',
        'summary': 'Plain-English explanation for whether waiting is acceptable.',
        'detail': 'Shows whether the player is safe to wait on, too costly to pass, too unlikely to survive, or only acceptable as a late specialist stash.',
    },
    'market_gap': {
        'label': 'Market gap',
        'summary': 'Difference between where the model and the market rank the player.',
        'detail': 'Positive values mean the model likes the player more than current market cost.',
    },
    'replacement_delta': {
        'label': 'Simple VOR proxy',
        'summary': 'Projected edge over replacement level at the position.',
        'detail': 'This is the dashboard’s baseline VOR-style comparison point when we contrast the contextual score against a simpler value-over-replacement view.',
    },
    'posterior_prob_beats_replacement': {
        'label': 'Posterior beat-replacement probability',
        'summary': 'Chance the model thinks this player beats replacement-level value.',
        'detail': 'Derived from the posterior mean and uncertainty width. Higher values mean the player clears the replacement bar more reliably.',
    },
}


def _safe_string(value: Any) -> str:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return ''
    return str(value)


def _coerce_float(value: Any, default: float = np.nan) -> float:
    try:
        if value is None:
            return default
        if isinstance(value, str) and not value.strip():
            return default
        return float(value)
    except Exception:
        return default


def _coerce_int(value: Any, default: int = 0) -> int:
    try:
        if value is None:
            return default
        if isinstance(value, str) and not value.strip():
            return default
        return int(float(value))
    except Exception:
        return default


def _zscore(series: pd.Series) -> pd.Series:
    values = pd.to_numeric(series, errors='coerce')
    mean = values.mean(skipna=True)
    std = values.std(ddof=0, skipna=True)
    if pd.isna(std) or np.isclose(std, 0.0):
        return pd.Series(np.zeros(len(values)), index=series.index)
    return (values - mean) / std


def _clamp(series: pd.Series, low: float = 0.0, high: float = 1.0) -> pd.Series:
    return series.clip(lower=low, upper=high)


def _normalize_position(value: Any) -> str:
    pos = _safe_string(value).upper().strip()
    if not pos:
        return 'UNKNOWN'
    if pos.startswith('RB'):
        return 'RB'
    if pos.startswith('WR'):
        return 'WR'
    if pos.startswith('QB'):
        return 'QB'
    if pos.startswith('TE'):
        return 'TE'
    if pos.startswith('DST') or pos in {'D/ST', 'DEF', 'DEFENSE'}:
        return 'DST'
    if pos.startswith('K'):
        return 'K'
    return 'UNKNOWN'


def _pick_first_row(series: pd.Series, fallback: Any = None) -> Any:
    for value in series:
        if pd.notna(value) and _safe_string(value):
            return value
    return fallback


@dataclass(frozen=True)
class LeagueSettings:
    """League-level configuration used to convert projections into decisions."""

    league_size: int = 10
    draft_position: int = 10
    scoring_type: str = 'PPR'
    ppr_value: float = 0.5
    risk_tolerance: str = 'medium'
    roster_spots: dict[str, int] = field(
        default_factory=lambda: dict(DEFAULT_ROSTER_TEMPLATE)
    )
    flex_weights: dict[str, float] = field(
        default_factory=lambda: dict(DEFAULT_FLEX_WEIGHTS)
    )
    bench_slots: int = 6

    @classmethod
    def from_mapping(cls, mapping: dict[str, Any] | None = None) -> 'LeagueSettings':
        mapping = mapping or {}
        league_settings = mapping.get('league_settings', mapping)
        roster = dict(DEFAULT_ROSTER_TEMPLATE)
        roster.update(
            {
                k.upper(): _coerce_int(v, roster.get(k.upper(), 0))
                for k, v in league_settings.get('roster_spots', {}).items()
            }
        )
        flex_weights = dict(DEFAULT_FLEX_WEIGHTS)
        flex_weights.update(
            {
                k.upper(): float(v)
                for k, v in league_settings.get('flex_weights', {}).items()
            }
        )
        return cls(
            league_size=_coerce_int(league_settings.get('league_size', 10), 10),
            draft_position=_coerce_int(league_settings.get('draft_position', 10), 10),
            scoring_type=_safe_string(league_settings.get('scoring_type', 'PPR'))
            or 'PPR',
            ppr_value=_coerce_float(
                league_settings.get('ppr_value', league_settings.get('ppr', 0.5)), 0.5
            ),
            risk_tolerance=_safe_string(league_settings.get('risk_tolerance', 'medium'))
            or 'medium',
            roster_spots=roster,
            flex_weights=flex_weights,
            bench_slots=_coerce_int(league_settings.get('bench_slots', 6), 6),
        )

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)

    def starters_by_position(self) -> dict[str, int]:
        return {
            pos: self.roster_spots.get(pos, 0) * self.league_size
            for pos in POSITION_ORDER
        }

    def effective_replacement_slots(self) -> dict[str, int]:
        starters = self.starters_by_position()
        flex_slots = self.roster_spots.get('FLEX', 0) * self.league_size
        flex_allocation = {
            'RB': int(round(flex_slots * self.flex_weights.get('RB', 0.0))),
            'WR': int(round(flex_slots * self.flex_weights.get('WR', 0.0))),
            'TE': int(round(flex_slots * self.flex_weights.get('TE', 0.0))),
        }
        replacement = dict(starters)
        replacement['RB'] = replacement.get('RB', 0) + flex_allocation['RB']
        replacement['WR'] = replacement.get('WR', 0) + flex_allocation['WR']
        replacement['TE'] = replacement.get('TE', 0) + flex_allocation['TE']
        return replacement

    def round_count(self) -> int:
        base = sum(
            self.roster_spots.get(pos, 0)
            for pos in ['QB', 'RB', 'WR', 'TE', 'DST', 'K']
        )
        base += self.bench_slots
        return max(1, base)


@dataclass
class DraftContext:
    """State of the live draft board when the recommendation is generated."""

    current_pick_number: int
    drafted_players: set[str] = field(default_factory=set)
    keepers: set[str] = field(default_factory=set)
    your_players: set[str] = field(default_factory=set)
    roster_counts: dict[str, int] = field(default_factory=dict)
    action_history: list[dict[str, Any]] = field(default_factory=list)
    selected_player: str | None = None

    def drafted_set(self) -> set[str]:
        drafted = {
            name.strip().lower()
            for name in self.drafted_players
            if isinstance(name, str) and name.strip()
        }
        drafted.update(
            {
                name.strip().lower()
                for name in self.keepers
                if isinstance(name, str) and name.strip()
            }
        )
        drafted.update(
            {
                name.strip().lower()
                for name in self.your_players
                if isinstance(name, str) and name.strip()
            }
        )
        return drafted


@dataclass(frozen=True)
class LiveDraftState:
    """Client-side draft state tracked by the dashboard."""

    current_pick_number: int
    next_pick_number: int
    taken_players: list[str] = field(default_factory=list)
    your_players: list[str] = field(default_factory=list)
    roster_counts: dict[str, int] = field(default_factory=dict)
    action_history: list[dict[str, Any]] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class DraftDecisionArtifacts:
    """Container for all draft-facing artifacts produced by the system."""

    league_settings: LeagueSettings
    decision_table: pd.DataFrame
    recommendations: pd.DataFrame
    roster_scenarios: pd.DataFrame
    tier_cliffs: pd.DataFrame
    source_freshness: pd.DataFrame
    backtest: dict[str, Any]
    dashboard_payload: dict[str, Any]
    metadata: dict[str, Any]

    def to_dict(self) -> dict[str, Any]:
        return {
            'league_settings': self.league_settings.to_dict(),
            'decision_table': self.decision_table.to_dict(orient='records'),
            'recommendations': self.recommendations.to_dict(orient='records'),
            'roster_scenarios': self.roster_scenarios.to_dict(orient='records'),
            'tier_cliffs': self.tier_cliffs.to_dict(orient='records'),
            'source_freshness': self.source_freshness.to_dict(orient='records'),
            'backtest': self.backtest,
            'dashboard_payload': self.dashboard_payload,
            'metadata': self.metadata,
        }


def _effective_ppr_value(settings: LeagueSettings) -> float:
    explicit = _coerce_float(getattr(settings, 'ppr_value', np.nan), np.nan)
    if np.isfinite(explicit):
        return explicit
    scoring_type = _safe_string(getattr(settings, 'scoring_type', '')).lower()
    if 'standard' in scoring_type:
        return 0.0
    if 'half' in scoring_type:
        return 0.5
    if 'ppr' in scoring_type:
        return 1.0
    return 0.5


def _resolve_scoring_preset_key(settings: LeagueSettings) -> str:
    ppr_value = _effective_ppr_value(settings)
    if np.isclose(ppr_value, 0.0):
        return 'standard'
    if np.isclose(ppr_value, 0.5):
        return 'half_ppr'
    if np.isclose(ppr_value, 1.0):
        return 'ppr'
    return 'custom'


def _scoring_settings_for_preset(
    settings: LeagueSettings, preset_key: str
) -> LeagueSettings | None:
    preset = SCORING_PRESETS.get(preset_key)
    if preset is None:
        return None
    return replace(
        settings,
        scoring_type=preset['scoring_type'],
        ppr_value=float(preset['ppr_value']),
    )


def _first_numeric_series(
    frame: pd.DataFrame, candidates: Iterable[str]
) -> pd.Series | None:
    for column in candidates:
        if column in frame.columns:
            return pd.to_numeric(frame[column], errors='coerce')
    return None


def _supports_scoring_presets(frame: pd.DataFrame) -> tuple[bool, str]:
    standard = _first_numeric_series(
        frame,
        [
            'FantPt',
            'fantasy_points',
            'fantasy_points_standard',
            'proj_points_standard',
            'proj_points_mean',
        ],
    )
    ppr = _first_numeric_series(
        frame,
        ['FantPtPPR', 'fantasy_points_ppr', 'proj_points_ppr'],
    )
    rec = _first_numeric_series(frame, ['REC', 'receptions', 'Receptions'])
    if standard is not None and ppr is not None:
        return True, ''
    if standard is not None and rec is not None:
        return True, ''
    return (
        False,
        'Only one scoring projection source is present, so alternate scoring presets would be guesswork.',
    )


def _projection_series_for_settings(
    frame: pd.DataFrame, settings: LeagueSettings
) -> pd.Series | None:
    base_projection = pd.to_numeric(frame.get('proj_points_mean'), errors='coerce')
    ppr_value = _effective_ppr_value(settings)
    standard = _first_numeric_series(
        frame,
        [
            'FantPt',
            'fantasy_points',
            'fantasy_points_standard',
            'proj_points_standard',
            'proj_points_mean',
        ],
    )
    ppr = _first_numeric_series(
        frame,
        ['FantPtPPR', 'fantasy_points_ppr', 'proj_points_ppr'],
    )
    rec = _first_numeric_series(frame, ['REC', 'receptions', 'Receptions'])

    projection = None
    if standard is not None and ppr is not None:
        projection = standard + ppr_value * (ppr - standard)
    elif standard is not None and rec is not None:
        projection = standard + ppr_value * rec
    elif ppr is not None and np.isclose(ppr_value, 1.0):
        projection = ppr
    elif standard is not None and np.isclose(ppr_value, 0.0):
        projection = standard

    if projection is None:
        return base_projection
    if base_projection is None:
        return projection
    return projection.combine_first(base_projection)


def _dashboard_supporting_math(decision_table: pd.DataFrame) -> dict[str, float]:
    return {
        'draft_score_mean': float(decision_table['draft_score'].mean())
        if not decision_table.empty
        else 0.0,
        'draft_score_std': float(decision_table['draft_score'].std(ddof=0))
        if not decision_table.empty
        else 0.0,
        'availability_mean': float(decision_table['availability_at_pick'].mean())
        if not decision_table.empty
        else 0.0,
        'top_draft_score': float(decision_table['draft_score'].max())
        if not decision_table.empty
        else 0.0,
        'top_simple_vor_proxy': float(decision_table['replacement_delta'].max())
        if not decision_table.empty
        else 0.0,
    }


def _dashboard_position_summary(decision_table: pd.DataFrame) -> list[dict[str, Any]]:
    if decision_table.empty:
        return []
    return (
        decision_table.groupby('position')
        .agg(
            player_count=('player_name', 'count'),
            mean_draft_score=('draft_score', 'mean'),
            mean_availability=('availability_at_pick', 'mean'),
            mean_upside=('upside_score', 'mean'),
            mean_fragility=('fragility_score', 'mean'),
            mean_proj=('proj_points_mean', 'mean'),
            mean_vor_proxy=('replacement_delta', 'mean'),
        )
        .reset_index()
        .to_dict(orient='records')
    )


def _build_bayesian_vor_summary(
    decision_table: pd.DataFrame, backtest: dict[str, Any]
) -> dict[str, Any]:
    table = decision_table.copy()
    if not table.empty:
        table['simple_vor_rank'] = (
            table['replacement_delta']
            .rank(method='first', ascending=False)
            .astype(int)
        )
        table['rank_gap_vs_vor'] = table['simple_vor_rank'] - table['draft_rank']
        disagreements = (
            table.assign(abs_rank_gap=lambda frame: frame['rank_gap_vs_vor'].abs())
            .sort_values(['abs_rank_gap', 'draft_score'], ascending=[False, False])
            .head(12)[
                [
                    'player_name',
                    'position',
                    'draft_rank',
                    'simple_vor_rank',
                    'rank_gap_vs_vor',
                    'draft_score',
                    'replacement_delta',
                    'availability_at_pick',
                    'why_flags',
                ]
            ]
            .to_dict(orient='records')
        )
    else:
        disagreements = []

    overall = backtest.get('overall', {}).get('by_strategy', []) if backtest else []
    by_strategy = {
        row.get('strategy'): row for row in overall if isinstance(row, dict)
    }
    draft_score_row = by_strategy.get('draft_score')
    vor_row = by_strategy.get('historical_vor_proxy')
    bootstrap = (
        backtest.get('overall', {}).get('draft_score_vs_historical_vor_proxy', {})
        if backtest
        else {}
    )
    if draft_score_row and vor_row:
        delta = float(draft_score_row['mean_lineup_points']) - float(
            vor_row['mean_lineup_points']
        )
        season_rows = []
        for season in backtest.get('by_season', []):
            season_map = season.get('by_strategy', {})
            draft_score_points = season_map.get('draft_score', {}).get(
                'our_team_lineup_points'
            )
            vor_points = season_map.get('historical_vor_proxy', {}).get(
                'our_team_lineup_points'
            )
            if draft_score_points is None or vor_points is None:
                continue
            season_rows.append(
                {
                    'holdout_year': season.get('holdout_year'),
                    'draft_score_lineup_points': float(draft_score_points),
                    'historical_vor_proxy_lineup_points': float(vor_points),
                    'delta_lineup_points': float(draft_score_points) - float(vor_points),
                    'winner': (
                        'draft_score'
                        if float(draft_score_points) >= float(vor_points)
                        else 'historical_vor_proxy'
                    ),
                }
            )
        return {
            'available': True,
            'comparison_scope': 'draft_level_internal_holdout_backtest',
            'headline': (
                'Contextual draft score outperforms the simple VOR proxy in backtests.'
                if delta > 0
                else 'Simple VOR proxy matches or beats the contextual draft score in backtests.'
            ),
            'winner': 'draft_score' if delta > 0 else 'historical_vor_proxy',
            'delta_mean_lineup_points': delta,
            'draft_score_mean_lineup_points': float(
                draft_score_row['mean_lineup_points']
            ),
            'historical_vor_proxy_mean_lineup_points': float(
                vor_row['mean_lineup_points']
            ),
            'season_count': min(
                int(draft_score_row.get('season_count', 0)),
                int(vor_row.get('season_count', 0)),
            ),
            'holdout_years': backtest.get('holdout_years', []),
            'bootstrap': bootstrap,
            'by_season': season_rows,
            'top_disagreements': disagreements,
            'limitations': [
                'This comparison is based on internal historical holdout seasons, not a live external validation set.',
                'The dashboard keeps the contextual draft score as a decision policy layered on posterior projections, not a pure posterior rank list.',
            ],
        }

    return {
        'available': False,
        'comparison_scope': 'board_only',
        'headline': 'No direct contextual-vs-VOR backtest summary is available for this export.',
        'reason_unavailable': 'No normalized backtest comparison rows were available.',
        'top_disagreements': disagreements,
        'limitations': [
            'Without the backtest comparison artifact, the board can show only current ranking disagreements, not outcome-level evidence.',
        ],
    }


def _build_decision_evidence(
    decision_table: pd.DataFrame,
    backtest: dict[str, Any],
    analysis_provenance: dict[str, Any] | None,
    bayesian_vor_summary: dict[str, Any] | None = None,
) -> dict[str, Any]:
    bayesian_vor_summary = bayesian_vor_summary or _build_bayesian_vor_summary(
        decision_table, backtest
    )
    analysis_provenance = analysis_provenance or {}
    freshness_context = (
        analysis_provenance.get('backtest_inputs')
        or analysis_provenance.get('overall_freshness')
        or {
            'status': 'unknown',
            'override_used': False,
            'warnings': [],
            'available_sources': [],
        }
    )
    strategy_summary = backtest.get('overall', {}).get('by_strategy', []) if backtest else []
    evaluation_scope = backtest.get('evaluation_scope', {}) if backtest else {}
    limitations = list(
        dict.fromkeys(
            [
                *(bayesian_vor_summary.get('limitations') or []),
                'Treat internal holdout backtests as directional decision evidence, not external validation.',
            ]
        )
    )

    status = 'available'
    reason_unavailable = None
    if not bayesian_vor_summary.get('available') or not strategy_summary:
        status = 'unavailable'
        reason_unavailable = bayesian_vor_summary.get(
            'reason_unavailable',
            'No backtest comparison artifact was available for this draft board.',
        )
    elif freshness_context.get('status') == 'degraded':
        status = 'degraded'

    return {
        'available': status != 'unavailable',
        'status': status,
        'headline': bayesian_vor_summary.get(
            'headline', 'Decision evidence is unavailable for this board.'
        ),
        'reason_unavailable': reason_unavailable,
        'evaluation_scope': evaluation_scope,
        'freshness': freshness_context,
        'strategy_summary': strategy_summary,
        'season_rows': bayesian_vor_summary.get('by_season', []),
        'top_disagreements': bayesian_vor_summary.get('top_disagreements', []),
        'winner': bayesian_vor_summary.get('winner'),
        'delta_mean_lineup_points': bayesian_vor_summary.get('delta_mean_lineup_points'),
        'season_count': bayesian_vor_summary.get('season_count'),
        'holdout_years': bayesian_vor_summary.get('holdout_years', []),
        'limitations': limitations,
        'interpretation_guidance': [
            'Use this panel to judge whether the board has internal evidence, not to treat it as externally validated truth.',
            'Freshness state applies to the evidence source, so degraded freshness weakens confidence even when a winner is shown.',
        ],
    }


def _build_scoring_preset_bundle(
    player_frame: pd.DataFrame,
    settings: LeagueSettings,
    context: DraftContext,
) -> dict[str, Any]:
    normalized = normalize_player_frame(player_frame)
    supported, reason = _supports_scoring_presets(normalized)
    active_key = _resolve_scoring_preset_key(settings)
    bundle: dict[str, Any] = {}
    for preset_key, preset in SCORING_PRESETS.items():
        preset_settings = _scoring_settings_for_preset(settings, preset_key)
        if preset_settings is None:
            continue
        if not supported and preset_key != active_key:
            bundle[preset_key] = {
                'key': preset_key,
                'label': preset['label'],
                'available': False,
                'reason_unavailable': reason,
                'league_settings': preset_settings.to_dict(),
            }
            continue
        preset_table = build_decision_table(player_frame, preset_settings, context)
        bundle[preset_key] = {
            'key': preset_key,
            'label': preset['label'],
            'available': True,
            'league_settings': preset_settings.to_dict(),
            'decision_table': preset_table.to_dict(orient='records'),
            'supporting_math': _dashboard_supporting_math(preset_table),
        }
    if active_key == 'custom':
        bundle['custom'] = {
            'key': 'custom',
            'label': f'Custom ({_effective_ppr_value(settings):.2f} PPR)',
            'available': True,
            'league_settings': settings.to_dict(),
            'decision_table': build_decision_table(player_frame, settings, context).to_dict(
                orient='records'
            ),
        }
    return bundle


def normalize_player_frame(player_frame: pd.DataFrame) -> pd.DataFrame:
    """Normalize a messy player frame into a standard draft table schema."""
    if player_frame is None or player_frame.empty:
        raise ValueError('player_frame is empty')

    df = player_frame.copy()

    rename_map = {}
    canonical_columns = set(df.columns)
    for column in df.columns:
        normalized = column.strip().lower().replace(' ', '_')
        if normalized in {'player', 'player_name', 'name', 'playername'}:
            if 'player_name' in canonical_columns:
                continue
            rename_map[column] = 'player_name'
            canonical_columns.add('player_name')
        elif normalized in {'pos', 'position', 'slot'}:
            if 'position' in canonical_columns:
                continue
            rename_map[column] = 'position'
            canonical_columns.add('position')
        elif normalized in {
            'fpts',
            'fantpt',
            'fantasy_points',
            'projected_points',
            'projected_fpts',
            'proj_points',
            'projection',
            'posterior_mean',
        }:
            if 'proj_points_mean' in canonical_columns:
                continue
            rename_map[column] = 'proj_points_mean'
            canonical_columns.add('proj_points_mean')
        elif normalized in {'adp', 'avg', 'average_draft_position', 'market_rank'}:
            if 'adp' in canonical_columns:
                continue
            rename_map[column] = 'adp'
            canonical_columns.add('adp')
        elif normalized in {'mean_projection', 'mean_proj', 'consensus_projection'}:
            if 'proj_points_mean' in canonical_columns:
                continue
            rename_map[column] = 'proj_points_mean'
            canonical_columns.add('proj_points_mean')
        elif normalized in {'std_projection', 'projection_std', 'projection_spread'}:
            if 'std_projection' in canonical_columns:
                continue
            rename_map[column] = 'std_projection'
            canonical_columns.add('std_projection')
        elif normalized in {'posterior_std', 'predictive_std'}:
            if 'std_projection' in canonical_columns:
                continue
            rename_map[column] = 'std_projection'
            canonical_columns.add('std_projection')
        elif normalized in {'uncertainty_score', 'risk_score', 'volatility_score'}:
            if 'uncertainty_score' in canonical_columns:
                continue
            rename_map[column] = 'uncertainty_score'
            canonical_columns.add('uncertainty_score')
        elif normalized in {'posterior_floor', 'projection_floor'}:
            if 'proj_points_floor' in canonical_columns:
                continue
            rename_map[column] = 'proj_points_floor'
            canonical_columns.add('proj_points_floor')
        elif normalized in {'posterior_ceiling', 'projection_ceiling'}:
            if 'proj_points_ceiling' in canonical_columns:
                continue
            rename_map[column] = 'proj_points_ceiling'
            canonical_columns.add('proj_points_ceiling')
        elif normalized in {'posterior_prob_beats_replacement', 'beat_replacement_prob'}:
            if 'posterior_prob_beats_replacement' in canonical_columns:
                continue
            rename_map[column] = 'posterior_prob_beats_replacement'
            canonical_columns.add('posterior_prob_beats_replacement')
        elif normalized in {'vor', 'value_over_replacement'}:
            if 'vor_value' in canonical_columns:
                continue
            rename_map[column] = 'vor_value'
            canonical_columns.add('vor_value')
        elif normalized in {'valuerank', 'vor_rank', 'market_rank_numeric'}:
            if 'market_rank' in canonical_columns:
                continue
            rename_map[column] = 'market_rank'
            canonical_columns.add('market_rank')

    df = df.rename(columns=rename_map)
    df = df.loc[:, ~df.columns.duplicated()]

    if 'player_name' not in df.columns:
        raise ValueError('player_frame must include a player name column')
    if 'position' not in df.columns:
        raise ValueError('player_frame must include a position column')

    df['player_name'] = df['player_name'].map(_safe_string)
    df['position'] = df['position'].map(_normalize_position)

    if 'proj_points_mean' not in df.columns:
        source_columns = [
            'mean_projection',
            'FPTS',
            'FantPt',
            'projected_points',
            'projection',
        ]
        for column in source_columns:
            if column in player_frame.columns:
                df['proj_points_mean'] = pd.to_numeric(
                    player_frame[column], errors='coerce'
                )
                break
    df['proj_points_mean'] = pd.to_numeric(df.get('proj_points_mean'), errors='coerce')

    if 'adp' in df.columns:
        df['adp'] = pd.to_numeric(df['adp'], errors='coerce')
    else:
        df['adp'] = np.nan

    if 'std_projection' in df.columns:
        df['std_projection'] = pd.to_numeric(df['std_projection'], errors='coerce')
    else:
        df['std_projection'] = np.nan

    if 'uncertainty_score' in df.columns:
        df['uncertainty_score'] = pd.to_numeric(
            df['uncertainty_score'], errors='coerce'
        )
    else:
        df['uncertainty_score'] = np.nan

    if 'proj_points_floor' in df.columns:
        df['proj_points_floor'] = pd.to_numeric(df['proj_points_floor'], errors='coerce')
    else:
        df['proj_points_floor'] = np.nan

    if 'proj_points_ceiling' in df.columns:
        df['proj_points_ceiling'] = pd.to_numeric(
            df['proj_points_ceiling'], errors='coerce'
        )
    else:
        df['proj_points_ceiling'] = np.nan

    if 'posterior_prob_beats_replacement' in df.columns:
        df['posterior_prob_beats_replacement'] = pd.to_numeric(
            df['posterior_prob_beats_replacement'], errors='coerce'
        )
    else:
        df['posterior_prob_beats_replacement'] = np.nan

    if 'vor_value' in df.columns:
        df['vor_value'] = pd.to_numeric(df['vor_value'], errors='coerce')
    else:
        df['vor_value'] = np.nan

    if 'market_rank' in df.columns:
        df['market_rank'] = pd.to_numeric(df['market_rank'], errors='coerce')
    else:
        df['market_rank'] = np.nan

    if 'season_count' not in df.columns:
        df['season_count'] = np.nan
    if 'games_missed' not in df.columns:
        df['games_missed'] = np.nan
    if 'age' not in df.columns:
        df['age'] = np.nan
    if 'years_in_league' not in df.columns:
        df['years_in_league'] = np.nan
    if 'team_change' not in df.columns:
        df['team_change'] = np.nan
    if 'role_volatility' not in df.columns:
        df['role_volatility'] = np.nan
    if 'role_label' not in df.columns:
        df['role_label'] = ''
    if 'team_pass_rate' not in df.columns:
        df['team_pass_rate'] = np.nan
    if 'neutral_pace' not in df.columns:
        df['neutral_pace'] = np.nan
    if 'vacated_targets' not in df.columns:
        df['vacated_targets'] = np.nan
    if 'teammate_competition' not in df.columns:
        df['teammate_competition'] = np.nan
    if 'site_disagreement' not in df.columns:
        df['site_disagreement'] = np.nan
    if 'adp_std' not in df.columns:
        df['adp_std'] = np.nan
    if 'source_name' not in df.columns:
        df['source_name'] = 'current_market_snapshot'
    if 'source_updated_at' not in df.columns:
        df['source_updated_at'] = pd.NaT

    df = df.drop_duplicates(subset=['player_name', 'position'], keep='last').reset_index(drop=True)
    return df


def _position_baseline(
    frame: pd.DataFrame, position: str, slot_count: int, fallback: float | None = None
) -> float:
    pos_values = pd.to_numeric(
        frame.loc[frame['position'] == position, 'proj_points_mean'], errors='coerce'
    ).dropna()
    if pos_values.empty:
        if fallback is not None and not pd.isna(fallback):
            return float(fallback)
        overall = pd.to_numeric(frame['proj_points_mean'], errors='coerce').dropna()
        return float(overall.mean()) if not overall.empty else 0.0

    slot_count = max(1, min(len(pos_values), int(round(slot_count))))
    sorted_values = pos_values.sort_values(ascending=False).reset_index(drop=True)
    return float(sorted_values.iloc[slot_count - 1])


def _compute_freshness(source_frame: pd.DataFrame) -> pd.DataFrame:
    if source_frame is None or source_frame.empty:
        return pd.DataFrame(
            columns=['source_name', 'source_updated_at', 'freshness_days', 'row_count']
        )

    summary_rows = []
    for source_name, sub in source_frame.groupby('source_name', dropna=False):
        updated = pd.to_datetime(sub['source_updated_at'], errors='coerce')
        freshest = updated.max() if not updated.isna().all() else pd.NaT
        freshness_days = (
            (pd.Timestamp.now(tz=None) - freshest).days
            if pd.notna(freshest)
            else np.nan
        )
        summary_rows.append(
            {
                'source_name': source_name if source_name else 'unknown',
                'source_updated_at': freshest,
                'freshness_days': freshness_days,
                'row_count': int(len(sub)),
                'missing_rate': float(sub.isna().mean().mean()),
            }
        )
    return pd.DataFrame(summary_rows)


def availability_probability(
    adp: float | int | None,
    target_pick: int,
    adp_std: float | int | None = None,
    uncertainty_score: float | int | None = None,
) -> float:
    """Estimate the probability a player survives to a future pick."""
    if adp is None or pd.isna(adp):
        return 0.5

    spread = _coerce_float(adp_std, np.nan)
    if pd.isna(spread) or spread <= 0:
        spread = 2.5

    if uncertainty_score is not None and not pd.isna(uncertainty_score):
        spread += 2.0 * float(uncertainty_score)

    z = (float(adp) - float(target_pick)) / max(1.0, spread)
    prob = 1.0 / (1.0 + math.exp(-z))
    return float(np.clip(prob, 0.0, 1.0))


def _filter_fantasy_player_pool(
    frame: pd.DataFrame,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    if frame is None:
        return (
            pd.DataFrame(),
            {
                'raw_player_count': 0,
                'eligible_player_count': 0,
                'removed_blank_name_count': 0,
                'removed_unknown_name_count': 0,
                'removed_non_fantasy_position_count': 0,
                'eligible_positions': list(FANTASY_DRAFT_POSITIONS),
            },
        )

    filtered = frame.copy()
    raw_player_count = int(len(filtered))
    names = filtered['player_name'].map(_safe_string).str.strip()
    positions = filtered['position'].map(_normalize_position)

    blank_name_mask = names.eq('')
    unknown_name_mask = names.str.upper().eq('UNKNOWN')
    invalid_position_mask = ~positions.isin(FANTASY_DRAFT_POSITIONS)
    keep_mask = ~(blank_name_mask | unknown_name_mask | invalid_position_mask)

    filtered = filtered.loc[keep_mask].copy()
    filtered['player_name'] = names.loc[keep_mask]
    filtered['position'] = positions.loc[keep_mask]

    return filtered.reset_index(drop=True), {
        'raw_player_count': raw_player_count,
        'eligible_player_count': int(len(filtered)),
        'removed_blank_name_count': int(blank_name_mask.sum()),
        'removed_unknown_name_count': int(unknown_name_mask.sum()),
        'removed_non_fantasy_position_count': int(invalid_position_mask.sum()),
        'eligible_positions': list(FANTASY_DRAFT_POSITIONS),
    }


def _current_round_number(pick_number: int, league_size: int) -> int:
    return max(1, math.ceil(max(1, int(pick_number)) / max(1, int(league_size))))


def _specialist_round_open(settings: LeagueSettings, round_number: int) -> bool:
    return round_number >= max(1, settings.round_count() - LATE_SPECIALIST_BUFFER_ROUNDS)


def _starter_position_needs(
    settings: LeagueSettings, roster_counts: dict[str, int]
) -> dict[str, int]:
    return {
        pos: max(0, settings.roster_spots.get(pos, 0) - roster_counts.get(pos, 0))
        for pos in POSITION_ORDER
    }


def _score_policy_rows(
    available: pd.DataFrame,
    league_settings: LeagueSettings,
    context: DraftContext,
) -> pd.DataFrame:
    if available.empty:
        return available.copy()

    rows = available.copy()
    next_turn_pick = next_pick_number(
        context.current_pick_number,
        league_settings.draft_position,
        league_settings.league_size,
    )
    current_round = _current_round_number(
        context.current_pick_number, league_settings.league_size
    )
    specialist_window_open = _specialist_round_open(league_settings, current_round)

    position_spread = (
        rows.groupby('position')['adp']
        .transform(lambda series: pd.to_numeric(series, errors='coerce').std(ddof=0))
        .fillna(rows['adp_std'])
        .fillna(2.5)
    )
    rows['availability_to_next_pick'] = [
        availability_probability(
            adp=row.adp,
            target_pick=next_turn_pick,
            adp_std=row.position_adp_spread,
            uncertainty_score=row.uncertainty_score,
        )
        for row in rows.assign(position_adp_spread=position_spread).itertuples(index=False)
    ]

    roster_counts = {pos: int(context.roster_counts.get(pos, 0)) for pos in POSITION_ORDER}
    starter_needs = _starter_position_needs(league_settings, roster_counts)
    offensive_needs = {pos: starter_needs.get(pos, 0) for pos in OFFENSIVE_POSITIONS}
    open_offensive_slots = sum(offensive_needs.values())
    total_need = sum(offensive_needs.values()) or 1
    position_counts_remaining = rows['position'].value_counts().to_dict()

    rows['starter_slot_urgency'] = rows['position'].map(
        lambda pos: (
            offensive_needs.get(pos, 0) / total_need
            if pos in OFFENSIVE_POSITIONS and open_offensive_slots > 0
            else 0.0
        )
    )
    rows['specialist_need_bonus'] = rows['position'].map(
        lambda pos: (
            2.5
            if pos in {'DST', 'K'}
            and specialist_window_open
            and starter_needs.get(pos, 0) > 0
            else 0.0
        )
    )
    rows['specialist_urgency'] = rows.apply(
        lambda row: float(
            rows.loc[rows['position'] == row['position'], 'player_name'].count()
        ),
        axis=1,
    )
    rows['specialist_urgency'] = rows.apply(
        lambda row: (
            (
                1.25 * (1.0 - float(row['availability_to_next_pick']))
                + 0.75
                * np.clip(
                    1.0 / max(1.0, float(row['specialist_urgency'])),
                    0.0,
                    1.0,
                )
            )
            if row['position'] in {'DST', 'K'}
            and specialist_window_open
            and starter_needs.get(row['position'], 0) > 0
            else 0.0
        ),
        axis=1,
    )
    rows['roster_fit_score'] = rows['starter_slot_urgency']
    rows['lineup_gain_now'] = rows.apply(
        lambda row: float(
            max(
                0.0,
                row['starter_delta']
                if row['position'] in OFFENSIVE_POSITIONS
                else row['replacement_delta'] * 0.35,
                row['replacement_delta'] * 0.70
                if row['position'] in OFFENSIVE_POSITIONS
                else row['replacement_delta'] * 0.20,
            )
        ),
        axis=1,
    )
    lineup_gain_rank = rows['lineup_gain_now'].rank(pct=True, method='average').fillna(0.0)

    def _position_run_risk(position: str) -> float:
        remaining = position_counts_remaining.get(position, 0)
        if position in OFFENSIVE_POSITIONS:
            demand = max(1, league_settings.league_size * max(1, offensive_needs.get(position, 0)))
        else:
            demand = max(1, league_settings.league_size)
        return float(np.clip(1.0 - remaining / demand, 0.0, 1.0))

    rows['position_run_risk'] = rows['position'].map(_position_run_risk)

    best_offensive_value = (
        float(
            rows.loc[rows['position'].isin(OFFENSIVE_POSITIONS), 'draft_score'].max()
        )
        if rows['position'].isin(OFFENSIVE_POSITIONS).any()
        else float(rows['draft_score'].max())
    )

    eligibility_reasons: list[str] = []
    policy_eligible: list[bool] = []
    for row in rows.itertuples(index=False):
        position = row.position
        current_count = roster_counts.get(position, 0)
        starter_requirement = league_settings.roster_spots.get(position, 0)
        reason = 'eligible'
        eligible = True
        if position in {'DST', 'K'} and not specialist_window_open:
            eligible = False
            reason = 'late_round_only'
        elif (
            position in {'QB', 'TE'}
            and open_offensive_slots > 0
            and current_count >= starter_requirement
            and float(row.draft_score) < (best_offensive_value + SECONDARY_QB_TE_VALUE_EDGE)
        ):
            eligible = False
            reason = 'starter_priority'
        policy_eligible.append(eligible)
        eligibility_reasons.append(reason)
    rows['policy_eligible'] = policy_eligible
    rows['policy_eligibility_reason'] = eligibility_reasons

    rows['expected_regret'] = (
        (0.55 * lineup_gain_rank)
        + (0.25 * rows['starter_slot_urgency'])
        + (0.20 * rows['position_run_risk'])
    ) * (1.0 - rows['availability_to_next_pick'])

    wait_signal: list[str] = []
    for row in rows.itertuples(index=False):
        if row.position in {'DST', 'K'} and specialist_window_open:
            wait_signal.append('late_round_stash_ok')
        elif float(row.availability_to_next_pick) < CONSERVATIVE_WAIT_SURVIVAL_THRESHOLD:
            wait_signal.append('low_survival')
        elif float(row.expected_regret) > CONSERVATIVE_WAIT_LINEUP_LOSS_THRESHOLD:
            wait_signal.append('too_much_lineup_loss')
        else:
            wait_signal.append('safe_to_wait')
    rows['wait_signal'] = wait_signal

    risk_bias = {'low': -0.03, 'medium': 0.0, 'high': 0.03}.get(
        league_settings.risk_tolerance.lower(), 0.0
    )
    rows['current_pick_utility'] = (
        rows['draft_score']
        + 0.32 * rows['starter_slot_urgency']
        + rows['specialist_need_bonus']
        + rows['specialist_urgency']
        + 0.22 * lineup_gain_rank
        + 0.08 * rows['posterior_prob_beats_replacement'].fillna(0.5)
        + 0.06 * rows['position_run_risk']
        + risk_bias * rows['upside_score']
    )
    rows.loc[~rows['policy_eligible'], 'current_pick_utility'] -= 2.5

    rows['wait_utility'] = (
        rows['draft_score'] * rows['availability_to_next_pick']
        + 0.06 * rows['upside_score']
        - 0.85 * rows['expected_regret']
    )
    return rows


def build_decision_table(
    player_frame: pd.DataFrame,
    league_settings: LeagueSettings | None = None,
    context: DraftContext | None = None,
) -> pd.DataFrame:
    """Build a player-level decision table aligned to the league context."""
    settings = league_settings or LeagueSettings()
    context = context or DraftContext(current_pick_number=settings.draft_position)

    df = normalize_player_frame(player_frame)
    df, fantasy_pool_metadata = _filter_fantasy_player_pool(df)
    df.attrs['fantasy_pool_metadata'] = fantasy_pool_metadata
    if df.empty:
        raise ValueError('player_frame does not contain eligible fantasy draft players')
    projection_series = _projection_series_for_settings(df, settings)
    if projection_series is not None:
        df['proj_points_mean'] = projection_series

    if df['proj_points_mean'].isna().all():
        raise ValueError('player_frame must contain a usable projection column')

    df['proj_points_mean'] = pd.to_numeric(df['proj_points_mean'], errors='coerce')

    # Project floors/ceilings from uncertainty or historical spread.
    uncertainty = df['uncertainty_score'].copy()
    if uncertainty.isna().all():
        uncertainty = (
            pd.to_numeric(df['std_projection'], errors='coerce')
            / df['proj_points_mean'].replace(0, np.nan)
        ).fillna(0.15)
    else:
        uncertainty = uncertainty.fillna(
            (
                pd.to_numeric(df['std_projection'], errors='coerce')
                / df['proj_points_mean'].replace(0, np.nan)
            ).fillna(0.15)
        )

    df['uncertainty_score'] = _clamp(uncertainty, 0.0, 1.0)

    spread_from_std = pd.to_numeric(df['std_projection'], errors='coerce')
    fallback_spread = (
        df['proj_points_mean'].abs() * (0.08 + 0.35 * df['uncertainty_score'])
    ).fillna(0.0)
    computed_floor = df['proj_points_mean'] - spread_from_std.fillna(fallback_spread)
    computed_ceiling = df['proj_points_mean'] + spread_from_std.fillna(fallback_spread) * 1.25
    df['proj_points_floor'] = pd.to_numeric(
        df.get('proj_points_floor'), errors='coerce'
    ).combine_first(computed_floor)
    df['proj_points_ceiling'] = pd.to_numeric(
        df.get('proj_points_ceiling'), errors='coerce'
    ).combine_first(computed_ceiling)

    # Replacement and starter baselines depend on league structure.
    starter_slots = settings.starters_by_position()
    replacement_slots = settings.effective_replacement_slots()
    df['starter_baseline'] = np.nan
    df['replacement_baseline'] = np.nan
    for position in df['position'].dropna().unique():
        starter_baseline = _position_baseline(
            df,
            position,
            starter_slots.get(position, 0),
            fallback=df['proj_points_mean'].mean(),
        )
        replacement_baseline = _position_baseline(
            df, position, replacement_slots.get(position, 0), fallback=starter_baseline
        )
        pos_mask = df['position'] == position
        df.loc[pos_mask, 'starter_baseline'] = starter_baseline
        df.loc[pos_mask, 'replacement_baseline'] = replacement_baseline

    df['starter_delta'] = df['proj_points_mean'] - df['starter_baseline']
    df['replacement_delta'] = df['proj_points_mean'] - df['replacement_baseline']
    df['posterior_prob_beats_replacement'] = pd.to_numeric(
        df.get('posterior_prob_beats_replacement'), errors='coerce'
    ).fillna(
        _clamp(
            pd.Series(
                1.0
                / (
                    1.0
                    + np.exp(
                        -(df['replacement_delta'] / np.maximum(spread_from_std.fillna(fallback_spread), 1.0))
                    )
                )
            ),
            0.0,
            1.0,
        )
    )

    # Market rank and relative market gap.
    if df['adp'].notna().any():
        df['market_rank'] = df['adp'].rank(method='first', ascending=True)
    elif df['market_rank'].notna().any():
        df['market_rank'] = pd.to_numeric(df['market_rank'], errors='coerce')
    else:
        df['market_rank'] = df['proj_points_mean'].rank(method='first', ascending=False)

    df['model_rank'] = df['proj_points_mean'].rank(method='first', ascending=False)
    df['market_gap'] = df['market_rank'] - df['model_rank']

    # Availability is based on the next time we expect to pick again in a snake draft.
    target_pick = next_pick_number(
        context.current_pick_number, settings.draft_position, settings.league_size
    )
    df['availability_at_pick'] = [
        availability_probability(
            adp=row.adp,
            target_pick=target_pick,
            adp_std=row.adp_std,
            uncertainty_score=row.uncertainty_score,
        )
        for row in df.itertuples(index=False)
    ]

    # Risk and upside signals.
    missing_history = pd.to_numeric(df['season_count'], errors='coerce').fillna(0.0)
    games_missed = pd.to_numeric(df['games_missed'], errors='coerce').fillna(0.0)
    age = pd.to_numeric(df['age'], errors='coerce').fillna(
        df['age'].median(skipna=True) if df['age'].notna().any() else 27.0
    )
    years = pd.to_numeric(df['years_in_league'], errors='coerce').fillna(0.0)
    team_change = pd.to_numeric(df['team_change'], errors='coerce').fillna(0.0)
    role_volatility = pd.to_numeric(df['role_volatility'], errors='coerce').fillna(0.0)
    site_disagreement = pd.to_numeric(df['site_disagreement'], errors='coerce').fillna(
        0.0
    )
    adp_std = pd.to_numeric(df['adp_std'], errors='coerce').fillna(0.0)

    history_penalty = _clamp(
        pd.Series(1.0 / np.maximum(1.0, missing_history + 1.0)), 0.0, 1.0
    )
    injury_penalty = _clamp(pd.Series(np.tanh(games_missed / 6.0)), 0.0, 1.0)
    age_penalty = _clamp(pd.Series(np.maximum(0.0, (age - 29.0) / 8.0)), 0.0, 1.0)
    role_penalty = _clamp(pd.Series(np.maximum(0.0, role_volatility)), 0.0, 1.0)
    team_penalty = _clamp(pd.Series(np.maximum(0.0, team_change)), 0.0, 1.0)
    disagreement_penalty = _clamp(
        pd.Series(np.maximum(0.0, np.maximum(site_disagreement, adp_std / 10.0))),
        0.0,
        1.0,
    )

    if 'projected_points_spread' in df.columns:
        spread = pd.to_numeric(df['projected_points_spread'], errors='coerce').fillna(
            0.0
        )
    else:
        spread = (df['proj_points_ceiling'] - df['proj_points_floor']) / 2.0
    upside_gap = (df['proj_points_ceiling'] - df['proj_points_mean']).fillna(0.0)

    posterior_reliability_penalty = 1.0 - df['posterior_prob_beats_replacement']
    df['fragility_score'] = _clamp(
        0.22 * history_penalty
        + 0.25 * injury_penalty
        + 0.12 * age_penalty
        + 0.16 * team_penalty
        + 0.15 * role_penalty
        + 0.10 * disagreement_penalty,
        0.0,
        1.0,
    )
    df['fragility_score'] = _clamp(
        0.65 * df['fragility_score'] + 0.35 * posterior_reliability_penalty,
        0.0,
        1.0,
    )

    df['upside_score'] = _clamp(
        _zscore(upside_gap).rank(pct=True)
        + 0.45 * df['posterior_prob_beats_replacement']
        + 0.35 * _zscore(df['availability_at_pick']).rank(pct=True)
        + 0.15 * _zscore(df['proj_points_mean']).rank(pct=True),
        0.0,
        1.0,
    )

    # Lean toward players who are both good and likely to survive.
    df['starter_need'] = 0.0
    roster_need = {
        'QB': max(
            0, settings.roster_spots.get('QB', 0) - context.roster_counts.get('QB', 0)
        ),
        'RB': max(
            0, settings.roster_spots.get('RB', 0) - context.roster_counts.get('RB', 0)
        ),
        'WR': max(
            0, settings.roster_spots.get('WR', 0) - context.roster_counts.get('WR', 0)
        ),
        'TE': max(
            0, settings.roster_spots.get('TE', 0) - context.roster_counts.get('TE', 0)
        ),
    }
    roster_need_total = sum(roster_need.values()) or 1
    for position, need in roster_need.items():
        if need > 0:
            df.loc[df['position'] == position, 'starter_need'] = (
                need / roster_need_total
            )

    scarcity_by_position = df.groupby('position')['player_name'].transform('count')
    df['position_scarcity'] = _clamp(
        1.0 / np.maximum(1.0, scarcity_by_position), 0.0, 1.0
    )

    risk_multiplier = {'low': 0.80, 'medium': 1.00, 'high': 1.18}.get(
        settings.risk_tolerance.lower(), 1.00
    )
    df['board_value_score'] = (
        0.40 * _zscore(df['proj_points_mean']).fillna(0.0)
        + 0.24 * _zscore(df['starter_delta']).fillna(0.0)
        + 0.18 * _zscore(df['replacement_delta']).fillna(0.0)
        + 0.10 * _zscore(df['posterior_prob_beats_replacement']).fillna(0.0)
        + 0.05 * _zscore(df['market_gap']).fillna(0.0)
        + 0.03 * _zscore(df['starter_need']).fillna(0.0)
        - (0.06 * risk_multiplier) * _zscore(df['fragility_score']).fillna(0.0)
    )
    df['draft_score'] = df['board_value_score'].fillna(0.0)
    df['draft_rank'] = (
        df['draft_score'].rank(method='first', ascending=False).astype(int)
    )
    df['simple_vor_proxy'] = df['replacement_delta']
    df['simple_vor_rank'] = (
        df['simple_vor_proxy'].rank(method='first', ascending=False).astype(int)
    )
    df['draft_tier'] = _assign_tiers(df['draft_score'])
    df['market_value_gap'] = df['model_rank'] - df['market_rank']
    df['why_flags'] = df.apply(_build_why_flags, axis=1)

    df = df.sort_values(
        ['draft_score', 'proj_points_mean'], ascending=[False, False]
    ).reset_index(drop=True)
    return df


def _assign_tiers(scores: pd.Series, num_tiers: int = 5) -> pd.Series:
    if scores.empty:
        return pd.Series(dtype='object')
    quantiles = pd.qcut(
        scores.rank(method='first', ascending=False),
        q=min(num_tiers, len(scores)),
        labels=False,
        duplicates='drop',
    )
    tiers = pd.Series(quantiles, index=scores.index).fillna(0).astype(int) + 1
    return tiers.map(lambda tier: f'Tier {tier}')


def _build_why_flags(row: pd.Series) -> str:
    flags: list[str] = []
    if (
        pd.notna(row.get('adp'))
        and pd.notna(row.get('market_rank'))
        and (row['market_rank'] - row['model_rank']) > 8
    ):
        flags.append('market_discount')
    if pd.notna(row.get('availability_at_pick')) and row['availability_at_pick'] < 0.40:
        flags.append('survival_risk')
    if pd.notna(row.get('upside_score')) and row['upside_score'] > 0.70:
        flags.append('upside_play')
    if pd.notna(row.get('fragility_score')) and row['fragility_score'] > 0.60:
        flags.append('fragile')
    if pd.notna(row.get('starter_delta')) and row['starter_delta'] > 0:
        flags.append('starter_gain')
    if (
        pd.notna(row.get('posterior_prob_beats_replacement'))
        and row['posterior_prob_beats_replacement'] > 0.70
    ):
        flags.append('posterior_confidence')
    if (
        row.get('position') in {'RB', 'WR'}
        and pd.notna(row.get('position_scarcity'))
        and row['position_scarcity'] > 0.10
    ):
        flags.append('scarce_position')
    if pd.notna(row.get('season_count')) and row['season_count'] <= 1:
        flags.append('cohort_prior')
    if pd.notna(row.get('site_disagreement')) and row['site_disagreement'] > 0.25:
        flags.append('projection_disagreement')
    return '|'.join(flags)


def next_pick_number(
    current_pick_number: int, draft_position: int, league_size: int
) -> int:
    """Return the next pick at which our team will draft in a snake draft."""
    current_pick_number = max(1, int(current_pick_number))
    draft_position = max(1, int(draft_position))
    league_size = max(1, int(league_size))

    picks = []
    rounds = max(1, math.ceil(current_pick_number / league_size) + 2)
    for round_num in range(1, rounds + 1):
        if round_num % 2 == 1:
            pick = (round_num - 1) * league_size + draft_position
        else:
            pick = round_num * league_size - draft_position + 1
        picks.append(pick)

    for pick in picks:
        if pick > current_pick_number:
            return pick
    return picks[-1]


def build_recommendations(
    decision_table: pd.DataFrame,
    league_settings: LeagueSettings,
    context: DraftContext,
    top_n: int = 5,
) -> pd.DataFrame:
    """Return the best current picks and survival-friendly wait options."""
    available = decision_table.copy()
    if context.drafted_players:
        drafted = context.drafted_set()
        available = available[
            ~available['player_name'].str.lower().isin(drafted)
        ].copy()
    if context.keepers:
        keepers = {name.lower() for name in context.keepers}
        available = available[
            ~available['player_name'].str.lower().isin(keepers)
        ].copy()

    if available.empty:
        return pd.DataFrame(
            columns=[
                'player_name',
                'position',
                'draft_score',
                'availability_at_pick',
                'expected_regret',
                'position_run_risk',
                'roster_fit_score',
                'wait_signal',
                'pick_mode',
                'recommendation_lane',
                'lane_rank',
                'rationale',
            ]
        )

    available = _score_policy_rows(available, league_settings, context)

    now = (
        available[available['policy_eligible']].sort_values(
            ['current_pick_utility', 'draft_score'], ascending=[False, False]
        )
        .head(top_n)
        .copy()
    )
    if now.empty:
        now = (
            available.sort_values(
                ['current_pick_utility', 'draft_score'], ascending=[False, False]
            )
            .head(top_n)
            .copy()
        )
    now['pick_mode'] = 'now'
    now['recommendation_lane'] = ['pick_now'] + ['fallback'] * max(0, len(now) - 1)
    now['lane_rank'] = list(range(1, len(now) + 1))
    now['rationale'] = now.apply(_rationale_now, axis=1)

    wait_candidates = available[
        available['wait_signal'].isin(['safe_to_wait', 'late_round_stash_ok'])
    ].copy()
    wait = (
        wait_candidates.sort_values(
            ['wait_utility', 'availability_to_next_pick'], ascending=[False, False]
        )
        .head(top_n)
        .copy()
    )
    wait = wait[~wait['player_name'].isin(now['player_name'])].copy()
    if wait.empty:
        wait = (
            available[~available['player_name'].isin(now['player_name'])]
            .sort_values(
                ['availability_to_next_pick', 'draft_score'],
                ascending=[False, False],
            )
            .head(top_n)
            .copy()
        )
    wait['pick_mode'] = 'wait'
    wait['recommendation_lane'] = 'can_wait'
    wait['lane_rank'] = list(range(1, len(wait) + 1))
    if wait.empty:
        wait['rationale'] = pd.Series(dtype='object')
    else:
        wait['rationale'] = wait.apply(_rationale_wait, axis=1)

    cols = [
        'player_name',
        'position',
        'proj_points_mean',
        'proj_points_floor',
        'proj_points_ceiling',
        'adp',
        'market_rank',
        'availability_to_next_pick',
        'replacement_delta',
        'starter_delta',
        'upside_score',
        'fragility_score',
        'posterior_prob_beats_replacement',
        'board_value_score',
        'draft_score',
        'draft_tier',
        'why_flags',
        'position_run_risk',
        'roster_fit_score',
        'expected_regret',
        'wait_signal',
        'policy_eligible',
        'current_pick_utility',
        'wait_utility',
        'pick_mode',
        'recommendation_lane',
        'lane_rank',
        'rationale',
    ]
    combined = pd.concat([now[cols], wait[cols]], ignore_index=True)
    lane_order = {'pick_now': 0, 'fallback': 1, 'can_wait': 2}
    combined['lane_order'] = combined['recommendation_lane'].map(lane_order).fillna(9)
    combined = combined.sort_values(
        ['lane_order', 'lane_rank', 'draft_score'], ascending=[True, True, False]
    ).drop(columns=['lane_order'])
    combined = combined.reset_index(drop=True)
    return combined


def _rationale_now(row: pd.Series) -> str:
    parts = [f'{row["position"]} value', f'score={row["draft_score"]:.2f}']
    if row.get('why_flags'):
        parts.append(row['why_flags'].replace('|', ', '))
    if pd.notna(row.get('availability_to_next_pick')):
        parts.append(f'survival={row["availability_to_next_pick"]:.0%}')
    if row.get('policy_eligibility_reason') and row['policy_eligibility_reason'] != 'eligible':
        parts.append(row['policy_eligibility_reason'].replace('_', ' '))
    return '; '.join(parts)


def _rationale_wait(row: pd.Series) -> str:
    parts = ['wait candidate', f'score={row["draft_score"]:.2f}']
    if pd.notna(row.get('availability_to_next_pick')):
        parts.append(f'survival={row["availability_to_next_pick"]:.0%}')
    if pd.notna(row.get('expected_regret')):
        parts.append(f'regret={row["expected_regret"]:.2f}')
    if row.get('wait_signal'):
        parts.append(row['wait_signal'].replace('_', ' '))
    return '; '.join(parts)


def _recommendation_rows_to_records(frame: pd.DataFrame) -> list[dict[str, Any]]:
    if frame is None or frame.empty:
        return []
    return frame.to_dict(orient='records')


def build_live_recommendation_snapshot(
    decision_table: pd.DataFrame,
    recommendations: pd.DataFrame,
    roster_scenarios: pd.DataFrame,
    league_settings: LeagueSettings,
    context: DraftContext,
) -> dict[str, Any]:
    """Build the live contingency view used by the dashboard and workbook."""
    next_turn_pick = next_pick_number(
        context.current_pick_number,
        league_settings.draft_position,
        league_settings.league_size,
    )
    recommendation_rows = recommendations.copy()
    if recommendation_rows.empty:
        recommendation_rows = pd.DataFrame(
            columns=[
                'player_name',
                'position',
                'draft_score',
                'availability_to_next_pick',
                'expected_regret',
                'roster_fit_score',
                'position_run_risk',
                'why_flags',
                'wait_signal',
                'recommendation_lane',
                'lane_rank',
            ]
        )

    pick_now = recommendation_rows[
        recommendation_rows['recommendation_lane'] == 'pick_now'
    ].head(1)
    fallbacks = recommendation_rows[
        recommendation_rows['recommendation_lane'] == 'fallback'
    ].head(5)
    can_wait = recommendation_rows[
        recommendation_rows['recommendation_lane'] == 'can_wait'
    ].sort_values(
        ['availability_to_next_pick', 'draft_score'], ascending=[False, False]
    )

    roster_need = {
        pos: max(
            0, league_settings.roster_spots.get(pos, 0) - context.roster_counts.get(pos, 0)
        )
        for pos in ['QB', 'RB', 'WR', 'TE', 'FLEX', 'DST', 'K']
    }

    best_roster_paths = roster_scenarios.head(3).to_dict(orient='records')
    wait_survival = can_wait.head(5).to_dict(orient='records')

    primary = pick_now.head(1).to_dict(orient='records')
    primary_pick = primary[0] if primary else {}

    return {
        'current_pick_number': context.current_pick_number,
        'next_pick_number': next_turn_pick,
        'roster_need': roster_need,
        'pick_now': primary_pick,
        'fallbacks': _recommendation_rows_to_records(fallbacks),
        'can_wait': _recommendation_rows_to_records(can_wait.head(5)),
        'wait_candidates': wait_survival,
        'best_roster_paths': best_roster_paths,
        'board_pressure': (
            decision_table.groupby('position')['player_name'].count().sort_values(
                ascending=False
            ).to_dict()
            if not decision_table.empty
            else {}
        ),
    }


def build_tier_cliffs(decision_table: pd.DataFrame) -> pd.DataFrame:
    """Compute position-consistent tier cliffs from the decision table."""
    rows = []
    for position, group in decision_table.groupby('position'):
        ordered = group.sort_values('proj_points_mean', ascending=False).reset_index(
            drop=True
        )
        if ordered.empty:
            continue
        diffs = ordered['proj_points_mean'].diff(-1).fillna(0.0)
        threshold = diffs.quantile(0.75) if len(diffs) > 3 else diffs.max()
        threshold = float(np.nan_to_num(threshold, nan=0.0))
        cliff_mask = (
            diffs >= threshold if threshold > 0 else pd.Series([False] * len(ordered))
        )
        for idx, row in ordered.iterrows():
            rows.append(
                {
                    'position': position,
                    'player_name': row['player_name'],
                    'draft_score': row['draft_score'],
                    'proj_points_mean': row['proj_points_mean'],
                    'tier_cliff_distance': float(diffs.iloc[idx]),
                    'is_tier_cliff': bool(cliff_mask.iloc[idx])
                    if len(cliff_mask) > idx
                    else False,
                    'draft_tier': row['draft_tier'],
                }
            )
    return pd.DataFrame(rows)


def build_roster_scenarios(
    decision_table: pd.DataFrame, league_settings: LeagueSettings
) -> pd.DataFrame:
    """Generate simple roster-construction archetypes and their utility tradeoffs."""
    archetypes = [
        ('balanced', {'RB': 1.0, 'WR': 1.0, 'QB': 0.8, 'TE': 0.8}),
        ('hero_rb', {'RB': 1.2, 'WR': 0.9, 'QB': 0.7, 'TE': 0.8}),
        ('zero_rb', {'RB': 0.8, 'WR': 1.2, 'QB': 0.7, 'TE': 0.8}),
        ('elite_qb', {'RB': 0.95, 'WR': 0.95, 'QB': 1.15, 'TE': 0.85}),
        ('tight_end_attack', {'RB': 0.95, 'WR': 0.95, 'QB': 0.8, 'TE': 1.15}),
    ]

    rows = []
    top = decision_table.head(80)
    for name, weights in archetypes:
        score = 0.0
        count = 0
        for pos, weight in weights.items():
            pos_df = top[top['position'] == pos].head(
                max(1, league_settings.roster_spots.get(pos, 0) + 1)
            )
            if pos_df.empty:
                continue
            score += float(pos_df['draft_score'].mean()) * weight
            count += 1
        rows.append(
            {
                'scenario': name,
                'utility_proxy': score / max(1, count),
                'mean_draft_score': float(top['draft_score'].mean()),
                'mean_upside_score': float(top['upside_score'].mean()),
                'mean_fragility_score': float(top['fragility_score'].mean()),
                'recommended_build': ', '.join(
                    [f'{pos}:{weight}' for pos, weight in weights.items()]
                ),
            }
        )
    return (
        pd.DataFrame(rows)
        .sort_values('utility_proxy', ascending=False)
        .reset_index(drop=True)
    )


def _starter_points_from_roster(
    team_roster: pd.DataFrame, league_settings: LeagueSettings
) -> float:
    if team_roster.empty:
        return 0.0
    total = 0.0
    chosen = pd.Series(False, index=team_roster.index)
    for pos in ['QB', 'RB', 'WR', 'TE', 'DST', 'K']:
        need = league_settings.roster_spots.get(pos, 0)
        pos_pool = team_roster[(team_roster['position'] == pos) & (~chosen)]
        if pos_pool.empty or need <= 0:
            continue
        take = pos_pool.nlargest(need, 'actual_points')
        chosen.loc[take.index] = True
        total += float(take['actual_points'].sum())

    flex_slots = league_settings.roster_spots.get('FLEX', 0)
    if flex_slots > 0:
        flex_pool = team_roster[
            (team_roster['position'].isin(['RB', 'WR', 'TE'])) & (~chosen)
        ]
        if not flex_pool.empty:
            take = flex_pool.nlargest(flex_slots, 'actual_points')
            chosen.loc[take.index] = True
            total += float(take['actual_points'].sum())
    return float(total)


def _team_actual_points(
    team_players: pd.DataFrame, league_settings: LeagueSettings
) -> float:
    if team_players.empty:
        return 0.0
    del league_settings
    return float(pd.to_numeric(team_players['actual_points'], errors='coerce').fillna(0.0).sum())


def _build_strategy_ranker(strategy_name: str):
    def _rank(frame: pd.DataFrame) -> pd.DataFrame:
        frame = frame.copy()
        if strategy_name == 'market':
            frame['strategy_rank'] = frame['market_rank'].rank(
                method='first', ascending=True
            )
            frame['strategy_score'] = -frame['market_rank']
        elif strategy_name in {'vor', 'historical_vor_proxy'}:
            strategy_series = pd.to_numeric(
                frame.get('historical_vor_proxy'), errors='coerce'
            ).fillna(frame['replacement_delta'])
            frame['strategy_rank'] = strategy_series.rank(
                method='first', ascending=False
            )
            frame['strategy_score'] = strategy_series
        elif strategy_name == 'consensus':
            frame['strategy_rank'] = (
                0.65 * frame['proj_points_mean'] + 0.35 * frame['availability_at_pick']
            ).rank(method='first', ascending=False)
            frame['strategy_score'] = (
                0.65 * frame['proj_points_mean'] + 0.35 * frame['availability_at_pick']
            )
        elif strategy_name == 'draft_score':
            frame['strategy_rank'] = frame['draft_score'].rank(
                method='first', ascending=False
            )
            frame['strategy_score'] = frame['draft_score']
        else:
            raise ValueError(f'Unknown strategy: {strategy_name}')
        return frame.sort_values(
            ['strategy_score', 'proj_points_mean'], ascending=[False, False]
        ).reset_index(drop=True)

    return _rank


def build_historical_vor_proxy_table(
    test_frame: pd.DataFrame, train_frame: pd.DataFrame
) -> pd.DataFrame:
    """Build a clearly labeled historical VOR proxy from pre-holdout information."""
    if test_frame is None or test_frame.empty:
        return pd.DataFrame(columns=['Season', 'Name', 'Position', 'historical_vor_proxy'])

    proxy = test_frame.copy()
    replacement_values: dict[str, float] = {}
    for position in proxy['Position'].dropna().unique():
        pos_train = pd.to_numeric(
            train_frame.loc[train_frame['Position'] == position, 'actual_points'],
            errors='coerce',
        ).dropna()
        if pos_train.empty:
            replacement_values[position] = 0.0
        else:
            slot_count = max(1, min(len(pos_train), 12))
            replacement_values[position] = float(
                pos_train.sort_values(ascending=False).reset_index(drop=True).iloc[
                    slot_count - 1
                ]
            )

    proxy['historical_vor_proxy'] = proxy.apply(
        lambda row: float(row['proj_points_mean'])
        - replacement_values.get(row['Position'], 0.0),
        axis=1,
    )
    proxy['historical_vor_proxy_rank'] = proxy.groupby('Position')[
        'historical_vor_proxy'
    ].rank(method='first', ascending=False)
    proxy['historical_vor_proxy_label'] = 'historical_vor_proxy'
    proxy['source_name'] = proxy.get('source_name', 'historical_vor_proxy')
    return proxy


def _roster_counts_from_team(team_roster: list[dict[str, Any]]) -> dict[str, int]:
    counts = {pos: 0 for pos in POSITION_ORDER}
    for row in team_roster:
        position = _normalize_position(row.get('position'))
        if position in counts:
            counts[position] += 1
    return counts


def _starter_fill_timing(
    pick_log: list[dict[str, Any]], league_settings: LeagueSettings
) -> dict[str, int | None]:
    counts = {pos: 0 for pos in POSITION_ORDER}
    fill_timing: dict[str, int | None] = {
        pos: None for pos in POSITION_ORDER if league_settings.roster_spots.get(pos, 0) > 0
    }
    for event in pick_log:
        position = _normalize_position(event.get('position'))
        if position not in counts:
            continue
        counts[position] += 1
        needed = league_settings.roster_spots.get(position, 0)
        if needed > 0 and counts[position] >= needed and fill_timing.get(position) is None:
            fill_timing[position] = int(event.get('pick_number', 0))
    return fill_timing


def _secondary_qb_te_before_starters_filled(
    pick_log: list[dict[str, Any]], league_settings: LeagueSettings
) -> int:
    counts = {pos: 0 for pos in POSITION_ORDER}
    total = 0
    for event in pick_log:
        position = _normalize_position(event.get('position'))
        open_offensive_slots = sum(
            max(0, league_settings.roster_spots.get(pos, 0) - counts.get(pos, 0))
            for pos in OFFENSIVE_POSITIONS
        )
        if (
            position in {'QB', 'TE'}
            and counts.get(position, 0) >= league_settings.roster_spots.get(position, 0)
            and open_offensive_slots > 0
        ):
            total += 1
        if position in counts:
            counts[position] += 1
    return total


def _draft_team_from_pool(
    available: pd.DataFrame,
    team_roster: list[dict[str, Any]],
    strategy_name: str,
    round_number: int,
    current_pick_number: int,
    league_settings: LeagueSettings,
) -> tuple[pd.DataFrame, list[dict[str, Any]], dict[str, Any]]:
    if available.empty:
        return available, team_roster, {}

    frame = available.copy()
    pick_event: dict[str, Any] = {
        'strategy': strategy_name,
        'pick_number': int(current_pick_number),
        'round_number': int(round_number),
    }

    if strategy_name == 'draft_score':
        roster_counts = _roster_counts_from_team(team_roster)
        starter_needs = _starter_position_needs(league_settings, roster_counts)
        context = DraftContext(
            current_pick_number=current_pick_number,
            your_players={_safe_string(row.get('player_name')) for row in team_roster},
            roster_counts=roster_counts,
        )
        scored = _score_policy_rows(frame, league_settings, context)
        eligible = scored[scored['policy_eligible']].copy()
        if eligible.empty:
            eligible = scored.copy()
        ranked = eligible.sort_values(
            ['current_pick_utility', 'draft_score'], ascending=[False, False]
        ).reset_index(drop=True)
        remaining_rounds = max(1, league_settings.round_count() - round_number + 1)
        open_specialists = [
            pos for pos in ['DST', 'K'] if starter_needs.get(pos, 0) > 0
        ]
        forced_specialists = ranked[ranked['position'].isin(open_specialists)]
        if (
            open_specialists
            and remaining_rounds <= (len(open_specialists) + 1)
            and not forced_specialists.empty
        ):
            choice = forced_specialists.sort_values(
                ['current_pick_utility', 'draft_score'],
                ascending=[False, False],
            ).iloc[0]
        else:
            choice = ranked.iloc[0]

        wait_candidates = scored[
            scored['wait_signal'].isin(['safe_to_wait', 'late_round_stash_ok'])
        ]
        wait_candidates = wait_candidates[
            wait_candidates['player_name'] != choice['player_name']
        ].sort_values(
            ['wait_utility', 'availability_to_next_pick'], ascending=[False, False]
        )
        if not wait_candidates.empty:
            top_wait = wait_candidates.iloc[0]
            pick_event['top_wait_candidate'] = {
                'player_name': top_wait['player_name'],
                'position': top_wait['position'],
                'actual_points': float(top_wait.get('actual_points', 0.0) or 0.0),
                'availability_to_next_pick': float(
                    top_wait.get('availability_to_next_pick', 0.0) or 0.0
                ),
                'expected_regret': float(top_wait.get('expected_regret', 0.0) or 0.0),
                'wait_signal': _safe_string(top_wait.get('wait_signal')),
            }
            pick_event['wait_candidate_survival'] = float(
                top_wait.get('availability_to_next_pick', 0.0) or 0.0
            )
            pick_event['wait_candidate_regret'] = float(
                top_wait.get('expected_regret', 0.0) or 0.0
            )
    else:
        ranker = _build_strategy_ranker(strategy_name)
        ranked = ranker(frame)
        choice = ranked.iloc[0]

    choice_record = choice.to_dict()
    team_roster.append(choice_record)
    pick_event.update(
        {
            'player_name': _safe_string(choice_record.get('player_name')),
            'position': _normalize_position(choice_record.get('position')),
            'actual_points': float(_coerce_float(choice_record.get('actual_points'), 0.0)),
            'availability_to_next_pick': float(
                _coerce_float(choice_record.get('availability_to_next_pick'), np.nan)
            )
            if pd.notna(choice_record.get('availability_to_next_pick'))
            else None,
            'expected_regret': float(
                _coerce_float(choice_record.get('expected_regret'), np.nan)
            )
            if pd.notna(choice_record.get('expected_regret'))
            else None,
            'wait_signal': _safe_string(choice_record.get('wait_signal')),
        }
    )
    available = frame[frame['player_name'] != choice['player_name']].reset_index(
        drop=True
    )
    return available, team_roster, pick_event


def _bootstrap_delta_summary(
    strategy_values: list[float],
    baseline_values: list[float],
    *,
    iterations: int = 2000,
    seed: int = 7,
) -> dict[str, Any]:
    if not strategy_values or not baseline_values:
        return {}
    deltas = np.asarray(strategy_values, dtype=float) - np.asarray(
        baseline_values, dtype=float
    )
    if deltas.size == 0:
        return {}
    rng = np.random.default_rng(seed)
    sample_indices = rng.integers(0, deltas.size, size=(iterations, deltas.size))
    sample_means = deltas[sample_indices].mean(axis=1)
    return {
        'mean_delta': float(deltas.mean()),
        'median_delta': float(np.median(deltas)),
        'ci_low': float(np.quantile(sample_means, 0.025)),
        'ci_high': float(np.quantile(sample_means, 0.975)),
        'season_win_rate': float((deltas > 0).mean()),
        'supports_positive_delta': bool(np.quantile(sample_means, 0.025) > 0),
    }


def run_draft_backtest(
    season_history: pd.DataFrame,
    league_settings: LeagueSettings | None = None,
    holdout_years: Iterable[int] | None = None,
) -> dict[str, Any]:
    """Run a snake-draft backtest across historical seasons."""
    settings = league_settings or LeagueSettings()
    if season_history is None or season_history.empty:
        raise ValueError('season_history is empty')

    df = season_history.copy()
    required = {'Season', 'Name', 'Position', 'FantPt'}
    missing = required.difference(df.columns)
    if missing:
        raise ValueError(
            f'season_history is missing required columns: {sorted(missing)}'
        )

    df['Season'] = pd.to_numeric(df['Season'], errors='coerce').astype('Int64')
    df['Name'] = df['Name'].map(_safe_string)
    df['Position'] = df['Position'].map(_normalize_position)
    df['FantPt'] = pd.to_numeric(df['FantPt'], errors='coerce')
    df = df.dropna(subset=['Season', 'Name', 'Position', 'FantPt']).copy()
    raw_pool_metadata = {
        'raw_player_count': int(len(df)),
    }
    df = df.rename(columns={'Name': 'player_name', 'Position': 'position'})
    df, fantasy_pool_metadata = _filter_fantasy_player_pool(df)
    df = df.rename(columns={'player_name': 'Name', 'position': 'Position'})
    raw_pool_metadata.update(fantasy_pool_metadata)
    if df.empty:
        raise ValueError('No eligible fantasy players available after pool filtering')

    season_table = aggregate_season_player_table(df).rename(
        columns={'fantasy_points': 'actual_points'}
    )

    seasons = sorted(int(s) for s in season_table['Season'].unique())
    if holdout_years is None:
        holdout_years = seasons[2:] if len(seasons) > 2 else seasons[1:]
    holdout_years = [int(year) for year in holdout_years]
    if not holdout_years:
        raise ValueError('No holdout seasons available')

    by_season = []
    for holdout_year in holdout_years:
        train = season_table[season_table['Season'] < holdout_year].copy()
        test = season_table[season_table['Season'] == holdout_year].copy()
        if train.empty or test.empty:
            continue

        posterior_inputs = build_posterior_projection_table(
            train_history=train.rename(columns={'actual_points': 'fantasy_points'}),
            target_frame=test.rename(columns={'actual_points': 'fantasy_points'}),
            holdout_year=holdout_year,
            replacement_quantile=0.2,
            min_history_seasons=0,
        )
        posterior_inputs['Name'] = posterior_inputs['player_name']
        posterior_inputs['Position'] = posterior_inputs['position']
        posterior_inputs['FantPt'] = posterior_inputs['actual_points']
        posterior_inputs['season_count'] = posterior_inputs['season_count'].fillna(0.0)
        posterior_inputs['years_in_league'] = posterior_inputs[
            'years_in_league'
        ].fillna(posterior_inputs['season_count'])
        posterior_inputs['games_missed'] = posterior_inputs['games_missed'].fillna(
            posterior_inputs['games_missed_mean']
        )
        posterior_inputs['site_disagreement'] = posterior_inputs[
            'site_disagreement'
        ].fillna(0.0)
        posterior_inputs['role_label'] = 'posterior'
        posterior_inputs['adp'] = posterior_inputs['adp'].fillna(
            posterior_inputs['market_proxy_score'].rank(
                method='first', ascending=False
            )
        )
        posterior_inputs['adp_std'] = posterior_inputs['adp_std'].fillna(
            np.maximum(2.0, posterior_inputs['posterior_std'] * 0.18)
        )

        context = DraftContext(
            current_pick_number=settings.draft_position, drafted_players=set()
        )
        decision_table = build_decision_table(posterior_inputs, settings, context)
        proxy_table = posterior_inputs[
            [
                'player_name',
                'position',
                'historical_vor_proxy_score',
                'historical_vor_proxy_point',
                'market_proxy_score',
            ]
        ].rename(
            columns={
                'historical_vor_proxy_score': 'historical_vor_proxy',
                'historical_vor_proxy_point': 'historical_vor_proxy_point',
            }
        )
        decision_table = decision_table.merge(
            proxy_table,
            on=['player_name', 'position'],
            how='left',
        )
        decision_table['historical_vor_proxy'] = decision_table[
            'historical_vor_proxy'
        ].fillna(decision_table['replacement_delta'])
        if 'actual_points' not in decision_table.columns:
            decision_table['actual_points'] = np.nan
        decision_table['actual_points'] = decision_table['actual_points'].fillna(
            pd.to_numeric(decision_table.get('FantPt'), errors='coerce')
        )
        by_strategy = {}
        for strategy in ['market', 'historical_vor_proxy', 'consensus', 'draft_score']:
            available = decision_table.copy()
            team_rosters = [[] for _ in range(settings.league_size)]
            our_pick_log: list[dict[str, Any]] = []
            pending_wait_candidate: dict[str, Any] | None = None
            picks_per_team = settings.round_count()
            draft_order = []
            for round_number in range(1, picks_per_team + 1):
                if round_number % 2 == 1:
                    order = list(range(settings.league_size))
                else:
                    order = list(reversed(range(settings.league_size)))
                draft_order.extend(order)

            available_pool = available.copy()
            for pick_index, team_idx in enumerate(
                draft_order[: settings.league_size * picks_per_team], start=1
            ):
                round_number = math.ceil(pick_index / settings.league_size)
                if team_idx == settings.draft_position - 1:
                    pending_candidate_was_available = (
                        pending_wait_candidate is not None
                        and pending_wait_candidate.get('player_name')
                        in set(available_pool['player_name'])
                    )
                    available_pool, team_rosters[team_idx], pick_event = _draft_team_from_pool(
                        available_pool,
                        team_rosters[team_idx],
                        strategy,
                        round_number,
                        pick_index,
                        settings,
                    )
                    if strategy == 'draft_score':
                        if pending_wait_candidate and not pending_candidate_was_available:
                            pick_event['realized_pass_regret'] = float(
                                max(
                                    0.0,
                                    float(
                                        pending_wait_candidate.get('actual_points', 0.0)
                                    )
                                    - float(pick_event.get('actual_points', 0.0)),
                                )
                            )
                        else:
                            pick_event['realized_pass_regret'] = 0.0
                        pending_wait_candidate = pick_event.get('top_wait_candidate')
                    our_pick_log.append(pick_event)
                else:
                    available_pool, team_rosters[team_idx], _ = _draft_team_from_pool(
                        available_pool,
                        team_rosters[team_idx],
                        'market',
                        round_number,
                        pick_index,
                        settings,
                    )
                if available_pool.empty:
                    break

            our_roster = (
                pd.DataFrame(team_rosters[settings.draft_position - 1])
                if team_rosters[settings.draft_position - 1]
                else pd.DataFrame(columns=decision_table.columns)
            )
            if not our_roster.empty and 'actual_points' not in our_roster.columns:
                our_roster['actual_points'] = (
                    test.set_index(['Name', 'Position'])
                    .reindex(
                        pd.MultiIndex.from_frame(our_roster[['Name', 'Position']])
                    )['actual_points']
                    .to_numpy()
            )
            roster_points = _team_actual_points(our_roster, settings)
            lineup_points = _starter_points_from_roster(our_roster, settings)
            fill_timing = _starter_fill_timing(our_pick_log, settings)
            early_specialist_cutoff = max(
                1, settings.round_count() - LATE_SPECIALIST_BUFFER_ROUNDS
            )
            early_specialist_picks = sum(
                1
                for event in our_pick_log
                if event.get('position') in {'DST', 'K'}
                and int(event.get('round_number', 0)) < early_specialist_cutoff
            )
            wait_survivals = [
                float(event.get('wait_candidate_survival'))
                for event in our_pick_log
                if pd.notna(event.get('wait_candidate_survival'))
            ]
            realized_regrets = [
                float(event.get('realized_pass_regret'))
                for event in our_pick_log
                if pd.notna(event.get('realized_pass_regret'))
            ]

            metric_rank = {}
            for metric_name, score_col in [
                ('market', 'market_rank'),
                ('historical_vor_proxy', 'historical_vor_proxy'),
                ('consensus', 'proj_points_mean'),
                ('draft_score', 'draft_score'),
            ]:
                test_eval = test.copy()
                test_eval['player_key'] = (
                    test_eval['Name'] + '||' + test_eval['Position']
                )
                pred = decision_table.copy()
                pred['player_key'] = pred['player_name'] + '||' + pred['position']
                if score_col not in pred.columns:
                    score_col = 'draft_score'
                pred_scores = pred[['player_key', score_col]].rename(
                    columns={score_col: 'predicted_score'}
                )
                merged = test_eval.merge(pred_scores, on='player_key', how='left')
                if (
                    merged['predicted_score'].notna().sum() > 1
                    and merged['predicted_score'].nunique(dropna=True) > 1
                    and merged['actual_points'].nunique(dropna=True) > 1
                ):
                    if metric_name == 'market':
                        corr_input = -merged['predicted_score']
                    else:
                        corr_input = merged['predicted_score']
                    corr = float(
                        spearmanr(corr_input, merged['actual_points']).correlation
                    )
                else:
                    corr = 0.0
                metric_rank[metric_name] = float(np.nan_to_num(corr, nan=0.0))

            by_strategy[strategy] = {
                'strategy': strategy,
                'holdout_year': holdout_year,
                'our_team_actual_points': float(roster_points),
                'our_team_lineup_points': float(lineup_points),
                'predicted_rank_correlation': metric_rank.get(strategy, 0.0),
                'drafted_players': [
                    row.get('player_name', '')
                    for row in team_rosters[settings.draft_position - 1]
                ],
                'position_counts': pd.Series(
                    [
                        row.get('position', 'UNKNOWN')
                        for row in team_rosters[settings.draft_position - 1]
                    ]
                )
                .value_counts()
                .to_dict(),
                'strategy_label': strategy,
                'policy_diagnostics': {
                    'starter_lineup_points': float(lineup_points),
                    'full_roster_points': float(roster_points),
                    'starter_fill_timing': fill_timing,
                    'early_specialist_picks': int(early_specialist_picks),
                    'secondary_qb_te_before_starters_filled': int(
                        _secondary_qb_te_before_starters_filled(our_pick_log, settings)
                    ),
                    'mean_wait_recommendation_survival': float(np.mean(wait_survivals))
                    if wait_survivals
                    else None,
                    'mean_realized_pass_regret': float(np.mean(realized_regrets))
                    if realized_regrets
                    else 0.0,
                },
            }

        best_strategy = max(
            by_strategy.values(), key=lambda item: item['our_team_lineup_points']
        )
        by_season.append(
            {
                'holdout_year': holdout_year,
                'by_strategy': by_strategy,
                'winner': best_strategy['strategy'],
                'lineup_points_winner': float(best_strategy['our_team_lineup_points']),
            }
        )

    if not by_season:
        raise ValueError('Unable to run backtest with the provided seasons')

    overall_rows = []
    for strategy in ['market', 'historical_vor_proxy', 'consensus', 'draft_score']:
        season_values = [
            season['by_strategy'][strategy]['our_team_lineup_points']
            for season in by_season
            if strategy in season['by_strategy']
        ]
        if not season_values:
            continue
        overall_rows.append(
            {
                'strategy': strategy,
                'mean_lineup_points': float(np.mean(season_values)),
                'median_lineup_points': float(np.median(season_values)),
                'season_count': len(season_values),
            }
        )
    overall = pd.DataFrame(overall_rows).sort_values(
        'mean_lineup_points', ascending=False
    )
    winner = (
        _pick_first_row(overall['strategy']) if not overall.empty else 'draft_score'
    )

    draft_score_values = [
        season['by_strategy']['draft_score']['our_team_lineup_points']
        for season in by_season
        if 'draft_score' in season['by_strategy']
    ]
    vor_values = [
        season['by_strategy']['historical_vor_proxy']['our_team_lineup_points']
        for season in by_season
        if 'historical_vor_proxy' in season['by_strategy']
    ]
    bootstrap = _bootstrap_delta_summary(draft_score_values, vor_values)
    consensus_values = [
        season['by_strategy']['consensus']['our_team_lineup_points']
        for season in by_season
        if 'consensus' in season['by_strategy']
    ]
    consensus_bootstrap = _bootstrap_delta_summary(draft_score_values, consensus_values)

    diagnostics_rollup = {}
    for strategy in ['market', 'historical_vor_proxy', 'consensus', 'draft_score']:
        strategy_diags = [
            season['by_strategy'][strategy].get('policy_diagnostics', {})
            for season in by_season
            if strategy in season['by_strategy']
        ]
        if not strategy_diags:
            continue
        diagnostics_rollup[strategy] = {
            'mean_early_specialist_picks': float(
                np.mean(
                    [diag.get('early_specialist_picks', 0) for diag in strategy_diags]
                )
            ),
            'mean_secondary_qb_te_before_starters_filled': float(
                np.mean(
                    [
                        diag.get('secondary_qb_te_before_starters_filled', 0)
                        for diag in strategy_diags
                    ]
                )
            ),
            'mean_wait_recommendation_survival': float(
                np.mean(
                    [
                        diag.get('mean_wait_recommendation_survival')
                        for diag in strategy_diags
                        if diag.get('mean_wait_recommendation_survival') is not None
                    ]
                )
            )
            if any(
                diag.get('mean_wait_recommendation_survival') is not None
                for diag in strategy_diags
            )
            else None,
            'mean_realized_pass_regret': float(
                np.mean([diag.get('mean_realized_pass_regret', 0.0) for diag in strategy_diags])
            ),
        }

    return {
        'model_type': 'draft_decision_backtest',
        'evaluation_scope': {
            'forecast_target': 'held_out_season_fantasy_points',
            'decision_target': 'snake_draft_lineup_points',
            'primary_objective': 'starter_lineup_points',
            'wait_policy': 'conservative',
            'draft_score_label': 'posterior_contextual_policy',
            'eligible_positions': list(FANTASY_DRAFT_POSITIONS),
            'eligible_player_universe': raw_pool_metadata,
        },
        'league_settings': settings.to_dict(),
        'holdout_years': holdout_years,
        'by_season': by_season,
        'overall': {
            'by_strategy': overall.to_dict(orient='records'),
            'winner': winner,
            'draft_score_vs_historical_vor_proxy': bootstrap,
            'draft_score_vs_consensus': consensus_bootstrap,
            'policy_diagnostics': diagnostics_rollup,
        },
    }


def build_dashboard_payload(
    decision_table: pd.DataFrame,
    recommendations: pd.DataFrame,
    tier_cliffs: pd.DataFrame,
    roster_scenarios: pd.DataFrame,
    source_freshness: pd.DataFrame,
    league_settings: LeagueSettings,
    backtest: dict[str, Any] | None = None,
    context: DraftContext | None = None,
    scoring_presets: dict[str, Any] | None = None,
    analysis_provenance: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Build a JSON-serializable payload for the local dashboard."""
    backtest = backtest or {}
    context = context or DraftContext(
        current_pick_number=league_settings.draft_position
    )
    live_state = build_live_recommendation_snapshot(
        decision_table,
        recommendations,
        roster_scenarios,
        league_settings,
        context,
    )
    recommendation_inputs_cols = [
        'player_name',
        'position',
        'proj_points_mean',
        'proj_points_floor',
        'proj_points_ceiling',
        'adp',
        'adp_std',
        'market_rank',
        'availability_at_pick',
        'availability_to_next_pick',
        'starter_delta',
        'replacement_delta',
        'posterior_prob_beats_replacement',
        'board_value_score',
        'upside_score',
        'fragility_score',
        'draft_score',
        'draft_tier',
        'why_flags',
        'current_pick_utility',
        'wait_utility',
        'position_run_risk',
        'roster_fit_score',
        'expected_regret',
        'wait_signal',
        'recommendation_lane',
        'lane_rank',
        'pick_mode',
        'rationale',
    ]
    recommendation_source = recommendations if not recommendations.empty else decision_table
    recommendation_inputs = recommendation_source[
        [
            column
            for column in recommendation_inputs_cols
            if column in recommendation_source.columns
        ]
    ].copy()
    recommendation_summary = recommendations.head(12).to_dict(orient='records')
    position_summary = _dashboard_position_summary(decision_table)
    active_scoring_preset = _resolve_scoring_preset_key(league_settings)
    analysis_provenance = analysis_provenance or {}
    bayesian_vor_summary = _build_bayesian_vor_summary(decision_table, backtest)
    decision_evidence = _build_decision_evidence(
        decision_table,
        backtest,
        analysis_provenance,
        bayesian_vor_summary=bayesian_vor_summary,
    )
    selected_player = (
        _pick_first_row(recommendations['player_name'])
        if not recommendations.empty
        else _pick_first_row(decision_table['player_name'])
    )
    return {
        'generated_at': datetime.now().isoformat(timespec='seconds'),
        'league_settings': league_settings.to_dict(),
        'runtime_controls': {
            'risk_tolerance_options': ['low', 'medium', 'high'],
            'supported_scoring_presets': list(SCORING_PRESETS.keys()),
            'active_scoring_preset': active_scoring_preset,
        },
        'current_pick_number': context.current_pick_number,
        'next_pick_number': next_pick_number(
            context.current_pick_number,
            league_settings.draft_position,
            league_settings.league_size,
        ),
        'current_draft_context_defaults': {
            'current_pick_number': context.current_pick_number,
            'next_pick_number': live_state['next_pick_number'],
            'taken_players': sorted(context.drafted_set()),
            'your_players': sorted(context.your_players),
            'roster_counts': context.roster_counts,
        },
        'selected_player': selected_player,
        'recommendation_inputs': recommendation_inputs.to_dict(orient='records'),
        'recommendation_summary': recommendation_summary,
        'live_state': live_state,
        'decision_table': decision_table.to_dict(orient='records'),
        'scoring_presets': scoring_presets or {},
        'position_summary': position_summary,
        'tier_cliffs': tier_cliffs.to_dict(orient='records'),
        'roster_scenarios': roster_scenarios.to_dict(orient='records'),
        'source_freshness': source_freshness.to_dict(orient='records'),
        'analysis_provenance': analysis_provenance,
        'decision_evidence': decision_evidence,
        'backtest': backtest,
        'supporting_math': _dashboard_supporting_math(decision_table),
        'metric_glossary': METRIC_GLOSSARY,
        'model_overview': {
            'headline': 'The draft board uses posterior player projections plus a starter-first decision policy.',
            'plain_english': [
                'Each player starts with a posterior mean, floor, ceiling, and uncertainty estimate built from historical player performance plus local feature signals.',
                'The board value score is a clean posterior-based ranking, and the live recommendation layer separately decides whether to pick now or wait based on starter urgency and survival risk.',
                'The simple VOR view is still available as a frozen baseline through replacement delta and rank-gap summaries.',
            ],
            'limitations': [
                'This is still a decision policy layered on posterior player estimates, not a full structural model of the entire draft room.',
                'Backtests here are internal holdout seasons and should be treated as directional evidence, not definitive proof.',
            ],
        },
        'bayesian_vor_summary': bayesian_vor_summary,
    }


def _format_sheet(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    header_fill = PatternFill(
        start_color='173F5F', end_color='173F5F', fill_type='solid'
    )
    header_font = Font(bold=True, color='FFFFFF', size=11)
    body_font = Font(size=10)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(
                horizontal='left', vertical='top', wrap_text=True
            )
    for column in worksheet.columns:
        max_len = 0
        for cell in column:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                continue
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = min(
            max_len + 2, 48
        )


def _write_dataframe_sheet(
    workbook: openpyxl.Workbook, title: str, df: pd.DataFrame
) -> None:
    ws = workbook.create_sheet(title)
    if df is None or df.empty:
        ws.append(['No data available'])
        return
    cols = list(df.columns)
    ws.append(cols)
    for _, row in df.iterrows():
        ws.append([row[col] for col in cols])
    _format_sheet(ws)


def export_workbook(
    decision_table: pd.DataFrame,
    recommendations: pd.DataFrame,
    tier_cliffs: pd.DataFrame,
    roster_scenarios: pd.DataFrame,
    source_freshness: pd.DataFrame,
    output_path: Path | str,
    league_settings: LeagueSettings,
    backtest: dict[str, Any] | None = None,
    context: DraftContext | None = None,
    dashboard_payload: dict[str, Any] | None = None,
) -> Path:
    """Write the portable draft workbook."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    board_cols = [
        'player_name',
        'position',
        'proj_points_mean',
        'proj_points_floor',
        'proj_points_ceiling',
        'adp',
        'market_rank',
        'availability_at_pick',
        'replacement_delta',
        'starter_delta',
        'upside_score',
        'fragility_score',
        'draft_score',
        'draft_rank',
        'draft_tier',
        'why_flags',
    ]
    board = decision_table[
        [col for col in board_cols if col in decision_table.columns]
    ].copy()
    _write_dataframe_sheet(wb, 'Big Board', board)

    by_position = decision_table.sort_values(
        ['position', 'draft_score'], ascending=[True, False]
    ).copy()
    _write_dataframe_sheet(wb, 'By Position', by_position)

    live_summary = pd.DataFrame(
        [
            {
                'metric': 'current_pick_number',
                'value': context.current_pick_number if context else league_settings.draft_position,
            },
            {
                'metric': 'next_pick_number',
                'value': next_pick_number(
                    context.current_pick_number
                    if context
                    else league_settings.draft_position,
                    league_settings.draft_position,
                    league_settings.league_size,
                ),
            },
            {
                'metric': 'draft_position',
                'value': league_settings.draft_position,
            },
            {
                'metric': 'league_size',
                'value': league_settings.league_size,
            },
        ]
    )
    _write_dataframe_sheet(wb, 'Live Context', live_summary)

    pick_now = recommendations[
        recommendations.get('recommendation_lane', pd.Series(dtype=object)) == 'pick_now'
    ].copy()
    if pick_now.empty and not recommendations.empty:
        pick_now = recommendations.head(1).copy()
    _write_dataframe_sheet(wb, 'Pick Now', pick_now)

    fallbacks = recommendations[
        recommendations.get('recommendation_lane', pd.Series(dtype=object)) == 'fallback'
    ].copy()
    _write_dataframe_sheet(wb, 'Fallback Ladder', fallbacks)

    can_wait = recommendations[
        recommendations.get('recommendation_lane', pd.Series(dtype=object)) == 'can_wait'
    ].copy()
    _write_dataframe_sheet(wb, 'Can Wait', can_wait)

    _write_dataframe_sheet(wb, 'My Picks', recommendations.copy())

    _write_dataframe_sheet(wb, 'Tier Cliffs', tier_cliffs)
    availability = decision_table[
        [
            'player_name',
            'position',
            'adp',
            'availability_at_pick',
            'draft_score',
            'why_flags',
        ]
    ].copy()
    _write_dataframe_sheet(wb, 'Availability', availability)

    round_rows = []
    current_round_count = league_settings.round_count()
    for rnd in range(1, current_round_count + 1):
        row = {'round': rnd}
        cut = max(1, league_settings.league_size * rnd)
        top = decision_table.head(cut)
        row['best_player'] = _pick_first_row(top['player_name'])
        row['best_position'] = _pick_first_row(top['position'])
        row['mean_draft_score'] = (
            float(top['draft_score'].mean()) if not top.empty else np.nan
        )
        round_rows.append(row)
    _write_dataframe_sheet(wb, 'Targets By Round', pd.DataFrame(round_rows))

    _write_dataframe_sheet(wb, 'Roster Construction Scenarios', roster_scenarios)
    _write_dataframe_sheet(
        wb,
        'Player Notes',
        decision_table[
            [
                'player_name',
                'position',
                'why_flags',
                'draft_tier',
                'fragility_score',
                'upside_score',
            ]
        ],
    )

    diagnostics = pd.DataFrame(
        [
            {'metric': 'player_count', 'value': int(len(decision_table))},
            {'metric': 'recommendation_count', 'value': int(len(recommendations))},
            {
                'metric': 'draft_score_mean',
                'value': float(decision_table['draft_score'].mean()),
            },
            {
                'metric': 'draft_score_std',
                'value': float(decision_table['draft_score'].std(ddof=0)),
            },
            {
                'metric': 'top_draft_score',
                'value': float(decision_table['draft_score'].max()),
            },
            {
                'metric': 'availability_mean',
                'value': float(decision_table['availability_at_pick'].mean()),
            },
            {
                'metric': 'source_count',
                'value': int(source_freshness['source_name'].nunique())
                if not source_freshness.empty
                else 0,
            },
        ]
    )
    if backtest:
        diagnostics = pd.concat(
            [
                diagnostics,
                pd.DataFrame(
                    [
                        {
                            'metric': 'backtest_model',
                            'value': backtest.get('model_type', ''),
                        }
                    ]
                ),
            ],
            ignore_index=True,
    )
    _write_dataframe_sheet(wb, 'Model Diagnostics', diagnostics)
    _write_dataframe_sheet(wb, 'Source Freshness', source_freshness)

    payload = dashboard_payload or {}
    decision_evidence = payload.get('decision_evidence') or _build_decision_evidence(
        decision_table,
        backtest or {},
        payload.get('analysis_provenance') if isinstance(payload, dict) else None,
    )
    evidence_rows = pd.DataFrame(
        [
            {'metric': 'status', 'value': decision_evidence.get('status', 'unknown')},
            {'metric': 'headline', 'value': decision_evidence.get('headline', '')},
            {'metric': 'winner', 'value': decision_evidence.get('winner', '')},
            {
                'metric': 'delta_mean_lineup_points',
                'value': decision_evidence.get('delta_mean_lineup_points'),
            },
            {'metric': 'season_count', 'value': decision_evidence.get('season_count')},
            {
                'metric': 'freshness_status',
                'value': decision_evidence.get('freshness', {}).get('status', 'unknown'),
            },
            {
                'metric': 'override_used',
                'value': decision_evidence.get('freshness', {}).get('override_used', False),
            },
            {
                'metric': 'reason_unavailable',
                'value': decision_evidence.get('reason_unavailable', ''),
            },
        ]
    )
    _write_dataframe_sheet(wb, 'Decision Evidence', evidence_rows)

    strategy_summary = pd.DataFrame(decision_evidence.get('strategy_summary', []))
    if not strategy_summary.empty:
        _write_dataframe_sheet(wb, 'Backtest Summary', strategy_summary)

    provenance = payload.get('analysis_provenance', {}) if isinstance(payload, dict) else {}
    provenance_rows = []
    overall_freshness = provenance.get('overall_freshness') or {}
    provenance_rows.append(
        {
            'scope': 'overall',
            'source_name': 'analysis_provenance',
            'status': overall_freshness.get('status', 'unknown'),
            'override_used': overall_freshness.get('override_used', False),
            'latest_expected_year': '',
            'latest_found_year': '',
            'warnings': ' | '.join(overall_freshness.get('warnings', [])),
        }
    )
    for label, entry in (provenance.get('sources') or {}).items():
        provenance_rows.append(
            {
                'scope': label,
                'source_name': entry.get('source_name', label),
                'status': entry.get('status', 'unknown'),
                'override_used': entry.get('override_used', False),
                'latest_expected_year': entry.get('latest_expected_year', ''),
                'latest_found_year': entry.get('latest_found_year', ''),
                'warnings': ' | '.join(entry.get('warnings', [])),
            }
        )
    _write_dataframe_sheet(wb, 'Artifact Provenance', pd.DataFrame(provenance_rows))

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _build_charts(
    decision_table: pd.DataFrame,
    recommendations: pd.DataFrame,
    league_settings: LeagueSettings,
) -> list[Any]:
    import plotly.graph_objects as go

    top = decision_table.head(120).copy()
    fig1 = go.Figure()
    for position in sorted(top['position'].dropna().unique()):
        sub = top[top['position'] == position]
        fig1.add_trace(
            go.Scatter(
                x=sub['adp'],
                y=sub['draft_score'],
                mode='markers',
                name=position,
                text=sub['player_name'],
                marker=dict(size=9, opacity=0.8),
            )
        )
    fig1.update_layout(
        title='ADP vs Draft Score',
        xaxis_title='ADP',
        yaxis_title='Draft Score',
        height=420,
    )

    fig2 = go.Figure()
    if not recommendations.empty:
        fig2.add_trace(
            go.Bar(
                x=recommendations['player_name'].head(10),
                y=recommendations['availability_to_next_pick'].head(10),
                name='Survival to next pick',
            )
        )
    fig2.update_layout(
        title='Availability by Pick', yaxis_title='Probability', height=420
    )

    fig3 = go.Figure()
    fig3.add_trace(
        go.Scatter(
            x=top['fragility_score'],
            y=top['upside_score'],
            mode='markers',
            text=top['player_name'],
            marker=dict(
                color=top['draft_score'], colorscale='Viridis', showscale=True, size=10
            ),
        )
    )
    fig3.update_layout(
        title='Upside vs Fragility',
        xaxis_title='Fragility',
        yaxis_title='Upside',
        height=420,
    )

    return [fig1, fig2, fig3]


def export_dashboard_html(
    decision_table: pd.DataFrame,
    recommendations: pd.DataFrame,
    output_path: Path | str,
    league_settings: LeagueSettings,
    backtest: dict[str, Any] | None = None,
    source_freshness: pd.DataFrame | None = None,
    dashboard_payload: dict[str, Any] | None = None,
    generated_label: str | None = None,
) -> Path:
    """Write a local interactive HTML dashboard."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    backtest = backtest or {}
    source_freshness = (
        source_freshness if source_freshness is not None else pd.DataFrame()
    )
    payload = dashboard_payload or build_dashboard_payload(
        decision_table,
        recommendations,
        build_tier_cliffs(decision_table),
        build_roster_scenarios(decision_table, league_settings),
        source_freshness,
        league_settings,
        backtest=backtest,
        context=DraftContext(current_pick_number=league_settings.draft_position),
    )
    payload_json = json.dumps(payload, default=str)
    payload_generated_at = payload.get('generated_at')
    resolved_generated_label = generated_label
    if resolved_generated_label is None:
        try:
            if payload_generated_at:
                resolved_generated_label = datetime.fromisoformat(
                    str(payload_generated_at)
                ).strftime('%Y-%m-%d %H:%M')
        except Exception:
            resolved_generated_label = None
    if resolved_generated_label is None:
        resolved_generated_label = datetime.now().strftime('%Y-%m-%d %H:%M')
    html = """
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>FFBayes Draft War Room</title>
  <style>
    :root {
      color-scheme: dark;
      --bg: #07111f;
      --panel: rgba(15, 23, 42, 0.88);
      --panel-strong: rgba(17, 24, 39, 0.98);
      --panel-soft: rgba(15, 23, 42, 0.62);
      --text: #f8fafc;
      --muted: #94a3b8;
      --border: rgba(148, 163, 184, 0.18);
      --accent: #38bdf8;
      --accent-soft: rgba(56, 189, 248, 0.16);
      --good: #34d399;
      --warn: #f59e0b;
      --bad: #f87171;
      --queue: #a78bfa;
      --shadow: 0 24px 60px rgba(0, 0, 0, 0.28);
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      min-height: 100vh;
      background:
        radial-gradient(circle at top left, rgba(56, 189, 248, 0.18), transparent 30%),
        radial-gradient(circle at top right, rgba(129, 140, 248, 0.14), transparent 22%),
        linear-gradient(180deg, #030712 0%, #0b1324 100%);
      color: var(--text);
      font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, sans-serif;
      padding: 18px;
    }
    h1, h2, h3, h4, p { margin: 0; }
    button, input, select {
      font: inherit;
      color: var(--text);
    }
    button, select, input[type="number"], input[type="search"] {
      background: rgba(255, 255, 255, 0.05);
      border: 1px solid rgba(148, 163, 184, 0.20);
      border-radius: 12px;
      padding: 10px 12px;
    }
    button {
      cursor: pointer;
      transition: background 0.15s ease, border-color 0.15s ease;
    }
    button:hover, button.is-active {
      background: rgba(255, 255, 255, 0.10);
      border-color: rgba(56, 189, 248, 0.40);
    }
    button:disabled {
      cursor: not-allowed;
      opacity: 0.45;
    }
    input[type="checkbox"] {
      accent-color: var(--accent);
    }
    .shell {
      max-width: 1880px;
      margin: 0 auto;
      display: grid;
      gap: 16px;
    }
    .topbar, .panel {
      background: linear-gradient(180deg, rgba(15, 23, 42, 0.96), rgba(15, 23, 42, 0.92));
      border: 1px solid var(--border);
      border-radius: 22px;
      box-shadow: var(--shadow);
    }
    .topbar {
      padding: 20px 22px;
      display: grid;
      gap: 16px;
    }
    .topbar-head {
      display: flex;
      flex-wrap: wrap;
      justify-content: space-between;
      gap: 12px;
      align-items: flex-start;
    }
    .title-wrap {
      display: grid;
      gap: 8px;
    }
    .title-row {
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      align-items: center;
    }
    .title-row h1 {
      font-size: 30px;
      letter-spacing: -0.03em;
    }
    .subtitle {
      color: var(--muted);
      font-size: 14px;
      max-width: 900px;
      line-height: 1.45;
    }
    .pill-row, .toolbar-row, .filter-row {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
    }
    .toolbar-stack {
      display: grid;
      gap: 8px;
      justify-items: end;
    }
    .toolbar-note {
      max-width: 340px;
      text-align: right;
    }
    .pill {
      display: inline-flex;
      align-items: center;
      gap: 6px;
      padding: 6px 10px;
      border-radius: 999px;
      border: 1px solid rgba(56, 189, 248, 0.24);
      background: var(--accent-soft);
      color: #d7f5ff;
      font-size: 12px;
    }
    .pill.good { border-color: rgba(52, 211, 153, 0.30); background: rgba(52, 211, 153, 0.14); color: #d2faec; }
    .pill.warn { border-color: rgba(245, 158, 11, 0.34); background: rgba(245, 158, 11, 0.14); color: #fde7b0; }
    .pill.bad { border-color: rgba(248, 113, 113, 0.32); background: rgba(248, 113, 113, 0.12); color: #fecaca; }
    .layout {
      display: grid;
      grid-template-columns: minmax(290px, 0.9fr) minmax(0, 1.7fr) minmax(330px, 1fr);
      gap: 16px;
      align-items: start;
    }
    .column, .stack {
      display: grid;
      gap: 16px;
      align-content: start;
    }
    .panel {
      padding: 18px;
    }
    .panel.strong {
      background: var(--panel-strong);
    }
    .panel h2 {
      font-size: 18px;
      margin-bottom: 6px;
    }
    .subtle {
      color: var(--muted);
      font-size: 12px;
      line-height: 1.5;
    }
    .hero-pick {
      display: grid;
      gap: 12px;
    }
    .hero-name {
      font-size: 28px;
      letter-spacing: -0.03em;
      font-weight: 700;
    }
    .metric-grid {
      display: grid;
      gap: 10px;
      grid-template-columns: repeat(2, minmax(0, 1fr));
    }
    .metric {
      background: rgba(255, 255, 255, 0.04);
      border: 1px solid rgba(148, 163, 184, 0.12);
      border-radius: 14px;
      padding: 10px 12px;
    }
    .metric .label {
      display: block;
      margin-bottom: 4px;
      font-size: 11px;
      text-transform: uppercase;
      letter-spacing: 0.06em;
      color: var(--muted);
    }
    .metric .value {
      display: block;
      font-size: 15px;
      font-weight: 600;
    }
    .lane-list, .mini-list, .details-stack, .settings-grid {
      display: grid;
      gap: 10px;
    }
    .lane-item, .mini-item {
      padding: 12px;
      border-radius: 14px;
      border: 1px solid rgba(148, 163, 184, 0.12);
      background: rgba(255, 255, 255, 0.04);
      display: grid;
      gap: 6px;
    }
    .lane-item-header, .mini-item-header, .split {
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 8px;
    }
    .item-title {
      font-weight: 600;
    }
    .item-meta {
      color: var(--muted);
      font-size: 12px;
      line-height: 1.4;
    }
    .board-controls {
      display: grid;
      gap: 10px;
      margin-bottom: 12px;
    }
    .search-wrap {
      display: flex;
      gap: 10px;
      flex-wrap: wrap;
      align-items: center;
    }
    .search-wrap input[type="search"] {
      flex: 1 1 300px;
    }
    .board-summary {
      display: flex;
      flex-wrap: wrap;
      justify-content: space-between;
      gap: 10px;
      color: var(--muted);
      font-size: 12px;
    }
    .board-table-wrap {
      overflow: auto;
      border: 1px solid rgba(148, 163, 184, 0.12);
      border-radius: 16px;
      max-height: 74vh;
    }
    table {
      width: 100%;
      border-collapse: collapse;
      font-size: 13px;
    }
    th, td {
      text-align: left;
      vertical-align: top;
      padding: 10px 10px;
      border-bottom: 1px solid rgba(148, 163, 184, 0.10);
    }
    th {
      position: sticky;
      top: 0;
      z-index: 1;
      background: rgba(9, 16, 32, 0.96);
      color: #9bdcff;
      font-size: 12px;
      text-transform: uppercase;
      letter-spacing: 0.05em;
    }
    tr.is-selected {
      outline: 1px solid rgba(56, 189, 248, 0.55);
      background: rgba(56, 189, 248, 0.10);
    }
    tr.is-taken {
      opacity: 0.55;
    }
    tr.is-mine {
      background: rgba(52, 211, 153, 0.10);
    }
    tr.is-queued {
      background: rgba(167, 139, 250, 0.10);
    }
    .status-badge {
      display: inline-flex;
      align-items: center;
      border-radius: 999px;
      padding: 4px 8px;
      font-size: 11px;
      border: 1px solid rgba(148, 163, 184, 0.18);
      background: rgba(148, 163, 184, 0.12);
    }
    .status-badge.available { color: #dbeafe; }
    .status-badge.taken { color: #fecaca; border-color: rgba(248, 113, 113, 0.30); background: rgba(248, 113, 113, 0.10); }
    .status-badge.mine { color: #d2faec; border-color: rgba(52, 211, 153, 0.30); background: rgba(52, 211, 153, 0.12); }
    .status-badge.queued { color: #e9ddff; border-color: rgba(167, 139, 250, 0.30); background: rgba(167, 139, 250, 0.12); }
    .tiny {
      font-size: 12px;
      color: var(--muted);
    }
    .action-group {
      display: flex;
      flex-wrap: wrap;
      gap: 6px;
    }
    .action-group button {
      padding: 6px 8px;
      border-radius: 10px;
      font-size: 12px;
    }
    .settings-grid {
      grid-template-columns: repeat(2, minmax(0, 1fr));
    }
    .field {
      display: grid;
      gap: 6px;
    }
    .field label {
      font-size: 12px;
      color: var(--muted);
    }
    .field.full {
      grid-column: 1 / -1;
    }
    .inspector-title {
      display: grid;
      gap: 6px;
      margin-bottom: 12px;
    }
    .summary-box {
      padding: 12px;
      border-radius: 14px;
      background: rgba(255, 255, 255, 0.04);
      border: 1px solid rgba(148, 163, 184, 0.12);
      color: #dbeafe;
      line-height: 1.5;
      font-size: 14px;
    }
    .bar-stack {
      display: grid;
      gap: 10px;
    }
    .bar-row {
      display: grid;
      gap: 5px;
    }
    .bar-head {
      display: flex;
      justify-content: space-between;
      gap: 8px;
      font-size: 12px;
    }
    .bar-track {
      height: 10px;
      border-radius: 999px;
      overflow: hidden;
      background: rgba(148, 163, 184, 0.16);
    }
    .bar-fill {
      height: 100%;
      border-radius: 999px;
      background: linear-gradient(90deg, rgba(56, 189, 248, 0.75), rgba(14, 165, 233, 0.98));
    }
    .bar-fill.warn {
      background: linear-gradient(90deg, rgba(245, 158, 11, 0.75), rgba(234, 88, 12, 0.98));
    }
    .bar-fill.good {
      background: linear-gradient(90deg, rgba(52, 211, 153, 0.78), rgba(5, 150, 105, 0.98));
    }
    .bar-fill.bad {
      background: linear-gradient(90deg, rgba(248, 113, 113, 0.78), rgba(220, 38, 38, 0.98));
    }
    details {
      border: 1px solid rgba(148, 163, 184, 0.12);
      border-radius: 14px;
      background: rgba(255, 255, 255, 0.03);
      padding: 10px 12px;
    }
    summary {
      cursor: pointer;
      font-weight: 600;
      color: #dbeafe;
      list-style: none;
    }
    summary::-webkit-details-marker {
      display: none;
    }
    .details-body {
      margin-top: 12px;
      display: grid;
      gap: 12px;
    }
    .glossary-list {
      display: grid;
      gap: 8px;
    }
    .glossary-item {
      padding: 10px;
      border-radius: 12px;
      border: 1px solid rgba(148, 163, 184, 0.10);
      background: rgba(255, 255, 255, 0.03);
    }
    .glossary-item strong {
      display: block;
      margin-bottom: 4px;
    }
    .roster-chip-row {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
    }
    .empty {
      color: var(--muted);
      font-size: 13px;
    }
    .notice {
      padding: 10px 12px;
      border-radius: 12px;
      border: 1px solid rgba(245, 158, 11, 0.24);
      background: rgba(245, 158, 11, 0.10);
      color: #fde7b0;
      font-size: 13px;
      line-height: 1.45;
    }
    @media (max-width: 1440px) {
      .layout {
        grid-template-columns: 1fr;
      }
      .board-table-wrap {
        max-height: 58vh;
      }
    }
    @media (max-width: 760px) {
      body {
        padding: 10px;
      }
      .topbar, .panel {
        border-radius: 18px;
      }
      .metric-grid, .settings-grid {
        grid-template-columns: 1fr;
      }
      .title-row h1 {
        font-size: 24px;
      }
    }
  </style>
</head>
<body>
  <div class="shell">
    <section class="topbar">
      <div class="topbar-head">
        <div class="title-wrap">
          <div class="title-row">
            <h1>FFBayes Draft War Room</h1>
            <span class="pill">Live draft mode</span>
            <span class="pill" id="storage-pill">Saved locally</span>
          </div>
          <p class="subtitle">Operate the draft from this board: update league shape, scoring preset, current pick, queue, taken players, and your roster without leaving the dashboard.</p>
        </div>
        <div class="toolbar-stack">
          <div class="toolbar-row">
            <button type="button" id="undo-button">Undo</button>
            <button type="button" id="redo-button">Redo</button>
            <button type="button" id="finalize-button">Finalize Draft</button>
          </div>
          <div class="tiny toolbar-note" id="finalize-note"></div>
        </div>
      </div>
      <div class="pill-row" id="status-pills"></div>
      <div class="tiny">Generated __GENERATED_LABEL__</div>
    </section>

    <section class="layout">
      <div class="column">
        <section class="panel strong">
          <h2>Pick Now</h2>
          <p class="subtle">Best current move after adjusting for your league, roster, and next-pick survival.</p>
          <div class="hero-pick" id="primary-card"></div>
        </section>

        <section class="panel">
          <div class="split">
            <div>
              <h2>Fallback Ladder</h2>
              <p class="subtle">If the top target goes, pivot here.</p>
            </div>
            <span class="pill">Immediate pivots</span>
          </div>
          <div class="lane-list" id="fallback-list"></div>
        </section>

        <section class="panel">
          <div class="split">
            <div>
              <h2>Can Wait</h2>
              <p class="subtle">Strong values with the best shot to survive to your next turn.</p>
            </div>
            <span class="pill">Patience lane</span>
          </div>
          <div class="lane-list" id="wait-list"></div>
        </section>

        <section class="panel">
          <h2>Queue & Roster</h2>
          <p class="subtle">Keep a short watchlist and confirm your current build at a glance.</p>
          <div class="mini-list" id="queue-list"></div>
          <div class="details-stack" style="margin-top: 12px;">
            <div class="roster-chip-row" id="my-roster"></div>
            <div class="metric-grid" id="roster-need-grid"></div>
          </div>
        </section>
      </div>

      <div class="column">
        <section class="panel strong">
          <div class="split">
            <div>
              <h2>Full Player Board</h2>
              <p class="subtle">Search every player, keep taken rows visible if you want, and click a row to inspect the model reasoning.</p>
            </div>
            <span class="pill" id="board-count-pill">0 players</span>
          </div>
          <div class="board-controls">
            <div class="search-wrap">
              <input id="player-search" type="search" placeholder="Search name, team, or position" />
            </div>
            <div class="board-summary">
              <span id="board-summary-text">Showing players</span>
              <span id="preset-summary-text"></span>
              <span>Marking Taken or Mine advances the draft automatically.</span>
            </div>
          </div>
          <div class="board-table-wrap">
            <table>
              <thead>
                <tr>
                  <th>Player</th>
                  <th>Status</th>
                  <th>Board value score</th>
                  <th>VOR proxy</th>
                  <th>Next-pick survival</th>
                  <th>Why now</th>
                  <th>Actions</th>
                </tr>
              </thead>
              <tbody id="board-table"></tbody>
            </table>
          </div>
        </section>
      </div>

      <div class="column">
        <section class="panel">
          <h2>Draft Controls</h2>
          <p class="subtle">These settings are now dashboard-first. Changes update the board immediately.</p>
          <div class="settings-grid">
            <div class="field">
              <label for="scoring-preset">Scoring preset</label>
              <select id="scoring-preset"></select>
            </div>
            <div class="field">
              <label for="risk-tolerance">Risk tolerance</label>
              <select id="risk-tolerance">
                <option value="low">Low</option>
                <option value="medium">Medium</option>
                <option value="high">High</option>
              </select>
            </div>
            <div class="field">
              <label for="league-size">League size</label>
              <input id="league-size" type="number" min="2" max="20" />
            </div>
            <div class="field">
              <label for="draft-position">Draft position</label>
              <input id="draft-position" type="number" min="1" max="20" />
            </div>
            <div class="field">
              <label for="current-pick-number">Current pick</label>
              <input id="current-pick-number" type="number" min="1" />
            </div>
            <div class="field">
              <label for="bench-slots">Bench slots</label>
              <input id="bench-slots" type="number" min="0" max="12" />
            </div>
            <div class="field">
              <label for="roster-qb">QB starters</label>
              <input id="roster-qb" type="number" min="0" max="4" />
            </div>
            <div class="field">
              <label for="roster-rb">RB starters</label>
              <input id="roster-rb" type="number" min="0" max="6" />
            </div>
            <div class="field">
              <label for="roster-wr">WR starters</label>
              <input id="roster-wr" type="number" min="0" max="6" />
            </div>
            <div class="field">
              <label for="roster-te">TE starters</label>
              <input id="roster-te" type="number" min="0" max="4" />
            </div>
            <div class="field">
              <label for="roster-flex">Flex starters</label>
              <input id="roster-flex" type="number" min="0" max="4" />
            </div>
            <div class="field full">
              <div class="notice" id="preset-notice"></div>
            </div>
          </div>
        </section>

        <section class="panel strong">
          <h2>Selected Player</h2>
          <p class="subtle">Click any row to inspect plain-English reasoning, score components, and where the model disagrees with simple VOR.</p>
          <div id="player-inspector"></div>
        </section>

        <section class="panel">
          <h2>Model Notes</h2>
          <div class="details-stack">
            <details>
              <summary>What the model is doing</summary>
              <div class="details-body" id="model-overview"></div>
            </details>
            <details>
              <summary>Decision evidence</summary>
              <div class="details-body" id="decision-evidence"></div>
            </details>
            <details>
              <summary>Metric glossary</summary>
              <div class="details-body" id="metric-glossary"></div>
            </details>
            <details>
              <summary>Freshness and provenance</summary>
              <div class="details-body">
                <div id="provenance-panel"></div>
                <div class="metric-grid" id="support-metrics"></div>
                <div class="board-table-wrap" style="max-height: 220px;">
                  <table>
                    <thead>
                      <tr><th>Source</th><th>Status</th><th>Override</th><th>Expected / found</th></tr>
                    </thead>
                    <tbody id="freshness-table"></tbody>
                  </table>
                </div>
              </div>
            </details>
          </div>
        </section>
      </div>
    </section>
  </div>

  <script>
    window.FFBAYES_DASHBOARD = __PAYLOAD_JSON__;

    (() => {
      const data = window.FFBAYES_DASHBOARD || {};
      const STORAGE_KEY = 'ffbayes-dashboard-state-v2';
      const POSITION_KEYS = ['QB', 'RB', 'WR', 'TE', 'FLEX', 'DST', 'K'];
      const FANTASY_DRAFT_POSITIONS = ['QB', 'RB', 'WR', 'TE', 'DST', 'K'];
      const OFFENSIVE_POSITIONS = ['QB', 'RB', 'WR', 'TE'];
      const FLEX_WEIGHTS = Object.assign({ RB: 0.45, WR: 0.45, TE: 0.10 }, (data.league_settings && data.league_settings.flex_weights) || {});
      const LATE_SPECIALIST_BUFFER_ROUNDS = 2;
      const CONSERVATIVE_WAIT_SURVIVAL_THRESHOLD = 0.74;
      const CONSERVATIVE_WAIT_LINEUP_LOSS_THRESHOLD = 0.28;
      const SECONDARY_QB_TE_VALUE_EDGE = 0.45;
      const FINALIZED_SCHEMA_VERSION = 'finalized_draft_v1';
      const scoringPresets = data.scoring_presets || {};
      const presetEntries = Object.values(scoringPresets);
      const availablePreset = presetEntries.find((entry) => entry && entry.available);
      const defaultPreset = (data.runtime_controls && data.runtime_controls.active_scoring_preset) || (availablePreset && availablePreset.key) || 'half_ppr';
      const defaultRosterSpots = Object.assign({}, (data.league_settings && data.league_settings.roster_spots) || {});
      const defaultState = {
        version: 2,
        currentPickNumber: (data.current_pick_number || (data.league_settings && data.league_settings.draft_position) || 1),
        draftPosition: (data.league_settings && data.league_settings.draft_position) || 1,
        leagueSize: (data.league_settings && data.league_settings.league_size) || 10,
        scoringPreset: defaultPreset,
        riskTolerance: ((data.league_settings && data.league_settings.risk_tolerance) || 'medium').toLowerCase(),
        benchSlots: (data.league_settings && data.league_settings.bench_slots) || 6,
        rosterSpots: Object.assign({}, defaultRosterSpots),
        takenPlayers: ((data.current_draft_context_defaults && data.current_draft_context_defaults.taken_players) || []).slice(),
        yourPlayers: ((data.current_draft_context_defaults && data.current_draft_context_defaults.your_players) || []).slice(),
        queuePlayers: [],
        history: [],
        redoHistory: [],
        pickLog: [],
        search: '',
        selectedPlayer: data.selected_player || '',
      };

      const state = loadState();
      bindControls();
      render();

      function safeLower(value) {
        return (value || '').toString().trim().toLowerCase();
      }

      function clone(value) {
        return JSON.parse(JSON.stringify(value));
      }

      function escapeHtml(value) {
        return (value ?? '').toString()
          .replace(/&/g, '&amp;')
          .replace(/</g, '&lt;')
          .replace(/>/g, '&gt;')
          .replace(/"/g, '&quot;')
          .replace(/'/g, '&#39;');
      }

      function serializeForHtml(value) {
        return JSON.stringify(value).replace(/</g, '\\u003c');
      }

      function formatSignedNumber(value) {
        const numeric = Number(value || 0);
        return `${numeric > 0 ? '+' : ''}${formatNumber(numeric)}`;
      }

      function explainFreshnessState(status, overrideUsed) {
        const normalized = safeLower(status);
        if (!normalized || normalized === 'fresh') {
          return '';
        }
        if (normalized === 'mixed') {
          return 'Some analysis inputs are current while others are degraded, so trust the board with extra caution.';
        }
        if (normalized === 'stale') {
          return 'The rendered dashboard surface is out of sync with its paired payload or expected source artifacts.';
        }
        if (normalized === 'degraded') {
          return 'The board was generated with degraded inputs, so some recommendation support may be weaker than usual.';
        }
        if (overrideUsed) {
          return 'A freshness override was used to force a degraded analysis window.';
        }
        return `Freshness status is ${status}.`;
      }

      function renderFreshnessNotice(freshness) {
        const warnings = Array.isArray(freshness && freshness.warnings) ? freshness.warnings : [];
        if (warnings.length) {
          return `<div class="notice">${warnings.join(' ')}</div>`;
        }
        const explanation = explainFreshnessState(freshness && freshness.status, freshness && freshness.override_used);
        return explanation ? `<div class="notice">${explanation} Detailed warning text is unavailable in this payload.</div>` : '';
      }

      function downloadTextFile(filename, text, mimeType) {
        const blob = new Blob([text], { type: mimeType });
        const url = window.URL.createObjectURL(blob);
        const anchor = document.createElement('a');
        anchor.href = url;
        anchor.download = filename;
        anchor.style.display = 'none';
        document.body.appendChild(anchor);
        anchor.click();
        anchor.remove();
        window.setTimeout(() => window.URL.revokeObjectURL(url), 0);
      }

      function isLocalFinalizeSupported() {
        return window.location.protocol === 'file:';
      }

      function rosterCapacity() {
        return Object.values(state.rosterSpots || {}).reduce((sum, value) => sum + (Number(value) || 0), 0) + (Number(state.benchSlots) || 0);
      }

      function draftYearLabel() {
        return String((data.generated_at || '').slice(0, 4) || new Date().getFullYear());
      }

      function timestampLabel() {
        return new Date().toISOString().replace(/[-:]/g, '').replace(/\..+$/, '').replace('T', '_');
      }

      function summarizeValueIndicator(row) {
        const draftRank = Number(row.draft_rank || 0);
        const marketRank = Number(row.market_rank || row.adp || 0);
        if (!Number.isFinite(draftRank) || !Number.isFinite(marketRank) || !draftRank || !marketRank) {
          return 'Fair';
        }
        const edge = marketRank - draftRank;
        if (edge >= 12) {
          return 'Value';
        }
        if (edge <= -12) {
          return 'Reach';
        }
        return 'Fair';
      }

      function summarizeRiskStyle(avgFragility, avgUpside) {
        if (avgFragility >= 0.55 && avgUpside >= 0.65) {
          return 'Ceiling-heavy and fragile';
        }
        if (avgFragility >= 0.55) {
          return 'Fragile';
        }
        if (avgUpside >= 0.65) {
          return 'Ceiling-heavy';
        }
        return 'Balanced';
      }

      function loadState() {
        try {
          const parsed = JSON.parse(window.localStorage.getItem(STORAGE_KEY) || 'null');
          if (!parsed || parsed.version !== defaultState.version) {
            return clone(defaultState);
          }
          return {
            ...clone(defaultState),
            ...parsed,
            rosterSpots: { ...clone(defaultState).rosterSpots, ...(parsed.rosterSpots || {}) },
            takenPlayers: Array.isArray(parsed.takenPlayers) ? parsed.takenPlayers : clone(defaultState).takenPlayers,
            yourPlayers: Array.isArray(parsed.yourPlayers) ? parsed.yourPlayers : clone(defaultState).yourPlayers,
            queuePlayers: Array.isArray(parsed.queuePlayers) ? parsed.queuePlayers : [],
            history: Array.isArray(parsed.history) ? parsed.history : [],
            redoHistory: Array.isArray(parsed.redoHistory) ? parsed.redoHistory : [],
            pickLog: Array.isArray(parsed.pickLog) ? parsed.pickLog : [],
          };
        } catch (_) {
          return clone(defaultState);
        }
      }

      function persistState() {
        window.localStorage.setItem(STORAGE_KEY, JSON.stringify(state));
      }

      function captureDraftSnapshot() {
        return {
          currentPickNumber: state.currentPickNumber,
          takenPlayers: state.takenPlayers.slice(),
          yourPlayers: state.yourPlayers.slice(),
          queuePlayers: state.queuePlayers.slice(),
          pickLog: state.pickLog.slice(),
          selectedPlayer: state.selectedPlayer,
        };
      }

      function pushSnapshot(stack, snapshot) {
        stack.push(snapshot);
        if (stack.length > 60) {
          stack.shift();
        }
      }

      function pushHistory() {
        pushSnapshot(state.history, captureDraftSnapshot());
        state.redoHistory = [];
      }

      function restoreDraftSnapshot(snapshot) {
        if (!snapshot) {
          return;
        }
        state.currentPickNumber = snapshot.currentPickNumber;
        state.takenPlayers = snapshot.takenPlayers;
        state.yourPlayers = snapshot.yourPlayers;
        state.queuePlayers = snapshot.queuePlayers;
        state.pickLog = snapshot.pickLog || [];
        state.selectedPlayer = snapshot.selectedPlayer;
      }

      function nextPickNumber(currentPickNumber, draftPosition, leagueSize) {
        const current = Math.max(1, Number(currentPickNumber) || 1);
        const draft = Math.max(1, Number(draftPosition) || 1);
        const size = Math.max(1, Number(leagueSize) || 1);
        const rounds = Math.max(1, Math.ceil(current / size) + 2);
        for (let round = 1; round <= rounds; round += 1) {
          const pick = round % 2 === 1 ? ((round - 1) * size) + draft : (round * size) - draft + 1;
          if (pick > current) {
            return pick;
          }
        }
        return current + size;
      }

      function currentRoundNumber(currentPickNumber, leagueSize) {
        return Math.max(1, Math.ceil(Math.max(1, Number(currentPickNumber) || 1) / Math.max(1, Number(leagueSize) || 1)));
      }

      function specialistWindowOpen() {
        const round = currentRoundNumber(state.currentPickNumber, state.leagueSize);
        const totalRounds = Object.values(state.rosterSpots || {}).reduce((sum, value) => sum + (Number(value) || 0), 0) + (Number(state.benchSlots) || 0);
        return round >= Math.max(1, totalRounds - LATE_SPECIALIST_BUFFER_ROUNDS);
      }

      function availabilityProbability(adp, targetPick, adpStd, uncertaintyScore) {
        const adpValue = Number(adp);
        if (!Number.isFinite(adpValue)) {
          return 0.5;
        }
        let spread = Number(adpStd);
        if (!Number.isFinite(spread) || spread <= 0) {
          spread = 2.5;
        }
        const uncertainty = Number(uncertaintyScore);
        if (Number.isFinite(uncertainty)) {
          spread += 2.0 * uncertainty;
        }
        const z = (adpValue - Number(targetPick)) / Math.max(1, spread);
        return Math.max(0, Math.min(1, 1 / (1 + Math.exp(-z))));
      }

      function formatNumber(value, digits = 2) {
        const num = Number(value);
        return Number.isFinite(num) ? num.toFixed(digits) : '0.00';
      }

      function formatPercent(value) {
        return `${Math.round((Number(value) || 0) * 100)}%`;
      }

      function formatTimestamp(value) {
        if (!value) {
          return 'n/a';
        }
        const dt = new Date(value);
        return Number.isNaN(dt.getTime()) ? String(value) : dt.toLocaleString();
      }

      function getPresetEntry() {
        const requested = scoringPresets[state.scoringPreset];
        if (requested && requested.available) {
          return requested;
        }
        return Object.values(scoringPresets).find((entry) => entry && entry.available) || { decision_table: data.decision_table || [], available: true, key: defaultPreset, label: 'Current preset' };
      }

      function activeRows() {
        const presetEntry = getPresetEntry();
        return Array.isArray(presetEntry.decision_table) && presetEntry.decision_table.length
          ? presetEntry.decision_table.map((row) => ({ ...row }))
          : (Array.isArray(data.decision_table) ? data.decision_table.map((row) => ({ ...row })) : []);
      }

      function rosterCounts(rows) {
        const byPlayer = new Map(rows.map((row) => [safeLower(row.player_name), row]));
        return state.yourPlayers.reduce((counts, playerName) => {
          const row = byPlayer.get(safeLower(playerName));
          if (row) {
            counts[row.position] = (counts[row.position] || 0) + 1;
          }
          return counts;
        }, {});
      }

      function rosterNeed(counts) {
        const need = POSITION_KEYS.reduce((acc, position) => {
          acc[position] = Math.max(0, Number(state.rosterSpots[position] || 0) - Number(counts[position] || 0));
          return acc;
        }, {});
        const flexEligibleExtras = ['RB', 'WR', 'TE'].reduce((sum, position) => {
          const baseSlots = Number(state.rosterSpots[position] || 0);
          const filled = Number(counts[position] || 0);
          return sum + Math.max(0, filled - baseSlots);
        }, 0);
        need.FLEX = Math.max(0, Number(state.rosterSpots.FLEX || 0) - flexEligibleExtras);
        return need;
      }

      function offensiveNeedByPosition(need, position) {
        const baseNeed = Number(need[position] || 0);
        if (position === 'RB' || position === 'WR' || position === 'TE') {
          return baseNeed + (Number(need.FLEX || 0) * Number(FLEX_WEIGHTS[position] || 0));
        }
        return baseNeed;
      }

      function isInspectableStatus(status) {
        return status === 'available' || status === 'queued';
      }

      function startersByPosition() {
        return POSITION_KEYS.reduce((acc, position) => {
          acc[position] = Number(state.rosterSpots[position] || 0) * Math.max(1, Number(state.leagueSize || 1));
          return acc;
        }, {});
      }

      function replacementSlots() {
        const starters = startersByPosition();
        const flexSlots = Number(state.rosterSpots.FLEX || 0) * Math.max(1, Number(state.leagueSize || 1));
        const replacement = { ...starters };
        replacement.RB = (replacement.RB || 0) + Math.round(flexSlots * (FLEX_WEIGHTS.RB || 0));
        replacement.WR = (replacement.WR || 0) + Math.round(flexSlots * (FLEX_WEIGHTS.WR || 0));
        replacement.TE = (replacement.TE || 0) + Math.round(flexSlots * (FLEX_WEIGHTS.TE || 0));
        return replacement;
      }

      function positionBaseline(rows, position, slotCount, fallback) {
        const values = rows
          .filter((row) => row.position === position)
          .map((row) => Number(row.proj_points_mean))
          .filter((value) => Number.isFinite(value))
          .sort((a, b) => b - a);
        if (!values.length) {
          return fallback;
        }
        const index = Math.max(0, Math.min(values.length - 1, Math.round(slotCount || 1) - 1));
        return values[index];
      }

      function rankPercentiles(values) {
        const finite = values.map((value, index) => ({ value: Number(value), index })).filter((entry) => Number.isFinite(entry.value));
        if (!finite.length) {
          return values.map(() => 0.5);
        }
        const sorted = [...finite].sort((a, b) => a.value - b.value);
        const result = values.map(() => 0.5);
        sorted.forEach((entry, idx) => {
          result[entry.index] = sorted.length === 1 ? 0.5 : idx / (sorted.length - 1);
        });
        return result;
      }

      function zScores(values) {
        const numeric = values.map((value) => Number(value));
        const finite = numeric.filter((value) => Number.isFinite(value));
        if (!finite.length) {
          return numeric.map(() => 0);
        }
        const mean = finite.reduce((sum, value) => sum + value, 0) / finite.length;
        const variance = finite.reduce((sum, value) => sum + ((value - mean) ** 2), 0) / finite.length;
        const std = Math.sqrt(variance);
        if (!Number.isFinite(std) || std === 0) {
          return numeric.map(() => 0);
        }
        return numeric.map((value) => Number.isFinite(value) ? (value - mean) / std : 0);
      }

      function assignTiers(rows) {
        const total = rows.length || 1;
        rows.forEach((row, index) => {
          const bucket = Math.min(5, Math.floor((index / total) * 5) + 1);
          row.draft_tier = `Tier ${bucket}`;
        });
      }

      function getStatus(row, takenSet, yoursSet, queueSet) {
        const key = safeLower(row.player_name);
        if (yoursSet.has(key)) {
          return 'mine';
        }
        if (takenSet.has(key)) {
          return 'taken';
        }
        if (queueSet.has(key)) {
          return 'queued';
        }
        return 'available';
      }

      function buildBoardState() {
        const rows = activeRows().filter((row) => {
          const name = (row.player_name || '').toString().trim();
          return name && name.toUpperCase() !== 'UNKNOWN' && FANTASY_DRAFT_POSITIONS.includes(row.position);
        });
        const takenSet = new Set((state.takenPlayers || []).map(safeLower));
        const yoursSet = new Set((state.yourPlayers || []).map(safeLower));
        const queueSet = new Set((state.queuePlayers || []).map(safeLower));
        const nextPick = nextPickNumber(state.currentPickNumber, state.draftPosition, state.leagueSize);
        const roundNumber = currentRoundNumber(state.currentPickNumber, state.leagueSize);
        const lateSpecialistsOk = specialistWindowOpen();
        const counts = rosterCounts(rows);
        const need = rosterNeed(counts);
        const effectiveNeed = OFFENSIVE_POSITIONS.reduce((acc, position) => {
          acc[position] = offensiveNeedByPosition(need, position);
          return acc;
        }, {});
        const totalRounds = Object.values(state.rosterSpots || {}).reduce((sum, value) => sum + (Number(value) || 0), 0) + (Number(state.benchSlots) || 0);
        const remainingRounds = Math.max(1, totalRounds - roundNumber + 1);
        const openSpecialists = ['DST', 'K'].filter((position) => Number(need[position] || 0) > 0);
        const starterCounts = startersByPosition();
        const replacementCounts = replacementSlots();
        const overallMeanProjection = rows
          .map((row) => Number(row.proj_points_mean))
          .filter((value) => Number.isFinite(value))
          .reduce((sum, value, _, values) => sum + value / values.length, 0) || 0;

        const baselines = {};
        ['QB', 'RB', 'WR', 'TE', 'DST', 'K'].forEach((position) => {
          baselines[position] = {
            starter: positionBaseline(rows, position, starterCounts[position] || 1, overallMeanProjection),
            replacement: positionBaseline(rows, position, replacementCounts[position] || 1, overallMeanProjection),
            scarcity: 1 / Math.max(1, rows.filter((row) => row.position === position).length),
          };
        });

        const totalNeed = OFFENSIVE_POSITIONS.reduce((sum, position) => sum + (effectiveNeed[position] || 0), 0) || 1;
        const openOffensiveSlots = OFFENSIVE_POSITIONS.reduce((sum, position) => sum + (effectiveNeed[position] || 0), 0);
        const projectionValues = [];
        const starterValues = [];
        const replacementValues = [];
        const posteriorConfidenceValues = [];
        const fragilityValues = [];
        const upsideValues = [];
        const starterNeedValues = [];
        const marketGapValues = [];
        const availableCounts = {};

        rows.forEach((row) => {
          const baseline = baselines[row.position] || { starter: overallMeanProjection, replacement: overallMeanProjection, scarcity: 0 };
          row.availability_at_pick = availabilityProbability(row.adp, nextPick, row.adp_std, row.uncertainty_score);
          row.availability_to_next_pick = row.availability_at_pick;
          row.starter_baseline = baseline.starter;
          row.replacement_baseline = baseline.replacement;
          row.starter_delta = Number(row.proj_points_mean || 0) - baseline.starter;
          row.replacement_delta = Number(row.proj_points_mean || 0) - baseline.replacement;
          row.simple_vor_proxy = row.replacement_delta;
          row.position_scarcity = baseline.scarcity;
          row.starter_need = OFFENSIVE_POSITIONS.includes(row.position) ? ((effectiveNeed[row.position] || 0) / totalNeed) : 0;
          row.status = getStatus(row, takenSet, yoursSet, queueSet);
          if (row.status === 'available' || row.status === 'queued') {
            availableCounts[row.position] = (availableCounts[row.position] || 0) + 1;
          }
          projectionValues.push(Number(row.proj_points_mean || 0));
          starterValues.push(Number(row.starter_delta || 0));
          replacementValues.push(Number(row.replacement_delta || 0));
          posteriorConfidenceValues.push(Number(row.posterior_prob_beats_replacement || 0));
          fragilityValues.push(Number(row.fragility_score || 0));
          starterNeedValues.push(Number(row.starter_need || 0));
          marketGapValues.push(Number(row.market_gap || 0));
          upsideValues.push((Number(row.proj_points_ceiling || 0) - Number(row.proj_points_mean || 0)));
        });

        const upsidePercentiles = rankPercentiles(upsideValues);
        const availabilityPercentiles = rankPercentiles(rows.map((row) => Number(row.availability_to_next_pick || 0)));
        const projectionPercentiles = rankPercentiles(projectionValues);
        rows.forEach((row, index) => {
          row.upside_score = Math.max(0, Math.min(1, upsidePercentiles[index] + (0.35 * availabilityPercentiles[index]) + (0.15 * projectionPercentiles[index]) + (0.45 * Number(row.posterior_prob_beats_replacement || 0))));
        });
        const finalUpsideValues = rows.map((row) => Number(row.upside_score || 0));
        const riskMultiplier = ({ low: 0.80, medium: 1.00, high: 1.18 })[(state.riskTolerance || 'medium').toLowerCase()] || 1.0;
        const componentZ = {
          proj_points_mean: zScores(projectionValues),
          starter_delta: zScores(starterValues),
          replacement_delta: zScores(replacementValues),
          posterior_prob_beats_replacement: zScores(posteriorConfidenceValues),
          starter_need: zScores(starterNeedValues),
          fragility_score: zScores(fragilityValues),
          market_gap: zScores(marketGapValues),
        };

        rows.forEach((row, index) => {
          row.component_terms = {
            proj_points_mean: 0.40 * componentZ.proj_points_mean[index],
            starter_delta: 0.24 * componentZ.starter_delta[index],
            replacement_delta: 0.18 * componentZ.replacement_delta[index],
            posterior_prob_beats_replacement: 0.10 * componentZ.posterior_prob_beats_replacement[index],
            starter_need: 0.03 * componentZ.starter_need[index],
            fragility_score: -(0.06 * riskMultiplier) * componentZ.fragility_score[index],
            market_gap: 0.06 * componentZ.market_gap[index],
          };
          row.board_value_score = Object.values(row.component_terms).reduce((sum, value) => sum + value, 0);
          row.draft_score = row.board_value_score;
        });

        rows.sort((a, b) => (Number(b.draft_score) - Number(a.draft_score)) || (Number(b.proj_points_mean) - Number(a.proj_points_mean)));
        assignTiers(rows);
        rows.forEach((row, index) => {
          row.draft_rank = index + 1;
        });
        const simpleVorSorted = [...rows].sort((a, b) => (Number(b.simple_vor_proxy) - Number(a.simple_vor_proxy)) || (Number(b.proj_points_mean) - Number(a.proj_points_mean)));
        simpleVorSorted.forEach((row, index) => {
          row.simple_vor_rank = index + 1;
        });

        const availableRows = rows.filter((row) => row.status === 'available' || row.status === 'queued');
        const bestOffensiveValue = availableRows.filter((row) => OFFENSIVE_POSITIONS.includes(row.position)).reduce((best, row) => Math.max(best, Number(row.draft_score) || 0), Number.NEGATIVE_INFINITY);
        const lineupGainValues = [];
        availableRows.forEach((row) => {
          const posNeed = effectiveNeed[row.position] || 0;
          const demand = OFFENSIVE_POSITIONS.includes(row.position)
            ? Math.max(1, Number(state.leagueSize || 1) * Math.max(1, posNeed))
            : Math.max(1, Number(state.leagueSize || 1));
          row.position_run_risk = Math.max(0, Math.min(1, 1 - ((availableCounts[row.position] || 0) / demand)));
          row.starter_slot_urgency = OFFENSIVE_POSITIONS.includes(row.position) ? (posNeed / totalNeed) : 0;
          row.specialist_need_bonus = ((row.position === 'DST' || row.position === 'K') && lateSpecialistsOk && Number(need[row.position] || 0) > 0) ? 2.5 : 0;
          if ((row.position === 'DST' || row.position === 'K') && openSpecialists.includes(row.position) && remainingRounds <= (openSpecialists.length + 1)) {
            row.specialist_need_bonus += 3.0;
          }
          row.specialist_urgency = ((row.position === 'DST' || row.position === 'K') && lateSpecialistsOk && Number(need[row.position] || 0) > 0)
            ? ((1.25 * (1 - Number(row.availability_to_next_pick || 0))) + (0.75 * Math.max(0, Math.min(1, 1 / Math.max(1, availableCounts[row.position] || 1)))))
            : 0;
          row.roster_fit_score = row.starter_slot_urgency;
          row.lineup_gain_now = Math.max(
            0,
            OFFENSIVE_POSITIONS.includes(row.position) ? Number(row.starter_delta || 0) : (Number(row.replacement_delta || 0) * 0.35),
            OFFENSIVE_POSITIONS.includes(row.position) ? (Number(row.replacement_delta || 0) * 0.70) : (Number(row.replacement_delta || 0) * 0.20),
          );
          lineupGainValues.push(Number(row.lineup_gain_now || 0));
          row.policy_eligible = true;
          row.policy_eligibility_reason = 'eligible';
          if ((row.position === 'DST' || row.position === 'K') && !lateSpecialistsOk) {
            row.policy_eligible = false;
            row.policy_eligibility_reason = 'late_round_only';
          } else if ((row.position === 'QB' || row.position === 'TE') && openOffensiveSlots > 0 && Number(counts[row.position] || 0) >= Number(state.rosterSpots[row.position] || 0) && Number(row.draft_score || 0) < (bestOffensiveValue + SECONDARY_QB_TE_VALUE_EDGE)) {
            row.policy_eligible = false;
            row.policy_eligibility_reason = 'starter_priority';
          }
        });
        const lineupGainPercentiles = rankPercentiles(lineupGainValues);
        availableRows.forEach((row, index) => {
          row.expected_regret = ((0.55 * lineupGainPercentiles[index]) + (0.25 * Number(row.starter_slot_urgency || 0)) + (0.20 * Number(row.position_run_risk || 0))) * (1 - Number(row.availability_to_next_pick || 0));
          if ((row.position === 'DST' || row.position === 'K') && lateSpecialistsOk) {
            row.wait_signal = 'late_round_stash_ok';
          } else if (Number(row.availability_to_next_pick || 0) < CONSERVATIVE_WAIT_SURVIVAL_THRESHOLD) {
            row.wait_signal = 'low_survival';
          } else if (Number(row.expected_regret || 0) > CONSERVATIVE_WAIT_LINEUP_LOSS_THRESHOLD) {
            row.wait_signal = 'too_much_lineup_loss';
          } else {
            row.wait_signal = 'safe_to_wait';
          }
          const riskBias = ({ low: -0.03, medium: 0.0, high: 0.03 })[(state.riskTolerance || 'medium').toLowerCase()] || 0;
          row.current_pick_utility = Number(row.draft_score || 0) + Number(row.specialist_need_bonus || 0) + Number(row.specialist_urgency || 0) + (0.32 * Number(row.starter_slot_urgency || 0)) + (0.22 * lineupGainPercentiles[index]) + (0.08 * Number(row.posterior_prob_beats_replacement || 0)) + (0.06 * Number(row.position_run_risk || 0)) + (riskBias * Number(row.upside_score || 0));
          if (!row.policy_eligible) {
            row.current_pick_utility -= 2.5;
          }
          row.wait_utility = (Number(row.draft_score || 0) * Number(row.availability_to_next_pick || 0)) + (0.06 * Number(row.upside_score || 0)) - (0.85 * Number(row.expected_regret || 0));
          row.rank_gap_vs_vor = Number(row.simple_vor_rank || 0) - Number(row.draft_rank || 0);
          row.rationale_live = buildPlayerSummary(row);
        });

        const recommendedNow = [...availableRows].filter((row) => row.policy_eligible).sort((a, b) => (Number(b.current_pick_utility) - Number(a.current_pick_utility)) || (Number(b.draft_score) - Number(a.draft_score)));
        const recommendedWait = [...availableRows].filter((row) => row.wait_signal === 'safe_to_wait' || row.wait_signal === 'late_round_stash_ok').sort((a, b) => (Number(b.wait_utility) - Number(a.wait_utility)) || (Number(b.availability_to_next_pick) - Number(a.availability_to_next_pick)));
        const pickNow = recommendedNow.slice(0, 1);
        const fallbacks = (recommendedNow.length ? recommendedNow : [...availableRows].sort((a, b) => (Number(b.current_pick_utility) - Number(a.current_pick_utility)) || (Number(b.draft_score) - Number(a.draft_score)))).slice(1, 6);
        const canWait = (recommendedWait.length ? recommendedWait : [...availableRows].sort((a, b) => (Number(b.availability_to_next_pick) - Number(a.availability_to_next_pick)) || (Number(b.draft_score) - Number(a.draft_score)))).filter((row) => !pickNow.some((candidate) => safeLower(candidate.player_name) === safeLower(row.player_name))).slice(0, 5);

        const selectedKey = safeLower(state.selectedPlayer);
        const selectedCandidate = rows.find((row) => safeLower(row.player_name) === selectedKey) || null;
        const selectedRow = (
          selectedCandidate && isInspectableStatus(selectedCandidate.status)
        ) ? selectedCandidate : (pickNow[0] || availableRows[0] || rows[0] || null);
        if (selectedRow) {
          state.selectedPlayer = selectedRow.player_name;
        } else {
          state.selectedPlayer = '';
        }

        return {
          rows,
          availableRows,
          nextPick,
          rosterCounts: counts,
          rosterNeed: need,
          pickNow,
          fallbacks,
          canWait,
          selectedRow,
          presetEntry: getPresetEntry(),
        };
      }

      function buildPlayerSummary(row) {
        if (!row) {
          return 'No player selected.';
        }
        const reasons = [];
        if (Number(row.starter_delta) > 0) {
          reasons.push(`adds ${formatNumber(row.starter_delta)} points over a typical ${row.position} starter`);
        }
        if (Number(row.rank_gap_vs_vor) < 0) {
          reasons.push(`the contextual model likes this profile more than the simple VOR baseline`);
        } else if (Number(row.rank_gap_vs_vor) > 0) {
          reasons.push(`simple VOR likes this player a bit more than the contextual score`);
        }
        if (Number(row.availability_to_next_pick) < 0.35) {
          reasons.push(`is unlikely to reach your next pick`);
        } else if (Number(row.availability_to_next_pick) > 0.70) {
          reasons.push(`has a realistic chance to survive to your next turn`);
        }
        if (Number(row.fragility_score) > 0.60) {
          reasons.push(`comes with elevated fragility risk`);
        } else if (Number(row.upside_score) > 0.70) {
          reasons.push(`brings strong upside if you want ceiling`);
        }
        return `${row.player_name} is recommended because ${reasons.slice(0, 3).join(', ')}.`.replace(' because .', '.');
      }

      function filterRows(rows) {
        const query = safeLower(state.search);
        return rows.filter((row) => {
          if (query) {
            const haystack = [row.player_name, row.position, row.team].map(safeLower).join(' ');
            if (!haystack.includes(query)) {
              return false;
            }
          }
          return true;
        });
      }

      function percentileRows(rows, key, invert = false) {
        const percentiles = rankPercentiles(rows.map((row) => Number(row[key] || 0)));
        return rows.reduce((acc, row, index) => {
          acc[safeLower(row.player_name)] = invert ? (1 - percentiles[index]) : percentiles[index];
          return acc;
        }, {});
      }

      function render() {
        const boardState = buildBoardState();
        persistState();
        renderTopbar(boardState);
        renderRecommendations(boardState);
        renderQueue(boardState);
        renderBoard(boardState);
        renderControls(boardState);
        renderInspector(boardState);
        renderModelNotes(boardState);
      }

      function renderTopbar(boardState) {
        const presetEntry = boardState.presetEntry;
        const finalizeButton = document.getElementById('finalize-button');
        const finalizeNote = document.getElementById('finalize-note');
        document.getElementById('storage-pill').textContent = 'Saved locally';
        document.getElementById('status-pills').innerHTML = [
          ['Current pick', state.currentPickNumber],
          ['Next pick', boardState.nextPick],
          ['League', `${state.leagueSize}-team`],
          ['Draft slot', state.draftPosition],
          ['Preset', presetEntry.label || presetEntry.key || 'Current'],
          ['Taken', state.takenPlayers.length],
          ['Yours', state.yourPlayers.length],
        ].map(([label, value]) => `<span class="pill">${label}: ${value}</span>`).join('');
        document.getElementById('undo-button').disabled = !state.history.length;
        document.getElementById('redo-button').disabled = !state.redoHistory.length;
        finalizeButton.hidden = !isLocalFinalizeSupported();
        finalizeNote.textContent = isLocalFinalizeSupported()
          ? 'Finalize downloads a locked HTML snapshot, JSON export, and post-draft summary from this local dashboard.'
          : 'Finalize downloads are only supported from the local generated dashboard opened as a file.';
      }

      function renderRecommendations(boardState) {
        const primary = boardState.pickNow[0];
        document.getElementById('primary-card').innerHTML = primary ? `
          <div class="pill-row">
            <span class="pill">${primary.position}</span>
            <span class="pill">Board value ${formatNumber(primary.draft_score)}</span>
            <span class="pill">Simple VOR rank ${primary.simple_vor_rank}</span>
          </div>
          <div class="hero-name">${primary.player_name}</div>
          <div class="summary-box">${buildPlayerSummary(primary)}</div>
          <div class="metric-grid">
            <div class="metric"><span class="label">Availability to next pick</span><span class="value">${formatPercent(primary.availability_to_next_pick)}</span></div>
            <div class="metric"><span class="label">Expected regret</span><span class="value">${formatNumber(primary.expected_regret)}</span></div>
            <div class="metric"><span class="label">Position run risk</span><span class="value">${formatNumber(primary.position_run_risk)}</span></div>
            <div class="metric"><span class="label">Roster fit</span><span class="value">${formatNumber(primary.roster_fit_score)}</span></div>
          </div>
        ` : '<div class="empty">No available recommendation.</div>';
        renderLane('fallback-list', boardState.fallbacks, 'No fallback options right now.');
        renderLane('wait-list', boardState.canWait, 'No wait candidates right now.');
      }

      function renderLane(elementId, rows, emptyMessage) {
        const container = document.getElementById(elementId);
        if (!rows.length) {
          container.innerHTML = `<div class="empty">${emptyMessage}</div>`;
          return;
        }
        container.innerHTML = rows.map((row) => `
          <div class="lane-item">
            <div class="lane-item-header">
              <div class="item-title">${row.player_name} <span class="tiny">• ${row.position}</span></div>
              <span class="status-badge ${row.status}">${row.status}</span>
            </div>
            <div class="item-meta">Board value ${formatNumber(row.draft_score)} • Survival ${formatPercent(row.availability_to_next_pick)} • Regret ${formatNumber(row.expected_regret)}</div>
            <div class="tiny">${buildPlayerSummary(row)}</div>
          </div>
        `).join('');
      }

      function renderQueue(boardState) {
        const queueContainer = document.getElementById('queue-list');
        const rosterContainer = document.getElementById('my-roster');
        const rosterNeedGrid = document.getElementById('roster-need-grid');
        const queueRows = boardState.rows.filter((row) => safeLowerArray(state.queuePlayers).includes(safeLower(row.player_name)));
        queueContainer.innerHTML = queueRows.length ? queueRows.map((row) => `
          <div class="mini-item">
            <div class="mini-item-header">
              <div class="item-title">${row.player_name} <span class="tiny">• ${row.position}</span></div>
              <span class="pill">${formatPercent(row.availability_to_next_pick)} survival</span>
            </div>
            <div class="item-meta">${buildPlayerSummary(row)}</div>
          </div>
        `).join('') : '<div class="empty">Queue players with the “Queue” action in the board.</div>';
        const yourRows = boardState.rows.filter((row) => safeLowerArray(state.yourPlayers).includes(safeLower(row.player_name)));
        rosterContainer.innerHTML = yourRows.length ? yourRows.map((row) => `<span class="pill good">${row.player_name} • ${row.position}</span>`).join('') : '<span class="empty">Your roster will appear here as you mark picks.</span>';
        rosterNeedGrid.innerHTML = POSITION_KEYS.map((position) => `
          <div class="metric">
            <span class="label">${position} need</span>
            <span class="value">${boardState.rosterNeed[position] || 0}</span>
          </div>
        `).join('');
      }

      function renderBoard(boardState) {
        const filteredRows = filterRows(boardState.rows);
        const tableBody = document.getElementById('board-table');
        document.getElementById('board-count-pill').textContent = `${filteredRows.length} shown`;
        document.getElementById('board-summary-text').textContent = `Showing ${filteredRows.length} of ${boardState.rows.length} players`;
        document.getElementById('preset-summary-text').textContent = `Preset: ${boardState.presetEntry.label || boardState.presetEntry.key || 'Current'} • Risk: ${(state.riskTolerance || 'medium').toUpperCase()}`;
        if (!filteredRows.length) {
          tableBody.innerHTML = '<tr><td colspan="7" class="empty">No players match the current filters.</td></tr>';
          return;
        }
        tableBody.innerHTML = filteredRows.map((row) => {
          const rowClasses = [
            safeLower(row.player_name) === safeLower(state.selectedPlayer) ? 'is-selected' : '',
            row.status === 'taken' ? 'is-taken' : '',
            row.status === 'mine' ? 'is-mine' : '',
            row.status === 'queued' ? 'is-queued' : '',
          ].filter(Boolean).join(' ');
          return `
            <tr class="${rowClasses}" data-player-row="${row.player_name}">
              <td>
                <div class="item-title">${row.player_name}</div>
                <div class="tiny">${row.position}${row.team ? ` • ${row.team}` : ''} • ${row.draft_tier || ''}</div>
              </td>
              <td><span class="status-badge ${row.status}">${row.status}</span></td>
              <td>${formatNumber(row.draft_score)}</td>
              <td>${formatNumber(row.simple_vor_proxy)} <span class="tiny">(rank ${row.simple_vor_rank || '-'})</span></td>
              <td>${formatPercent(row.availability_to_next_pick)}</td>
              <td class="tiny">${buildPlayerSummary(row)}</td>
              <td>
                <div class="action-group">
                  <button type="button" data-action="queue" data-player="${row.player_name}">${row.status === 'queued' ? 'Unqueue' : 'Queue'}</button>
                  <button type="button" data-action="taken" data-player="${row.player_name}">${row.status === 'taken' ? 'Clear' : 'Taken'}</button>
                  <button type="button" data-action="mine" data-player="${row.player_name}">${row.status === 'mine' ? 'Unmark' : 'Mine'}</button>
                </div>
              </td>
            </tr>
          `;
        }).join('');
        tableBody.querySelectorAll('[data-player-row]').forEach((rowElement) => {
          rowElement.addEventListener('click', (event) => {
            if (event.target.closest('button')) {
              return;
            }
            state.selectedPlayer = rowElement.getAttribute('data-player-row');
            render();
          });
        });
        tableBody.querySelectorAll('button[data-action]').forEach((button) => {
          button.addEventListener('click', (event) => {
            event.stopPropagation();
            applyAction(button.getAttribute('data-action'), button.getAttribute('data-player'));
          });
        });
      }

      function renderControls(boardState) {
        const presetSelect = document.getElementById('scoring-preset');
        const currentPreset = getPresetEntry();
        presetSelect.innerHTML = Object.entries(scoringPresets).map(([key, entry]) => `
          <option value="${key}" ${key === currentPreset.key ? 'selected' : ''} ${entry.available ? '' : 'disabled'}>
            ${entry.label}${entry.available ? '' : ' (regen needed)'}
          </option>
        `).join('');
        document.getElementById('risk-tolerance').value = state.riskTolerance || 'medium';
        document.getElementById('league-size').value = state.leagueSize;
        document.getElementById('draft-position').value = state.draftPosition;
        document.getElementById('current-pick-number').value = state.currentPickNumber;
        document.getElementById('bench-slots').value = state.benchSlots;
        document.getElementById('roster-qb').value = state.rosterSpots.QB || 0;
        document.getElementById('roster-rb').value = state.rosterSpots.RB || 0;
        document.getElementById('roster-wr').value = state.rosterSpots.WR || 0;
        document.getElementById('roster-te').value = state.rosterSpots.TE || 0;
        document.getElementById('roster-flex').value = state.rosterSpots.FLEX || 0;
        document.getElementById('preset-notice').textContent = currentPreset.available
          ? `Using ${currentPreset.label}. This board can recompute live draft context in-browser.`
          : (currentPreset.reason_unavailable || 'This preset is unavailable for the exported payload.');
      }

      function renderInspector(boardState) {
        const container = document.getElementById('player-inspector');
        const row = boardState.selectedRow;
        if (!row) {
          container.innerHTML = '<div class="empty">Select a player from the board.</div>';
          return;
        }
        const percentileMaps = {
          projection: percentileRows(boardState.rows, 'proj_points_mean'),
          upside: percentileRows(boardState.rows, 'upside_score'),
          fragility: percentileRows(boardState.rows, 'fragility_score', true),
          survival: percentileRows(boardState.rows, 'availability_to_next_pick'),
          market: percentileRows(boardState.rows, 'market_gap'),
        };
        const contributions = Object.entries(row.component_terms || {}).map(([key, value]) => ({
          key,
          value: Number(value || 0),
          magnitude: Math.abs(Number(value || 0)),
        }));
        const totalContribution = contributions.reduce((sum, item) => sum + item.magnitude, 0) || 1;
        const componentLabels = {
          starter_delta: 'Starter edge',
          replacement_delta: 'Simple VOR',
          proj_points_mean: 'Projection',
          availability_to_next_pick: 'Draft timing',
          upside_score: 'Upside',
          starter_need: 'Roster need',
          position_scarcity: 'Scarcity',
          fragility_score: 'Fragility',
          market_gap: 'Market gap',
        };
        container.innerHTML = `
          <div class="inspector-title">
            <div class="pill-row">
              <span class="pill">${row.position}</span>
              <span class="pill">${row.status}</span>
              <span class="pill">Draft rank ${row.draft_rank}</span>
              <span class="pill">VOR rank ${row.simple_vor_rank}</span>
            </div>
            <div class="hero-name" style="font-size: 24px;">${row.player_name}</div>
            <div class="summary-box">${buildPlayerSummary(row)}</div>
          </div>
          <div class="metric-grid">
            <div class="metric"><span class="label">Board value score</span><span class="value">${formatNumber(row.draft_score)}</span></div>
            <div class="metric"><span class="label">Simple VOR proxy</span><span class="value">${formatNumber(row.simple_vor_proxy)}</span></div>
            <div class="metric"><span class="label">Availability to next pick</span><span class="value">${formatPercent(row.availability_to_next_pick)}</span></div>
            <div class="metric"><span class="label">Expected regret</span><span class="value">${formatNumber(row.expected_regret)}</span></div>
            <div class="metric"><span class="label">Upside score</span><span class="value">${formatPercent(row.upside_score)}</span></div>
            <div class="metric"><span class="label">Fragility score</span><span class="value">${formatPercent(row.fragility_score)}</span></div>
          </div>
          <details open>
            <summary>Why this player / why not wait?</summary>
            <div class="details-body">
              <div class="bar-stack">
                ${contributions.sort((a, b) => b.magnitude - a.magnitude).map((item) => `
                  <div class="bar-row">
                    <div class="bar-head">
                      <span>${componentLabels[item.key] || item.key}</span>
                      <span>${formatNumber(item.value)}</span>
                    </div>
                    <div class="bar-track">
                      <div class="bar-fill ${item.key === 'fragility_score' ? 'bad' : (item.value >= 0 ? 'good' : 'warn')}" style="width: ${Math.max(6, Math.round((item.magnitude / totalContribution) * 100))}%"></div>
                    </div>
                  </div>
                `).join('')}
              </div>
            </div>
          </details>
          <details>
            <summary>Player fingerprint</summary>
            <div class="details-body">
              <div class="bar-stack">
                ${[
                  ['Projection', percentileMaps.projection[safeLower(row.player_name)] || 0],
                  ['Upside', percentileMaps.upside[safeLower(row.player_name)] || 0],
                  ['Safety', percentileMaps.fragility[safeLower(row.player_name)] || 0],
                  ['Next-pick survival', percentileMaps.survival[safeLower(row.player_name)] || 0],
                  ['Model vs market gap', percentileMaps.market[safeLower(row.player_name)] || 0],
                ].map(([label, value]) => `
                  <div class="bar-row">
                    <div class="bar-head"><span>${label}</span><span>${formatPercent(value)}</span></div>
                    <div class="bar-track"><div class="bar-fill" style="width: ${Math.round(value * 100)}%"></div></div>
                  </div>
                `).join('')}
              </div>
            </div>
          </details>
        `;
      }

      function renderModelNotes(boardState) {
        const modelOverview = data.model_overview || {};
        const decisionEvidence = data.decision_evidence || {};
        const analysisProvenance = data.analysis_provenance || {};
        const publishProvenance = data.publish_provenance || {};
        const analysisFreshness = analysisProvenance.overall_freshness || decisionEvidence.freshness || {};
        const freshnessRows = Object.values(analysisProvenance.sources || {});
        const fallbackFreshnessRows = Array.isArray(data.source_freshness) ? data.source_freshness : [];
        const renderedFreshnessRows = freshnessRows.length ? freshnessRows : fallbackFreshnessRows;
        const backtestRows = Array.isArray(decisionEvidence.strategy_summary)
          ? decisionEvidence.strategy_summary
          : (data.backtest && data.backtest.overall && Array.isArray(data.backtest.overall.by_strategy)
            ? data.backtest.overall.by_strategy
            : []);
        document.getElementById('model-overview').innerHTML = `
          <div class="summary-box">${modelOverview.headline || 'The model mixes projection, VOR-style value, timing, and fragility.'}</div>
          ${(modelOverview.plain_english || []).map((line) => `<div class="tiny">• ${line}</div>`).join('')}
          ${(modelOverview.limitations || []).length ? `<div class="notice">${(modelOverview.limitations || []).join(' ')}</div>` : ''}
        `;
        document.getElementById('decision-evidence').innerHTML = decisionEvidence.available ? `
          <div class="summary-box">${decisionEvidence.headline || 'Decision evidence is available for this board.'}</div>
          <div class="metric-grid">
            <div class="metric"><span class="label">Evidence status</span><span class="value">${decisionEvidence.status || 'available'}</span></div>
            <div class="metric"><span class="label">Freshness</span><span class="value">${(decisionEvidence.freshness && decisionEvidence.freshness.status) || 'unknown'}</span></div>
            <div class="metric"><span class="label">Winner</span><span class="value">${decisionEvidence.winner || 'n/a'}</span></div>
            <div class="metric"><span class="label">Mean lineup delta</span><span class="value">${formatNumber(decisionEvidence.delta_mean_lineup_points)}</span></div>
          </div>
          ${renderFreshnessNotice(decisionEvidence.freshness || {})}
          <div class="board-table-wrap" style="max-height: 220px;">
            <table>
              <thead>
                <tr><th>Strategy</th><th>Mean lineup points</th><th>Seasons</th></tr>
              </thead>
              <tbody>
                ${backtestRows.map((row) => `
                  <tr>
                    <td>${row.strategy}</td>
                    <td>${formatNumber(row.mean_lineup_points)}</td>
                    <td>${row.season_count ?? 'n/a'}</td>
                  </tr>
                `).join('') || '<tr><td colspan="3" class="empty">No strategy rows available.</td></tr>'}
              </tbody>
            </table>
          </div>
          <div class="board-table-wrap" style="max-height: 220px;">
            <table>
              <thead>
                <tr><th>Year</th><th>Board value score</th><th>Simple VOR proxy</th><th>Delta</th></tr>
              </thead>
              <tbody>
                ${(decisionEvidence.season_rows || []).map((row) => `
                  <tr>
                    <td>${row.holdout_year}</td>
                    <td>${formatNumber(row.draft_score_lineup_points)}</td>
                    <td>${formatNumber(row.historical_vor_proxy_lineup_points)}</td>
                    <td>${formatNumber(row.delta_lineup_points)}</td>
                  </tr>
                `).join('') || '<tr><td colspan="4" class="empty">No season-level rows available.</td></tr>'}
              </tbody>
            </table>
          </div>
          <div class="board-table-wrap" style="max-height: 220px;">
            <table>
              <thead>
                <tr><th>Player</th><th>Pos</th><th>Draft rank</th><th>VOR rank</th><th>Gap</th></tr>
              </thead>
              <tbody>
                ${(decisionEvidence.top_disagreements || []).map((row) => `
                  <tr>
                    <td>${row.player_name}</td>
                    <td>${row.position}</td>
                    <td>${row.draft_rank}</td>
                    <td>${row.simple_vor_rank}</td>
                    <td>${row.rank_gap_vs_vor}</td>
                  </tr>
                `).join('') || '<tr><td colspan="5" class="empty">No disagreement rows available.</td></tr>'}
              </tbody>
            </table>
          </div>
          ${(decisionEvidence.limitations || []).length ? `<div class="notice">${decisionEvidence.limitations.join(' ')}</div>` : ''}
        ` : `
          <div class="notice">${decisionEvidence.reason_unavailable || decisionEvidence.headline || 'No decision evidence is available in this payload.'}</div>
          <div class="board-table-wrap" style="max-height: 220px;">
            <table>
              <thead>
                <tr><th>Player</th><th>Pos</th><th>Draft rank</th><th>VOR rank</th><th>Gap</th></tr>
              </thead>
              <tbody>
                ${(decisionEvidence.top_disagreements || []).map((row) => `
                  <tr>
                    <td>${row.player_name}</td>
                    <td>${row.position}</td>
                    <td>${row.draft_rank}</td>
                    <td>${row.simple_vor_rank}</td>
                    <td>${row.rank_gap_vs_vor}</td>
                  </tr>
                `).join('') || '<tr><td colspan="5" class="empty">No disagreement rows available.</td></tr>'}
              </tbody>
            </table>
          </div>
        `;
        document.getElementById('metric-glossary').innerHTML = `<div class="glossary-list">${
          Object.entries(data.metric_glossary || {}).map(([key, item]) => `
            <div class="glossary-item">
              <strong>${item.label || key}</strong>
              <div class="tiny">${item.summary || ''}</div>
              <div class="tiny" style="margin-top: 4px;">${item.detail || ''}</div>
            </div>
          `).join('')
        }</div>`;
        document.getElementById('provenance-panel').innerHTML = `
          <div class="summary-box">${publishProvenance.published_at
            ? `Pages staged ${formatTimestamp(publishProvenance.published_at)} from ${publishProvenance.source_html || 'index.html'}.`
            : 'Publish provenance will appear after `ffbayes publish-pages` stages this dashboard.'}</div>
          <div class="tiny">Dashboard generated: ${formatTimestamp(data.generated_at)}</div>
          <div class="tiny">Analysis freshness: ${(analysisFreshness.status || 'unknown')}${analysisFreshness.override_used ? ' (explicit override used)' : ''}</div>
          <div class="tiny">Staged surface sync: ${(publishProvenance.surface_sync && publishProvenance.surface_sync.status) || (publishProvenance.published_at ? 'status detail unavailable' : 'local-only')}</div>
          ${(publishProvenance.dashboard_generated_at || publishProvenance.source_payload)
            ? `<div class="tiny">Staged payload: ${publishProvenance.source_payload || 'dashboard_payload.json'} • generated ${formatTimestamp(publishProvenance.dashboard_generated_at)}</div>`
            : ''}
          ${renderFreshnessNotice(analysisFreshness)}
          ${(publishProvenance.surface_sync && publishProvenance.surface_sync.detail)
            ? `<div class="tiny">${publishProvenance.surface_sync.detail}</div>`
            : ''}
        `;
        document.getElementById('support-metrics').innerHTML = [
          ['Evidence status', decisionEvidence.status || 'unavailable'],
          ['Analysis freshness', analysisFreshness.status || 'unknown'],
          ['Published at', publishProvenance.published_at ? formatTimestamp(publishProvenance.published_at) : 'local-only'],
          ['Players in board', boardState.rows.length],
          ['Available players', boardState.availableRows.length],
          ['Current pick', state.currentPickNumber],
          ['League size', state.leagueSize],
          ['Bench slots', state.benchSlots],
        ].map(([label, value]) => `
          <div class="metric">
            <span class="label">${label}</span>
            <span class="value">${value}</span>
          </div>
        `).join('');
        document.getElementById('freshness-table').innerHTML = renderedFreshnessRows.map((row) => `
          <tr>
            <td>${row.source_name || row.label || 'unknown'}</td>
            <td>${row.status || (row.freshness_days != null ? `${row.freshness_days}d` : 'unknown')}</td>
            <td>${row.override_used ? 'yes' : 'no'}</td>
            <td>${row.latest_expected_year ?? '-'} / ${row.latest_found_year ?? '-'}</td>
          </tr>
        `).join('') || '<tr><td colspan="4" class="empty">No freshness rows available.</td></tr>';
      }

      function getDraftedRows(boardState) {
        const draftedSet = new Set((state.yourPlayers || []).map(safeLower));
        return boardState.rows
          .filter((row) => draftedSet.has(safeLower(row.player_name)))
          .map((row) => ({ ...row }));
      }

      function isRosterComplete(boardState) {
        const draftedRows = getDraftedRows(boardState);
        const openNeed = Object.values(boardState.rosterNeed || {}).reduce((sum, value) => sum + Number(value || 0), 0);
        return draftedRows.length >= rosterCapacity() && openNeed === 0;
      }

      function sortRowsForLineup(rows) {
        return [...rows].sort((a, b) => (
          (Number(b.proj_points_mean || 0) - Number(a.proj_points_mean || 0))
          || (Number(b.draft_score || 0) - Number(a.draft_score || 0))
        ));
      }

      function buildStarterBenchSplit(draftedRows) {
        const used = new Set();
        const starters = [];
        const addSlot = (position, slotCount, labelPrefix) => {
          const pool = sortRowsForLineup(
            draftedRows.filter((row) => row.position === position && !used.has(safeLower(row.player_name)))
          );
          for (let index = 0; index < Math.max(0, slotCount); index += 1) {
            const row = pool[index];
            if (!row) {
              break;
            }
            used.add(safeLower(row.player_name));
            starters.push({ ...row, lineup_slot: `${labelPrefix}${index + 1}` });
          }
        };

        addSlot('QB', Number(state.rosterSpots.QB || 0), 'QB');
        addSlot('RB', Number(state.rosterSpots.RB || 0), 'RB');
        addSlot('WR', Number(state.rosterSpots.WR || 0), 'WR');
        addSlot('TE', Number(state.rosterSpots.TE || 0), 'TE');
        addSlot('DST', Number(state.rosterSpots.DST || 0), 'DST');
        addSlot('K', Number(state.rosterSpots.K || 0), 'K');

        const flexPool = sortRowsForLineup(
          draftedRows.filter((row) => ['RB', 'WR', 'TE'].includes(row.position) && !used.has(safeLower(row.player_name)))
        );
        for (let index = 0; index < Math.max(0, Number(state.rosterSpots.FLEX || 0)); index += 1) {
          const row = flexPool[index];
          if (!row) {
            break;
          }
          used.add(safeLower(row.player_name));
          starters.push({ ...row, lineup_slot: `FLEX${index + 1}` });
        }

        const bench = sortRowsForLineup(
          draftedRows.filter((row) => !used.has(safeLower(row.player_name)))
        ).map((row, index) => ({ ...row, lineup_slot: `BENCH${index + 1}` }));

        return { starters, bench };
      }

      function sumMetric(rows, key) {
        return rows.reduce((sum, row) => sum + Number(row[key] || 0), 0);
      }

      function buildFinalizedDraftPayload(boardState) {
        const draftedRows = getDraftedRows(boardState);
        const split = buildStarterBenchSplit(draftedRows);
        const starterFragility = split.starters.length
          ? split.starters.reduce((sum, row) => sum + Number(row.fragility_score || 0), 0) / split.starters.length
          : 0;
        const fullFragility = draftedRows.length
          ? draftedRows.reduce((sum, row) => sum + Number(row.fragility_score || 0), 0) / draftedRows.length
          : 0;
        const starterUpside = split.starters.length
          ? split.starters.reduce((sum, row) => sum + Number(row.upside_score || 0), 0) / split.starters.length
          : 0;
        const fullUpside = draftedRows.length
          ? draftedRows.reduce((sum, row) => sum + Number(row.upside_score || 0), 0) / draftedRows.length
          : 0;
        const capacity = rosterCapacity();
        const rosterComplete = isRosterComplete(boardState);
        const projectedStarterMean = sumMetric(split.starters, 'proj_points_mean');
        const projectedStarterFloor = sumMetric(split.starters, 'proj_points_floor');
        const projectedStarterCeiling = sumMetric(split.starters, 'proj_points_ceiling');
        const benchProjection = sumMetric(split.bench, 'proj_points_mean');
        const draftedPlayerRows = draftedRows.map((row) => ({
          player_name: row.player_name,
          position: row.position,
          team: row.team || '',
          adp: row.adp,
          market_rank: row.market_rank,
          draft_rank: row.draft_rank,
          draft_score: row.draft_score,
          simple_vor_proxy: row.simple_vor_proxy,
          fragility_score: row.fragility_score,
          upside_score: row.upside_score,
          proj_points_mean: row.proj_points_mean,
          proj_points_floor: row.proj_points_floor,
          proj_points_ceiling: row.proj_points_ceiling,
          lineup_slot: [...split.starters, ...split.bench].find((item) => safeLower(item.player_name) === safeLower(row.player_name))?.lineup_slot || '',
          value_indicator: summarizeValueIndicator(row),
        }));
        const followedCount = (state.pickLog || []).filter((entry) => entry.followed_model).length;
        const latestPickedPlayer = state.pickLog.length
          ? draftedPlayerRows.find((row) => safeLower(row.player_name) === safeLower(state.pickLog[state.pickLog.length - 1].player_name))
          : null;
        return {
          schema_version: FINALIZED_SCHEMA_VERSION,
          season_year: Number(draftYearLabel()),
          exported_at: new Date().toISOString(),
          source_payload_generated_at: data.generated_at || '',
          title: 'FFBayes Finalized Draft Snapshot',
          league_settings: {
            league_size: state.leagueSize,
            draft_position: state.draftPosition,
            scoring_preset: state.scoringPreset,
            risk_tolerance: state.riskTolerance,
            bench_slots: state.benchSlots,
            roster_spots: { ...(state.rosterSpots || {}) },
          },
          final_state: {
            current_pick_number: state.currentPickNumber,
            next_pick_number: boardState.nextPick,
            taken_players: state.takenPlayers.slice(),
            your_players: state.yourPlayers.slice(),
            queue_players: state.queuePlayers.slice(),
            roster_capacity: capacity,
            drafted_player_count: draftedRows.length,
            roster_complete: rosterComplete,
          },
          drafted_players: draftedPlayerRows,
          starters: split.starters.map((row) => ({ player_name: row.player_name, position: row.position, lineup_slot: row.lineup_slot, proj_points_mean: row.proj_points_mean })),
          bench: split.bench.map((row) => ({ player_name: row.player_name, position: row.position, lineup_slot: row.lineup_slot, proj_points_mean: row.proj_points_mean })),
          summary_metrics: {
            starter_lineup_mean: projectedStarterMean,
            starter_lineup_floor: projectedStarterFloor,
            starter_lineup_ceiling: projectedStarterCeiling,
            bench_depth_mean: benchProjection,
            starter_fragility_avg: starterFragility,
            full_roster_fragility_avg: fullFragility,
            starter_upside_avg: starterUpside,
            full_roster_upside_avg: fullUpside,
            risk_style: summarizeRiskStyle(starterFragility, starterUpside),
            model_follow_count: followedCount,
            model_pivot_count: Math.max(0, (state.pickLog || []).length - followedCount),
          },
          value_recap: draftedPlayerRows.map((row) => ({
            player_name: row.player_name,
            position: row.position,
            adp: row.adp,
            market_rank: row.market_rank,
            draft_rank: row.draft_rank,
            draft_score: row.draft_score,
            value_indicator: row.value_indicator,
          })),
          pick_receipts: (state.pickLog || []).slice(),
          selected_player: latestPickedPlayer || draftedPlayerRows[0] || null,
        };
      }

      function tableRowsHtml(rows, columns) {
        return rows.length
          ? rows.map((row) => `
            <tr>${columns.map((column) => `<td>${escapeHtml(row[column.key] ?? '')}</td>`).join('')}</tr>
          `).join('')
          : `<tr><td colspan="${columns.length}" class="empty">No rows available.</td></tr>`;
      }

      function buildFinalizedSnapshotHtml(payload) {
        const draftedRows = payload.drafted_players || [];
        const starters = payload.starters || [];
        const bench = payload.bench || [];
        const selectedPlayer = payload.selected_player || draftedRows[0] || null;
        const metrics = payload.summary_metrics || {};
        return `<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>FFBayes Finalized Draft Snapshot</title>
  <style>
    :root { color-scheme: dark; --bg:#08111e; --panel:#0f172a; --border:rgba(148,163,184,0.18); --text:#f8fafc; --muted:#94a3b8; --accent:#38bdf8; --good:#34d399; --warn:#f59e0b; }
    * { box-sizing:border-box; }
    body { margin:0; background:linear-gradient(180deg,#030712 0%,#0b1324 100%); color:var(--text); font-family:Inter,ui-sans-serif,system-ui,sans-serif; padding:20px; }
    .shell { max-width:1480px; margin:0 auto; display:grid; gap:16px; }
    .panel { background:rgba(15,23,42,0.92); border:1px solid var(--border); border-radius:20px; padding:18px; }
    .layout { display:grid; grid-template-columns:1.1fr 0.9fr; gap:16px; }
    .pill { display:inline-flex; align-items:center; padding:6px 10px; border-radius:999px; border:1px solid rgba(56,189,248,0.24); background:rgba(56,189,248,0.14); color:#d7f5ff; font-size:12px; margin-right:8px; margin-bottom:8px; }
    .metric-grid { display:grid; gap:10px; grid-template-columns:repeat(2, minmax(0, 1fr)); }
    .metric { background:rgba(255,255,255,0.04); border:1px solid rgba(148,163,184,0.12); border-radius:14px; padding:10px 12px; }
    .label { display:block; color:var(--muted); font-size:11px; text-transform:uppercase; letter-spacing:0.06em; margin-bottom:4px; }
    .value { display:block; font-size:16px; font-weight:600; }
    table { width:100%; border-collapse:collapse; font-size:13px; }
    th, td { text-align:left; padding:10px; border-bottom:1px solid rgba(148,163,184,0.12); vertical-align:top; }
    th { color:#9bdcff; font-size:12px; text-transform:uppercase; letter-spacing:0.05em; }
    .summary-box { padding:12px; border-radius:14px; background:rgba(255,255,255,0.04); border:1px solid rgba(148,163,184,0.12); line-height:1.5; }
    .roster-grid { display:grid; gap:12px; grid-template-columns:repeat(2, minmax(0, 1fr)); }
    .empty { color:var(--muted); }
    @media (max-width: 980px) { .layout, .metric-grid, .roster-grid { grid-template-columns:1fr; } }
  </style>
</head>
<body>
  <div class="shell">
    <section class="panel">
      <div class="pill">Read-only snapshot</div>
      <div class="pill">Schema ${escapeHtml(payload.schema_version || '')}</div>
      <div class="pill">League ${escapeHtml(payload.league_settings?.league_size || '')}-team</div>
      <h1 style="margin:8px 0 6px 0;">FFBayes Finalized Draft Snapshot</h1>
      <p style="color:#94a3b8; margin:0;">Saved from the local live draft dashboard. This snapshot is locked and contains no draft controls.</p>
    </section>
    <section class="panel">
      <div class="metric-grid">
        <div class="metric"><span class="label">Starter lineup mean</span><span class="value">${escapeHtml(formatNumber(metrics.starter_lineup_mean || 0))}</span></div>
        <div class="metric"><span class="label">Starter lineup floor</span><span class="value">${escapeHtml(formatNumber(metrics.starter_lineup_floor || 0))}</span></div>
        <div class="metric"><span class="label">Starter lineup ceiling</span><span class="value">${escapeHtml(formatNumber(metrics.starter_lineup_ceiling || 0))}</span></div>
        <div class="metric"><span class="label">Bench depth mean</span><span class="value">${escapeHtml(formatNumber(metrics.bench_depth_mean || 0))}</span></div>
        <div class="metric"><span class="label">Risk style</span><span class="value">${escapeHtml(metrics.risk_style || 'Balanced')}</span></div>
        <div class="metric"><span class="label">Model pivots</span><span class="value">${escapeHtml(String(metrics.model_pivot_count || 0))}</span></div>
      </div>
    </section>
    <section class="layout">
      <section class="panel">
        <h2 style="margin-top:0;">Final Roster</h2>
        <div class="roster-grid">
          <div>
            <h3>Starters</h3>
            <table>
              <thead><tr><th>Slot</th><th>Player</th><th>Pos</th><th>Proj</th></tr></thead>
              <tbody>${tableRowsHtml(starters, [{ key: 'lineup_slot' }, { key: 'player_name' }, { key: 'position' }, { key: 'proj_points_mean' }])}</tbody>
            </table>
          </div>
          <div>
            <h3>Bench</h3>
            <table>
              <thead><tr><th>Slot</th><th>Player</th><th>Pos</th><th>Proj</th></tr></thead>
              <tbody>${tableRowsHtml(bench, [{ key: 'lineup_slot' }, { key: 'player_name' }, { key: 'position' }, { key: 'proj_points_mean' }])}</tbody>
            </table>
          </div>
        </div>
      </section>
      <section class="panel">
        <h2 style="margin-top:0;">Player Inspector</h2>
        ${selectedPlayer ? `
          <div class="summary-box">
            <strong>${escapeHtml(selectedPlayer.player_name || '')}</strong><br />
            ${escapeHtml(selectedPlayer.position || '')}${selectedPlayer.team ? ` • ${escapeHtml(selectedPlayer.team)}` : ''}<br />
            Draft rank ${escapeHtml(String(selectedPlayer.draft_rank || 'n/a'))} • ADP ${escapeHtml(formatNumber(selectedPlayer.adp || 0))}<br />
            Board value score ${escapeHtml(formatNumber(selectedPlayer.draft_score || 0))} • Simple VOR proxy ${escapeHtml(formatNumber(selectedPlayer.simple_vor_proxy || 0))}<br />
            Fragility ${escapeHtml(formatPercent(selectedPlayer.fragility_score || 0))} • Upside ${escapeHtml(formatPercent(selectedPlayer.upside_score || 0))}
          </div>
        ` : '<div class="empty">No drafted player available.</div>'}
        <div style="margin-top:12px;">
          <h3>Pick Receipts</h3>
          <table>
            <thead><tr><th>Pick</th><th>Player</th><th>Top recommendation</th><th>Decision</th></tr></thead>
            <tbody>${tableRowsHtml(payload.pick_receipts || [], [{ key: 'pick_number' }, { key: 'player_name' }, { key: 'top_recommendation' }, { key: 'decision_label' }])}</tbody>
          </table>
        </div>
      </section>
    </section>
  </div>
</body>
</html>`;
      }

      function buildFinalizedSummaryHtml(payload) {
        const metrics = payload.summary_metrics || {};
        return `<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>FFBayes Post-Draft Summary</title>
  <style>
    :root { color-scheme: dark; --bg:#08111e; --panel:#0f172a; --border:rgba(148,163,184,0.18); --text:#f8fafc; --muted:#94a3b8; }
    * { box-sizing:border-box; }
    body { margin:0; background:linear-gradient(180deg,#030712 0%,#0b1324 100%); color:var(--text); font-family:Inter,ui-sans-serif,system-ui,sans-serif; padding:20px; }
    .shell { max-width:1280px; margin:0 auto; display:grid; gap:16px; }
    .panel { background:rgba(15,23,42,0.92); border:1px solid var(--border); border-radius:20px; padding:18px; }
    .metric-grid { display:grid; gap:10px; grid-template-columns:repeat(2, minmax(0, 1fr)); }
    .metric { background:rgba(255,255,255,0.04); border:1px solid rgba(148,163,184,0.12); border-radius:14px; padding:10px 12px; }
    .label { display:block; color:var(--muted); font-size:11px; text-transform:uppercase; letter-spacing:0.06em; margin-bottom:4px; }
    .value { display:block; font-size:16px; font-weight:600; }
    table { width:100%; border-collapse:collapse; font-size:13px; }
    th, td { text-align:left; padding:10px; border-bottom:1px solid rgba(148,163,184,0.12); vertical-align:top; }
    th { color:#9bdcff; font-size:12px; text-transform:uppercase; letter-spacing:0.05em; }
    .summary-box { padding:12px; border-radius:14px; background:rgba(255,255,255,0.04); border:1px solid rgba(148,163,184,0.12); line-height:1.5; }
    @media (max-width: 980px) { .metric-grid { grid-template-columns:1fr; } }
  </style>
</head>
<body>
  <div class="shell">
    <section class="panel">
      <h1 style="margin:0 0 6px 0;">FFBayes Post-Draft Summary</h1>
      <p style="color:#94a3b8; margin:0;">A compact recap of the roster you drafted and how it looked through the draft-board model.</p>
    </section>
    <section class="panel">
      <h2 style="margin-top:0;">Final Roster Recap</h2>
      <div class="summary-box">
        Drafted ${escapeHtml(String(payload.final_state?.drafted_player_count || 0))} of ${escapeHtml(String(payload.final_state?.roster_capacity || 0))} total roster spots.
        Roster complete: ${escapeHtml(payload.final_state?.roster_complete ? 'Yes' : 'No')}.
      </div>
      <table style="margin-top:12px;">
        <thead><tr><th>Player</th><th>Pos</th><th>Slot</th><th>Proj</th></tr></thead>
        <tbody>${tableRowsHtml(payload.drafted_players || [], [{ key: 'player_name' }, { key: 'position' }, { key: 'lineup_slot' }, { key: 'proj_points_mean' }])}</tbody>
      </table>
    </section>
    <section class="panel">
      <h2 style="margin-top:0;">Team Projection Snapshot</h2>
      <div class="metric-grid">
        <div class="metric"><span class="label">Starter lineup mean</span><span class="value">${escapeHtml(formatNumber(metrics.starter_lineup_mean || 0))}</span></div>
        <div class="metric"><span class="label">Starter lineup floor</span><span class="value">${escapeHtml(formatNumber(metrics.starter_lineup_floor || 0))}</span></div>
        <div class="metric"><span class="label">Starter lineup ceiling</span><span class="value">${escapeHtml(formatNumber(metrics.starter_lineup_ceiling || 0))}</span></div>
        <div class="metric"><span class="label">Bench depth mean</span><span class="value">${escapeHtml(formatNumber(metrics.bench_depth_mean || 0))}</span></div>
      </div>
    </section>
    <section class="panel">
      <h2 style="margin-top:0;">Risk & Upside Profile</h2>
      <div class="summary-box">
        This roster profiles as <strong>${escapeHtml(metrics.risk_style || 'Balanced')}</strong>.
        Starter fragility averages ${escapeHtml(formatPercent(metrics.starter_fragility_avg || 0))} with starter upside averaging ${escapeHtml(formatPercent(metrics.starter_upside_avg || 0))}.
      </div>
      <div class="metric-grid" style="margin-top:12px;">
        <div class="metric"><span class="label">Starter fragility avg</span><span class="value">${escapeHtml(formatPercent(metrics.starter_fragility_avg || 0))}</span></div>
        <div class="metric"><span class="label">Full-roster fragility avg</span><span class="value">${escapeHtml(formatPercent(metrics.full_roster_fragility_avg || 0))}</span></div>
        <div class="metric"><span class="label">Starter upside avg</span><span class="value">${escapeHtml(formatPercent(metrics.starter_upside_avg || 0))}</span></div>
        <div class="metric"><span class="label">Full-roster upside avg</span><span class="value">${escapeHtml(formatPercent(metrics.full_roster_upside_avg || 0))}</span></div>
      </div>
    </section>
    <section class="panel">
      <h2 style="margin-top:0;">Draft Value Recap</h2>
      <table>
        <thead><tr><th>Player</th><th>Pos</th><th>ADP</th><th>Draft rank</th><th>Indicator</th></tr></thead>
        <tbody>${tableRowsHtml(payload.value_recap || [], [{ key: 'player_name' }, { key: 'position' }, { key: 'adp' }, { key: 'draft_rank' }, { key: 'value_indicator' }])}</tbody>
      </table>
    </section>
    <section class="panel">
      <h2 style="margin-top:0;">Pick-by-Pick Receipts</h2>
      <table>
        <thead><tr><th>Pick</th><th>Player</th><th>Top recommendation</th><th>Decision</th><th>Board value score</th></tr></thead>
        <tbody>${tableRowsHtml(payload.pick_receipts || [], [{ key: 'pick_number' }, { key: 'player_name' }, { key: 'top_recommendation' }, { key: 'decision_label' }, { key: 'draft_score' }])}</tbody>
      </table>
    </section>
  </div>
</body>
</html>`;
      }

      function finalizeDraft(boardState) {
        if (!isLocalFinalizeSupported()) {
          return;
        }
        if (!isRosterComplete(boardState)) {
          const confirmed = window.confirm('Your roster is not full yet. Download a finalized draft snapshot anyway?');
          if (!confirmed) {
            return;
          }
        }
        const payload = buildFinalizedDraftPayload(boardState);
        const year = draftYearLabel();
        const stamp = timestampLabel();
        window.__ffbayesLastFinalizedPayload = payload;
        window.__ffbayesFinalizedDownloads = [
          `ffbayes_finalized_draft_${year}_${stamp}.json`,
          `ffbayes_finalized_draft_${year}_${stamp}.html`,
          `ffbayes_finalized_summary_${year}_${stamp}.html`,
        ];
        downloadTextFile(`ffbayes_finalized_draft_${year}_${stamp}.json`, JSON.stringify(payload, null, 2), 'application/json');
        downloadTextFile(`ffbayes_finalized_draft_${year}_${stamp}.html`, buildFinalizedSnapshotHtml(payload), 'text/html');
        downloadTextFile(`ffbayes_finalized_summary_${year}_${stamp}.html`, buildFinalizedSummaryHtml(payload), 'text/html');
      }

      function bindControls() {
        document.getElementById('player-search').addEventListener('input', (event) => {
          state.search = event.target.value;
          render();
        });
        document.getElementById('undo-button').addEventListener('click', undoLast);
        document.getElementById('redo-button').addEventListener('click', redoLast);
        document.getElementById('finalize-button').addEventListener('click', () => {
          const boardState = buildBoardState();
          finalizeDraft(boardState);
        });
        [
          ['scoring-preset', (value) => { state.scoringPreset = value; }],
          ['risk-tolerance', (value) => { state.riskTolerance = value; }],
          ['league-size', (value) => { state.leagueSize = Math.max(2, Number(value || 10)); state.draftPosition = Math.min(state.draftPosition, state.leagueSize); }],
          ['draft-position', (value) => { state.draftPosition = Math.max(1, Number(value || 1)); }],
          ['current-pick-number', (value) => { state.currentPickNumber = Math.max(1, Number(value || 1)); }],
          ['bench-slots', (value) => { state.benchSlots = Math.max(0, Number(value || 0)); }],
          ['roster-qb', (value) => { state.rosterSpots.QB = Math.max(0, Number(value || 0)); }],
          ['roster-rb', (value) => { state.rosterSpots.RB = Math.max(0, Number(value || 0)); }],
          ['roster-wr', (value) => { state.rosterSpots.WR = Math.max(0, Number(value || 0)); }],
          ['roster-te', (value) => { state.rosterSpots.TE = Math.max(0, Number(value || 0)); }],
          ['roster-flex', (value) => { state.rosterSpots.FLEX = Math.max(0, Number(value || 0)); }],
        ].forEach(([id, setter]) => {
          document.getElementById(id).addEventListener('change', (event) => {
            setter(event.target.value);
            render();
          });
        });
      }

      function safeLowerArray(values) {
        return (values || []).map(safeLower);
      }

      function buildPickReceipt(row, boardState) {
        const topRecommendation = (boardState.pickNow && boardState.pickNow[0]) || null;
        const topWaitCandidate = (boardState.canWait && boardState.canWait[0]) || null;
        return {
          pick_number: state.currentPickNumber,
          player_name: row.player_name,
          position: row.position,
          team: row.team || '',
          adp: Number(row.adp || 0),
          market_rank: Number(row.market_rank || 0),
          draft_score: Number(row.draft_score || 0),
          simple_vor_proxy: Number(row.simple_vor_proxy || 0),
          fragility_score: Number(row.fragility_score || 0),
          upside_score: Number(row.upside_score || 0),
          top_recommendation: topRecommendation ? topRecommendation.player_name : '',
          recommended_draft_score: topRecommendation ? Number(topRecommendation.draft_score || 0) : null,
          top_wait_candidate: topWaitCandidate ? {
            player_name: topWaitCandidate.player_name,
            position: topWaitCandidate.position,
            wait_utility: Number(topWaitCandidate.wait_utility || 0),
            availability_to_next_pick: Number(topWaitCandidate.availability_to_next_pick || 0),
            expected_regret: Number(topWaitCandidate.expected_regret || 0),
            draft_score: Number(topWaitCandidate.draft_score || 0),
          } : null,
          followed_model: !!(topRecommendation && safeLower(topRecommendation.player_name) === safeLower(row.player_name)),
          decision_label: topRecommendation && safeLower(topRecommendation.player_name) === safeLower(row.player_name)
            ? 'Followed model'
            : 'Pivoted',
        };
      }

      function applyAction(action, playerName) {
        const normalized = safeLower(playerName);
        const boardState = buildBoardState();
        const targetRow = boardState.rows.find((row) => safeLower(row.player_name) === normalized) || null;
        pushHistory();
        let shouldAdvancePick = false;
        if (action === 'queue') {
          state.queuePlayers = state.queuePlayers.some((item) => safeLower(item) === normalized)
            ? state.queuePlayers.filter((item) => safeLower(item) !== normalized)
            : [...state.queuePlayers, playerName];
        } else if (action === 'taken') {
          const alreadyTaken = state.takenPlayers.some((item) => safeLower(item) === normalized);
          state.takenPlayers = alreadyTaken
            ? state.takenPlayers.filter((item) => safeLower(item) !== normalized)
            : [...state.takenPlayers, playerName];
          if (alreadyTaken) {
            state.yourPlayers = state.yourPlayers.filter((item) => safeLower(item) !== normalized);
          } else {
            state.queuePlayers = state.queuePlayers.filter((item) => safeLower(item) !== normalized);
            shouldAdvancePick = true;
          }
        } else if (action === 'mine') {
          const alreadyMine = state.yourPlayers.some((item) => safeLower(item) === normalized);
          state.yourPlayers = alreadyMine
            ? state.yourPlayers.filter((item) => safeLower(item) !== normalized)
            : [...state.yourPlayers, playerName];
          state.takenPlayers = alreadyMine
            ? state.takenPlayers.filter((item) => safeLower(item) !== normalized)
            : Array.from(new Set([...state.takenPlayers, playerName]));
          if (alreadyMine) {
            state.queuePlayers = state.queuePlayers.filter((item) => safeLower(item) !== normalized);
          } else {
            state.queuePlayers = state.queuePlayers.filter((item) => safeLower(item) !== normalized);
            if (targetRow) {
              state.pickLog = [...state.pickLog, buildPickReceipt(targetRow, boardState)];
            }
            shouldAdvancePick = true;
          }
        }
        state.pickLog = (state.pickLog || []).filter((entry) =>
          state.yourPlayers.some((item) => safeLower(item) === safeLower(entry.player_name))
        );
        if (shouldAdvancePick) {
          state.currentPickNumber = Math.max(1, Number(state.currentPickNumber || 1) + 1);
        }
        render();
      }

      function undoLast() {
        const snapshot = state.history.pop();
        if (!snapshot) {
          return;
        }
        pushSnapshot(state.redoHistory, captureDraftSnapshot());
        restoreDraftSnapshot(snapshot);
        render();
      }

      function redoLast() {
        const snapshot = state.redoHistory.pop();
        if (!snapshot) {
          return;
        }
        pushSnapshot(state.history, captureDraftSnapshot());
        restoreDraftSnapshot(snapshot);
        render();
      }
    })();
  </script>
</body>
</html>
"""
    html = html.replace('__PAYLOAD_JSON__', payload_json).replace(
        '__GENERATED_LABEL__', resolved_generated_label
    )

    output_path.write_text(html, encoding='utf-8')
    return output_path


def build_draft_decision_artifacts(
    player_frame: pd.DataFrame,
    league_settings: LeagueSettings | None = None,
    context: DraftContext | None = None,
    season_history: pd.DataFrame | None = None,
    analysis_provenance: dict[str, Any] | None = None,
) -> DraftDecisionArtifacts:
    """Build the full set of draft decision artifacts in memory."""
    settings = league_settings or LeagueSettings()
    context = context or DraftContext(current_pick_number=settings.draft_position)
    decision_table = build_decision_table(player_frame, settings, context)
    recommendations = build_recommendations(decision_table, settings, context)
    tier_cliffs = build_tier_cliffs(decision_table)
    roster_scenarios = build_roster_scenarios(decision_table, settings)
    source_freshness = _compute_freshness(normalize_player_frame(player_frame))
    scoring_presets = _build_scoring_preset_bundle(player_frame, settings, context)
    backtest = (
        run_draft_backtest(season_history, settings)
        if season_history is not None and not season_history.empty
        else {}
    )
    dashboard_payload = build_dashboard_payload(
        decision_table,
        recommendations,
        tier_cliffs,
        roster_scenarios,
        source_freshness,
        settings,
        backtest=backtest,
        context=context,
        scoring_presets=scoring_presets,
        analysis_provenance=analysis_provenance,
    )
    metadata = {
        'generated_at': datetime.now().isoformat(timespec='seconds'),
        'context': {
            'current_pick_number': context.current_pick_number,
            'draft_position': settings.draft_position,
            'league_size': settings.league_size,
        },
        'decision_table_columns': list(decision_table.columns),
        'analysis_provenance': analysis_provenance or {},
    }
    return DraftDecisionArtifacts(
        league_settings=settings,
        decision_table=decision_table,
        recommendations=recommendations,
        roster_scenarios=roster_scenarios,
        tier_cliffs=tier_cliffs,
        source_freshness=source_freshness,
        backtest=backtest,
        dashboard_payload=dashboard_payload,
        metadata=metadata,
    )


def _stage_runtime_dashboard_shortcuts(
    html_path: Path, payload_path: Path, year: int
) -> dict[str, Path]:
    """Copy dashboard artifacts into a shallow runtime location.

    Canonical, versioned artifacts stay under:
    `<runtime_root>/runs/<year>/pre_draft/artifacts/draft_strategy/`, where
    `<runtime_root>` defaults to `~/ProjectsRuntime/ffbayes`.

    For convenience, we also stage stable entrypoints:
    - runtime root: `<runtime_root>/dashboard/index.html`
    - repo root (if writable): `<repo>/dashboard/index.html`
    """

    from ffbayes.utils.path_constants import get_project_root, get_runtime_root

    shortcuts: dict[str, Path] = {}
    try:
        runtime_root = get_runtime_root()
        dashboard_dir = runtime_root / 'dashboard'
        dashboard_dir.mkdir(parents=True, exist_ok=True)

        for legacy_path in list(dashboard_dir.glob('draft_board_*.html')) + list(
            dashboard_dir.glob('dashboard_payload_*.json')
        ):
            legacy_path.unlink(missing_ok=True)

        index_path = dashboard_dir / 'index.html'
        if html_path.exists() and html_path.resolve() != index_path.resolve():
            shutil.copy2(html_path, index_path)

        payload_target = dashboard_dir / 'dashboard_payload.json'
        if payload_path.exists() and payload_path.resolve() != payload_target.resolve():
            shutil.copy2(payload_path, payload_target)

        shortcuts.update(
            {
            'runtime_dashboard_dir': dashboard_dir,
            'runtime_dashboard_index': index_path,
            'runtime_dashboard_payload': payload_target,
            }
        )
    except OSError:
        pass

    try:
        project_root = get_project_root()
        repo_dashboard_dir = project_root / 'dashboard'
        repo_dashboard_dir.mkdir(parents=True, exist_ok=True)

        for legacy_path in list(repo_dashboard_dir.glob('draft_board_*.html')) + list(
            repo_dashboard_dir.glob('dashboard_payload_*.json')
        ):
            legacy_path.unlink(missing_ok=True)

        repo_index = repo_dashboard_dir / 'index.html'
        if html_path.exists() and html_path.resolve() != repo_index.resolve():
            shutil.copy2(html_path, repo_index)

        repo_payload = repo_dashboard_dir / 'dashboard_payload.json'
        if payload_path.exists() and payload_path.resolve() != repo_payload.resolve():
            shutil.copy2(payload_path, repo_payload)

        shortcuts.update(
            {
                'repo_dashboard_dir': repo_dashboard_dir,
                'repo_dashboard_index': repo_index,
                'repo_dashboard_payload': repo_payload,
            }
        )
    except OSError:
        pass

    return shortcuts


def save_draft_decision_artifacts(
    artifacts: DraftDecisionArtifacts,
    output_dir: Path | str,
    year: int | None = None,
    filename_prefix: str = '',
    dashboard_dir: Path | str | None = None,
    diagnostics_dir: Path | str | None = None,
) -> dict[str, Path]:
    """Write workbook, dashboard payload, and HTML dashboard to disk."""
    year = year or datetime.now().year
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    dashboard_dir = Path(dashboard_dir) if dashboard_dir is not None else output_dir
    dashboard_dir.mkdir(parents=True, exist_ok=True)
    diagnostics_dir = (
        Path(diagnostics_dir) if diagnostics_dir is not None else output_dir
    )
    diagnostics_dir.mkdir(parents=True, exist_ok=True)

    workbook_path = output_dir / f'draft_board_{filename_prefix}{year}.xlsx'
    payload_path = dashboard_dir / f'dashboard_payload_{filename_prefix}{year}.json'
    html_path = dashboard_dir / f'draft_board_{filename_prefix}{year}.html'
    compat_path = dashboard_dir / f'draft_board_{filename_prefix}{year}.json'
    backtest_years = artifacts.backtest.get('holdout_years', [])
    backtest_suffix = (
        f'{min(backtest_years)}-{max(backtest_years)}' if backtest_years else str(year)
    )
    backtest_path = (
        output_dir / f'draft_decision_backtest_{filename_prefix}{backtest_suffix}.json'
    )
    model_output_dir = output_dir / 'model_outputs'
    model_output_dir.mkdir(parents=True, exist_ok=True)
    comparison_path = model_output_dir / f'current_year_model_comparison_{year}.json'

    export_workbook(
        artifacts.decision_table,
        artifacts.recommendations,
        artifacts.tier_cliffs,
        artifacts.roster_scenarios,
        artifacts.source_freshness,
        workbook_path,
        artifacts.league_settings,
        backtest=artifacts.backtest,
        dashboard_payload=artifacts.dashboard_payload,
    )
    payload_path.write_text(
        json.dumps(artifacts.dashboard_payload, default=str, indent=2), encoding='utf-8'
    )
    export_dashboard_html(
        artifacts.decision_table,
        artifacts.recommendations,
        html_path,
        artifacts.league_settings,
        backtest=artifacts.backtest,
        source_freshness=artifacts.source_freshness,
        dashboard_payload=artifacts.dashboard_payload,
    )
    compat_path.write_text(
        json.dumps(artifacts.dashboard_payload, default=str, indent=2),
        encoding='utf-8',
    )
    if artifacts.backtest:
        backtest_path.write_text(
            json.dumps(artifacts.backtest, default=str, indent=2), encoding='utf-8'
        )

    comparison_payload = {
        'model_type': 'current_year_model_comparison',
        'year': year,
        'top_players': artifacts.decision_table[
            [
                column
                for column in ['player_name', 'position', 'draft_score', 'draft_tier']
                if column in artifacts.decision_table.columns
            ]
        ]
        .head(25)
        .to_dict(orient='records'),
        'supporting_math': artifacts.dashboard_payload.get('supporting_math', {}),
        'league_settings': artifacts.league_settings.to_dict(),
    }
    comparison_path.write_text(
        json.dumps(comparison_payload, default=str, indent=2), encoding='utf-8'
    )

    shortcuts = _stage_runtime_dashboard_shortcuts(html_path, payload_path, year)
    return {
        'workbook_path': workbook_path,
        'payload_path': payload_path,
        'html_path': html_path,
        'compat_path': compat_path,
        'backtest_path': backtest_path,
        'comparison_path': comparison_path,
        **shortcuts,
    }
